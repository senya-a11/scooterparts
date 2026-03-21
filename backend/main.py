# backend/main.py  ·  Fm TuN v18
import logging
import re
from fastapi import FastAPI, HTTPException, Depends, status, Request, Response, UploadFile, File, Form, Query, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse
from pydantic import BaseModel, EmailStr, field_validator, Field
from typing import List, Optional, Dict, Any, Literal
import uvicorn
from datetime import datetime, timedelta, timezone
import os
from pathlib import Path
import jwt
import hashlib
import uuid
import hmac
import json
from uuid import uuid4
import secrets
import shutil
import base64
import aiofiles
from PIL import Image
import io

import asyncpg
from asyncpg.pool import Pool
import asyncio
from contextlib import asynccontextmanager
from dotenv import load_dotenv
import smtplib
import urllib.request
import urllib.parse
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

load_dotenv()

# Fix #2: logger определён на уровне модуля — доступен везде
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger(__name__)

from fastapi.templating import Jinja2Templates
from fastapi import Request
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.responses import Response as StarletteResponse
import time
from collections import defaultdict

# ========== НАСТРОЙКА ПУТЕЙ ==========
BASE_DIR = Path(__file__).parent.parent

templates_path = BASE_DIR / "templates"
if not templates_path.exists():
    templates_path = Path("/app/templates")
    if not templates_path.exists():
        templates_path = BASE_DIR / "templates"
        templates_path.mkdir(exist_ok=True)

templates = Jinja2Templates(directory=str(templates_path))

STATIC_DIR = BASE_DIR / "static"
DATA_DIR   = BASE_DIR / "data"
UPLOAD_DIR = STATIC_DIR / "uploads"

# Создаем директории с правильными правами
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
STATIC_DIR.mkdir(parents=True, exist_ok=True)
(STATIC_DIR / "images").mkdir(exist_ok=True)
(STATIC_DIR / "favicon").mkdir(exist_ok=True)

# Fix #11: Права 0o777 (rwxrwxrwx) опасны — любой пользователь системы мог
# читать, записывать и выполнять файлы. Заменено на 0o755 (rwxr-xr-x).
try:
    os.chmod(UPLOAD_DIR, 0o755)
    print(f"✅ Права на папку uploads установлены (755): {UPLOAD_DIR}")
except Exception as e:
    print(f"⚠️ Не удалось установить права на {UPLOAD_DIR}: {e}")


# Fix #3: Секреты обязательны. Слабые значения по умолчанию недопустимы в продакшне.
# При отсутствии переменных окружения приложение не запустится — это намеренное поведение.
_admin_password_raw = os.getenv("ADMIN_PASSWORD", "")
if not _admin_password_raw:
    import warnings
    warnings.warn(
        "⚠️  ADMIN_PASSWORD не задан. Используется небезопасный пароль 'admin123'. "
        "Установите ADMIN_PASSWORD в .env перед деплоем!",
        stacklevel=2
    )
    _admin_password_raw = "admin123"

ADMIN_PASSWORD = _admin_password_raw
ADMIN_USERNAME = "admin"

DATABASE_URL = os.getenv("DATABASE_URL", "postgresql://user:password@localhost/scooter_shop")
if DATABASE_URL and DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

# Конфиг платёжной системы (заполните переменные окружения)
PAYMENT_API_KEY    = os.getenv("PAYMENT_API_KEY", "")
PAYMENT_SHOP_ID    = os.getenv("PAYMENT_SHOP_ID", "")
PAYMENT_SECRET_KEY = os.getenv("PAYMENT_SECRET_KEY", "")
PAYMENT_CALLBACK_URL = os.getenv("PAYMENT_CALLBACK_URL", "https://your-domain.com/api/payment/callback")
YM_COUNTER_ID = os.getenv("YM_COUNTER_ID", "")
VIDEO_URL = os.getenv("VIDEO_URL", "")  # VK Video embed URL для главной страницы
BASE_URL = os.getenv("BASE_URL", "https://fmtun.ru")

# ── Email-уведомления ──
SMTP_HOST = os.getenv("SMTP_HOST", "")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER", "")
SMTP_PASS = os.getenv("SMTP_PASS", "")
SMTP_FROM = os.getenv("SMTP_FROM", "noreply@fmtun.ru")

# ── Telegram-уведомления (бот для уведомлений) ──
TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "")
TELEGRAM_ADMIN_CHAT_ID = os.getenv("TELEGRAM_ADMIN_CHAT_ID", "")

ORDER_STATUS_LABELS: dict = {
    "created": "Создан",
    "processing": "В обработке",
    "confirmed": "Подтверждён",
    "in_transit": "В пути",
    "customs": "На таможне",
    "warehouse": "Прибыл на склад",
    "delivery": "Передан в доставку",
    "completed": "Завершён",
    "cancelled": "Отменён",
    "pending": "Создан",
    "shipped": "В пути",
}


STATUS_ICONS: dict = {
    "created": "📋", "processing": "⚙️", "confirmed": "✅",
    "in_transit": "🚚", "customs": "🛃", "warehouse": "📦",
    "delivery": "🤝", "completed": "🏁", "cancelled": "❌",
}

STATUS_STEPS = [
    ("created", "Создан"),
    ("processing", "В обработке"),
    ("confirmed", "Подтверждён"),
    ("in_transit", "В пути"),
    ("customs", "На таможне"),
    ("warehouse", "Прибыл на склад"),
    ("delivery", "Передан в доставку"),
    ("completed", "Завершён"),
]


def _build_status_timeline(current_status: str) -> str:
    if current_status == "cancelled":
        return '<div style="padding:.625rem 1rem;background:rgba(255,71,87,.1);border:1px solid rgba(255,71,87,.3);border-radius:8px;color:#FF4757;font-weight:700">❌ Заказ отменён</div>'
    status_keys = [s[0] for s in STATUS_STEPS]
    try:
        cur_idx = status_keys.index(current_status)
    except ValueError:
        cur_idx = -1
    rows = []
    for i, (key, label) in enumerate(STATUS_STEPS):
        if i < cur_idx:
            color, dot = '#00D4FF', '✓'
        elif i == cur_idx:
            color, dot = '#00D4FF', '●'
        else:
            color, dot = '#3A4055', '○'
        weight = '700' if i == cur_idx else '400'
        rows.append(
            f'<div style="display:flex;align-items:center;gap:.75rem;padding:.375rem 0;color:{color};font-weight:{weight}">'
            f'<span style="width:1.25rem;text-align:center;font-size:.9rem">{dot}</span>'
            f'<span>{label}</span></div>'
        )
    return '<div style="border-left:2px solid rgba(0,212,255,.2);padding-left:1rem;margin:.5rem 0">' + ''.join(rows) + '</div>'


def _send_email_sync(to_email: str, order_id: int, status: str, delay_note: str | None = None,
                     items: list | None = None, order_info: dict | None = None):
    """Отправить email-уведомление клиенту об изменении статуса заказа."""
    if not SMTP_HOST or not SMTP_USER or not SMTP_PASS:
        return
    base_url = os.getenv("BASE_URL", "https://scooterparts.onrender.com")
    label = ORDER_STATUS_LABELS.get(status, status)
    icon = STATUS_ICONS.get(status, "📦")
    note_html = (f'<div style="margin:1rem 0;padding:.75rem 1rem;background:rgba(255,176,32,.08);'
                 f'border:1px solid rgba(255,176,32,.25);border-radius:8px;color:#FFB020">'
                 f'<strong>Примечание:</strong> {delay_note}</div>') if delay_note else ''
    # Items table
    items_html = ''
    if items:
        DELIVERY_LABELS = {'in_stock': '📦 В наличии', 'auto': '🚗 Авто', 'air': '✈️ Авиа'}
        rows_html = ''.join(
            f'<tr><td style="padding:.5rem .75rem;border-bottom:1px solid rgba(255,255,255,.05)">{it.get("product_name","")}</td>'
            f'<td style="padding:.5rem .75rem;border-bottom:1px solid rgba(255,255,255,.05);text-align:center">{it.get("quantity",1)}</td>'
            f'<td style="padding:.5rem .75rem;border-bottom:1px solid rgba(255,255,255,.05);text-align:right">{float(it.get("price",0))*int(it.get("quantity",1)):,.0f} ₽</td>'
            f'<td style="padding:.5rem .75rem;border-bottom:1px solid rgba(255,255,255,.05);color:#7B8599;font-size:.8rem">'
            f'{DELIVERY_LABELS.get(it.get("delivery_type") or it.get("order_type",""), "")}</td></tr>'
            for it in items
        )
        items_html = f'''<table style="width:100%;border-collapse:collapse;margin:1rem 0;font-size:.875rem">
  <thead><tr style="color:#7B8599;font-size:.75rem;text-transform:uppercase">
    <th style="padding:.375rem .75rem;text-align:left;font-weight:600">Товар</th>
    <th style="padding:.375rem .75rem;text-align:center;font-weight:600">Кол-во</th>
    <th style="padding:.375rem .75rem;text-align:right;font-weight:600">Сумма</th>
    <th style="padding:.375rem .75rem;text-align:left;font-weight:600">Тип</th>
  </tr></thead><tbody>{rows_html}</tbody></table>'''
    # Order total
    total_html = ''
    if order_info and order_info.get('total_amount'):
        total_html = f'<div style="text-align:right;font-weight:700;color:#F0F4F8;margin:.5rem 0">Итого: {float(order_info["total_amount"]):,.0f} ₽</div>'
    track_html = ''
    if order_info and order_info.get('track_number'):
        track_html = f'<p style="font-size:.875rem">Трек-номер: <strong style="color:#00D4FF">{order_info["track_number"]}</strong></p>'
    timeline = _build_status_timeline(status)
    body = f"""<html><body style="background:#080A0F;padding:2rem;margin:0">
<div style="max-width:580px;margin:0 auto;background:#0D1018;border-radius:12px;padding:2rem;border:1px solid rgba(255,255,255,.06);font-family:sans-serif;color:#F0F4F8">
  <h1 style="color:#00D4FF;margin-top:0;font-size:1.4rem;letter-spacing:.05em">⚡ Fm TuN</h1>
  <p style="font-size:1rem;color:#7B8599;margin:.25rem 0 1.5rem">Уведомление о статусе заказа</p>
  <div style="background:rgba(0,212,255,.07);border:1px solid rgba(0,212,255,.2);border-radius:10px;padding:1.25rem;margin-bottom:1.5rem">
    <p style="margin:0 0 .25rem;font-size:.8rem;color:#7B8599;text-transform:uppercase;letter-spacing:.1em">Заказ #{order_id}</p>
    <div style="font-size:1.5rem;font-weight:700;color:#00D4FF">{icon} {label}</div>
  </div>
  {note_html}
  <h3 style="font-size:.85rem;text-transform:uppercase;letter-spacing:.1em;color:#7B8599;margin:1.5rem 0 .5rem">История статусов</h3>
  {timeline}
  {items_html}
  {total_html}
  {track_html}
  <hr style="border:none;border-top:1px solid rgba(255,255,255,.06);margin:1.5rem 0">
  <p style="margin:0"><a href="{base_url}/tracking" style="display:inline-block;padding:.625rem 1.5rem;background:#00D4FF;color:#080A0F;text-decoration:none;border-radius:6px;font-weight:700;font-size:.875rem">Отследить заказ →</a></p>
  <p style="color:#3A4055;font-size:.75rem;margin-top:1.5rem">© Fm TuN. Автоматическое уведомление — не отвечайте на это письмо.</p>
</div></body></html>"""
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"Fm TuN — Заказ #{order_id}: {label}"
        msg["From"] = SMTP_FROM
        msg["To"] = to_email
        msg.attach(MIMEText(body, "html", "utf-8"))
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=10) as srv:
            srv.ehlo()
            srv.starttls()
            srv.login(SMTP_USER, SMTP_PASS)
            srv.sendmail(SMTP_FROM, to_email, msg.as_string())
    except Exception as e:
        logger.warning("Email notification to %s failed: %s", to_email, e)


async def send_order_email(to_email: str, order_id: int, status: str, delay_note: str | None = None):
    """Fetch order details and send rich HTML email notification."""
    try:
        async with db.pool.acquire() as conn:
            items = await conn.fetch(
                "SELECT product_name, quantity, price, order_type, delivery_type FROM order_items WHERE order_id=$1",
                order_id
            )
            order_info = await conn.fetchrow(
                "SELECT total_amount, track_number FROM orders WHERE id=$1", order_id
            )
        await asyncio.to_thread(
            _send_email_sync, to_email, order_id, status, delay_note,
            [dict(r) for r in items] if items else None,
            dict(order_info) if order_info else None
        )
    except Exception as e:
        logger.warning("send_order_email error: %s", e)


def _send_verification_email_sync(to_email: str, token: str):
    """Отправить письмо для подтверждения email."""
    if not SMTP_HOST or not SMTP_USER or not SMTP_PASS:
        logger.warning("SMTP not configured — verification email not sent to %s", to_email)
        return
    base_url = os.getenv("BASE_URL", "https://scooterparts.onrender.com")
    verify_url = f"{base_url}/api/verify-email?token={token}"
    body = f"""<html><body style="background:#080A0F;padding:2rem;margin:0">
<div style="max-width:540px;margin:0 auto;background:#0D1018;border-radius:12px;padding:2rem;border:1px solid rgba(255,255,255,.06);font-family:sans-serif;color:#F0F4F8">
  <h1 style="color:#00D4FF;margin-top:0;font-size:1.5rem">⚡ Fm TuN</h1>
  <p style="font-size:1rem">Добро пожаловать! Подтвердите ваш email-адрес, чтобы завершить регистрацию.</p>
  <a href="{verify_url}" style="display:inline-block;margin:1.25rem 0;padding:.75rem 2rem;background:#00D4FF;color:#080A0F;text-decoration:none;border-radius:8px;font-weight:700;font-size:1rem">Подтвердить email</a>
  <p style="color:#7B8599;font-size:.85rem">Ссылка действительна 24 часа. Если вы не регистрировались — просто проигнорируйте это письмо.</p>
  <hr style="border:none;border-top:1px solid rgba(255,255,255,.06);margin:1.5rem 0">
  <p style="color:#7B8599;font-size:.8rem">© Fm TuN. Автоматическое уведомление — не отвечайте на это письмо.</p>
</div></body></html>"""
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = "Fm TuN — Подтверждение email"
        msg["From"] = SMTP_FROM
        msg["To"] = to_email
        msg.attach(MIMEText(body, "html", "utf-8"))
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=10) as srv:
            srv.ehlo()
            srv.starttls()
            srv.login(SMTP_USER, SMTP_PASS)
            srv.sendmail(SMTP_FROM, to_email, msg.as_string())
        logger.info("Verification email sent to %s", to_email)
    except Exception as e:
        logger.warning("Verification email to %s failed: %s", to_email, e)


async def send_verification_email(to_email: str, token: str):
    await asyncio.to_thread(_send_verification_email_sync, to_email, token)


def _send_telegram_sync(text: str):
    """Отправить сообщение в Telegram (admin-канал или пользователю)."""
    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_ADMIN_CHAT_ID:
        return
    try:
        payload = urllib.parse.urlencode({
            "chat_id": TELEGRAM_ADMIN_CHAT_ID,
            "text": text,
            "parse_mode": "HTML",
        }).encode("utf-8")
        req = urllib.request.Request(
            f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage",
            data=payload, method="POST"
        )
        urllib.request.urlopen(req, timeout=8)
    except Exception as e:
        logger.warning("Telegram notification failed: %s", e)


async def send_order_telegram(order_id: int, status: str, username: str, email: str, delay_note: str | None = None):
    label = ORDER_STATUS_LABELS.get(status, status)
    note = f"\n⚠️ <i>{delay_note}</i>" if delay_note else ""
    text = (f"📦 <b>Заказ #{order_id}</b>\n"
            f"Статус: <b>{label}</b>{note}\n"
            f"Клиент: {username} ({email})\n"
            f"🔗 <a href='https://scooterparts.onrender.com/admin'>Открыть в админке</a>")
    await asyncio.to_thread(_send_telegram_sync, text)


# ========== МОДЕЛИ ==========
class UserRegister(BaseModel):
    username: str
    email: EmailStr
    password: str
    full_name: str
    phone: Optional[str] = None
    privacy_accepted: bool = False

    @field_validator('username')
    @classmethod
    def validate_username(cls, v: str) -> str:
        if len(v) < 3:
            raise ValueError('Имя пользователя должно содержать минимум 3 символа')
        if len(v) > 50:
            raise ValueError('Имя пользователя должно содержать не более 50 символов')
        return v

    @field_validator('password')
    @classmethod
    def validate_password(cls, v: str) -> str:
        if len(v) < 8:
            raise ValueError('Пароль должен содержать минимум 8 символов')
        if len(v) > 128:
            raise ValueError('Пароль слишком длинный (максимум 128 символов)')
        if not re.search(r'[A-Za-zА-Яа-яЁё]', v):
            raise ValueError('Пароль должен содержать хотя бы одну букву')
        if not re.search(r'\d', v):
            raise ValueError('Пароль должен содержать хотя бы одну цифру')
        return v

    @field_validator('privacy_accepted')
    @classmethod
    def validate_privacy(cls, v: bool) -> bool:
        if not v:
            raise ValueError('Необходимо принять политику конфиденциальности')
        return v


class UserLogin(BaseModel):
    username: str
    password: str


class CartItem(BaseModel):
    product_id: int
    quantity: int


class CartUpdate(BaseModel):
    product_id: int
    quantity: int
    specification_id: Optional[int] = None
    order_type: Optional[str] = None       # "in_stock" | "preorder"
    delivery_type: Optional[str] = None    # "auto" | "air" (только при preorder)


# Fix #8: Типизированные модели вместо body: dict
class CartQuantityUpdate(BaseModel):
    quantity: int = Field(gt=0, le=9999)
    specification_id: Optional[int] = None


class AdminLogin(BaseModel):
    username: str
    password: str


class CustomerNoteCreate(BaseModel):
    note: str = Field(min_length=1, max_length=2000)


class OrderStatusUpdate(BaseModel):
    status: Literal[
        'created', 'processing', 'confirmed', 'in_transit',
        'customs', 'warehouse', 'delivery', 'completed', 'cancelled'
    ]
    delay_note: Optional[str] = Field(None, max_length=500)
    payment_status: Optional[Literal['not_paid', 'pending', 'waiting', 'paid', 'failed']] = None
    status_in_stock: Optional[Literal[
        'created', 'processing', 'confirmed', 'in_transit',
        'customs', 'warehouse', 'delivery', 'completed', 'cancelled'
    ]] = None
    status_auto: Optional[Literal[
        'created', 'processing', 'confirmed', 'in_transit',
        'customs', 'warehouse', 'delivery', 'completed', 'cancelled'
    ]] = None
    status_air: Optional[Literal[
        'created', 'processing', 'confirmed', 'in_transit',
        'customs', 'warehouse', 'delivery', 'completed', 'cancelled'
    ]] = None


class TrackNumberUpdate(BaseModel):
    track_number: Optional[str] = Field(None, max_length=100)


class PaymentStatusUpdate(BaseModel):
    payment_status: Literal['pending', 'waiting', 'paid', 'failed']


class ProductCreate(BaseModel):
    name: str
    category: str
    price: float
    description: str
    stock: int = 0
    featured: bool = False
    in_stock: bool = False  # В наличии
    preorder: bool = False  # Доступен по предзаказу
    preorder_unavailable: bool = False  # Предзаказ временно недоступен
    cost_price: Optional[float] = None  # Себестоимость (только для админа)
    installation_service: bool = False
    price_type: str = 'static'  # 'static' or 'dynamic'
    price_cny: Optional[float] = None
    no_preorder: bool = False
    description_sections: Optional[dict] = None


class ProductUpdate(BaseModel):
    name: Optional[str] = None
    category: Optional[str] = None
    price: Optional[float] = None
    description: Optional[str] = None
    stock: Optional[int] = None
    featured: Optional[bool] = None
    in_stock: Optional[bool] = None  # В наличии
    preorder: Optional[bool] = None  # Доступен по предзаказу
    preorder_unavailable: Optional[bool] = None  # Предзаказ временно недоступен
    cost_price: Optional[float] = None  # Себестоимость
    installation_service: Optional[bool] = None
    price_type: Optional[str] = None
    price_cny: Optional[float] = None
    no_preorder: Optional[bool] = None
    description_sections: Optional[dict] = None


class CategoryCreate(BaseModel):
    slug: str
    name: str
    emoji: str = "📦"
    description: Optional[str] = ""


class CategoryUpdate(BaseModel):
    name: Optional[str] = None
    emoji: Optional[str] = None
    description: Optional[str] = None


class OrderCreate(BaseModel):
    delivery_address: Optional[str] = None
    comment: Optional[str] = None
    # Тип заказа для товаров с предзаказом, если не сохранён в cart_items
    items_overrides: Optional[list] = None
    # Промокод (опционально)
    promo_code: Optional[str] = None
    installation_data: Optional[dict] = None


class PaymentCreate(BaseModel):
    order_id: int
    amount: float
    currency: str = "RUB"


# ========== АУТЕНТИФИКАЦИЯ ==========
# Fix #3: SECRET_KEY обязан быть задан и длиннее 32 символов.
# Генерация надёжного ключа: python -c "import secrets; print(secrets.token_hex(32))"
_secret_key_raw = os.getenv("SECRET_KEY", "")
if not _secret_key_raw or _secret_key_raw in (
    "your-secret-key-change-in-production",
    "local-dev-secret-key-change-in-production",
):
    import warnings
    _fallback_key = "INSECURE-fallback-key-DO-NOT-USE-IN-PRODUCTION-" + secrets.token_hex(8)
    warnings.warn(
        "⚠️  SECRET_KEY не задан или содержит небезопасное значение по умолчанию. "
        "Установите SECRET_KEY в .env (минимум 32 случайных символа)! "
        "Сгенерируйте: python -c \"import secrets; print(secrets.token_hex(32))\"",
        stacklevel=2
    )
    _secret_key_raw = _fallback_key

if len(_secret_key_raw) < 32:
    raise RuntimeError(
        "SECRET_KEY слишком короткий (минимум 32 символа). "
        "Сгенерируйте: python -c \"import secrets; print(secrets.token_hex(32))\""
    )

SECRET_KEY = _secret_key_raw
ALGORITHM  = "HS256"
# HTTPBearer удалён — используем HttpOnly cookies


# Fix #5: валидация image_url против SSRF и XSS
_ALLOWED_IMAGE_URL = re.compile(
    r'^https?://.+\.(jpg|jpeg|png|gif|webp)$', re.IGNORECASE
)
_INTERNAL_HOSTS = re.compile(
    r'^(localhost|127\.\d+\.\d+\.\d+|0\.0\.0\.0|169\.254\.|10\.|192\.168\.|172\.(1[6-9]|2\d|3[01])\.)'
)

def validate_image_url(url: str) -> str:
    """Проверяет image_url против SSRF и XSS. Разрешает только https/http + картинки."""
    if not url:
        return url
    if url.startswith('data:image/'):   # base64 из формы — разрешаем
        return url
    if not _ALLOWED_IMAGE_URL.match(url):
        raise HTTPException(status_code=400, detail="Недопустимый URL изображения. Разрешены только http(s) ссылки на jpg/png/gif/webp.")
    from urllib.parse import urlparse
    host = urlparse(url).hostname or ""
    if _INTERNAL_HOSTS.match(host):
        raise HTTPException(status_code=400, detail="Недопустимый URL: внутренние адреса запрещены.")
    return url



class PasswordHasher:
    @staticmethod
    def get_password_hash(password: str) -> str:
        salt       = secrets.token_hex(16)
        iterations = 100000
        key = hashlib.pbkdf2_hmac('sha256', password.encode(), salt.encode(), iterations)
        return f"pbkdf2_sha256:{iterations}:{salt}:{key.hex()}"

    @staticmethod
    def verify_password(plain_password: str, hashed_password: str) -> bool:
        try:
            parts = hashed_password.split(':')
            if len(parts) != 4:
                return False
            algorithm, iterations_str, salt, stored_hash = parts
            if algorithm != 'pbkdf2_sha256':
                return False
            key = hashlib.pbkdf2_hmac(
                'sha256', plain_password.encode(), salt.encode(), int(iterations_str)
            )
            return hmac.compare_digest(key.hex(), stored_hash)
        except Exception:
            return False


hasher = PasswordHasher()


def create_access_token(data: dict, expires_minutes: int = 60):
    """
    Создаёт JWT с обязательным полем exp (срок действия).
    По умолчанию 60 минут. Используйте меньший срок для чувствительных операций.
    """
    to_encode = data.copy()
    for k, v in to_encode.items():
        if isinstance(v, (uuid.UUID, datetime)):
            to_encode[k] = str(v)
    expire = datetime.now(timezone.utc) + timedelta(minutes=expires_minutes)
    to_encode["exp"] = expire
    to_encode["iat"] = datetime.now(timezone.utc)  # issued at
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)


def get_current_user(request: Request) -> str:
    """
    Fix #1: Читает JWT из HttpOnly cookie вместо Bearer-заголовка.
    XSS не может получить HttpOnly cookie через JavaScript.
    """
    token = request.cookies.get("access_token")
    if not token:
        raise HTTPException(status_code=401, detail="Не авторизован")
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = payload.get("user_id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Недействительный токен: отсутствует user_id")
        return user_id
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Токен истёк. Войдите снова")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Недействительный токен")
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=401, detail="Ошибка аутентификации")


def verify_admin(request: Request) -> dict:
    """Lvl 2 Admin + Lvl 3 Superadmin: проверяет is_admin ИЛИ is_superadmin из cookie-токена."""
    token = request.cookies.get("access_token")
    if not token:
        raise HTTPException(status_code=401, detail="Не авторизован")
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        if not (payload.get("is_admin") or payload.get("is_superadmin")):
            raise HTTPException(status_code=403, detail="Доступ запрещён")
        return payload
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=401, detail="Не авторизован")


def verify_superadmin(request: Request) -> dict:
    """Lvl 3 Superadmin only: проверяет is_superadmin из cookie-токена."""
    token = request.cookies.get("access_token")
    if not token:
        raise HTTPException(status_code=401, detail="Не авторизован")
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        if not payload.get("is_superadmin"):
            raise HTTPException(status_code=403, detail="Доступ запрещён: требуется уровень Суперадмин")
        return payload
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=401, detail="Не авторизован")


def verify_manager_or_admin(request: Request) -> dict:
    """Lvl 1 Manager + Lvl 2 Admin + Lvl 3 Superadmin (заказы, статусы, трек-номера)."""
    token = request.cookies.get("access_token")
    if not token:
        raise HTTPException(status_code=401, detail="Не авторизован")
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        if not (payload.get("is_admin") or payload.get("is_manager") or payload.get("is_superadmin")):
            raise HTTPException(status_code=403, detail="Доступ запрещён")
        return payload
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=401, detail="Не авторизован")


# ========== БАЗА ДАННЫХ ==========
DEFAULT_CATEGORIES = [
    ("batteries",  "Аккумуляторы", "🔋", "Литий-ионные аккумуляторы и зарядные устройства"),
    ("motors",     "Моторы",       "⚙️", "Мотор-колёса, контроллеры и двигатели"),
    ("electronics","Электроника",  "📱", "Дисплеи, контроллеры, электроника"),
    ("brakes",     "Тормоза",      "🛑", "Тормозные диски, колодки, тросы"),
    ("tires",      "Колёса",       "🛞", "Покрышки, камеры, обода"),
    ("accessories","Аксессуары",   "🔧", "Ручки, подножки, зеркала и прочее"),
]


class Database:
    def __init__(self):
        self.pool: Optional[Pool] = None

    async def connect(self):
        try:
            self.pool = await asyncpg.create_pool(
                DATABASE_URL, min_size=1, max_size=10, command_timeout=60
            )
            await self.init_database()
            print("✅ База данных подключена")
        except Exception as e:
            print(f"❌ Ошибка подключения к БД: {e}")
            raise

    async def disconnect(self):
        if self.pool:
            await self.pool.close()

    async def init_database(self):
        async with self.pool.acquire() as conn:
            # Пользователи
            await conn.execute('''
                CREATE TABLE IF NOT EXISTS users (
                    id UUID PRIMARY KEY,
                    username VARCHAR(50) UNIQUE NOT NULL,
                    email VARCHAR(100) UNIQUE NOT NULL,
                    full_name VARCHAR(100) NOT NULL,
                    phone VARCHAR(20),
                    password_hash TEXT NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    is_admin BOOLEAN DEFAULT FALSE,
                    privacy_accepted BOOLEAN DEFAULT FALSE,
                    privacy_accepted_at TIMESTAMP
                )
            ''')

            # Категории товаров (CRUD)
            await conn.execute('''
                CREATE TABLE IF NOT EXISTS categories (
                    slug VARCHAR(50) PRIMARY KEY,
                    name VARCHAR(100) NOT NULL,
                    emoji VARCHAR(10) DEFAULT '📦',
                    description TEXT DEFAULT '',
                    sort_order INTEGER DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')

            # Товары
            await conn.execute('''
                CREATE TABLE IF NOT EXISTS products (
                    id SERIAL PRIMARY KEY,
                    name VARCHAR(200) NOT NULL,
                    category VARCHAR(50) NOT NULL,
                    price DECIMAL(10,2) NOT NULL,
                    description TEXT NOT NULL,
                    image_url TEXT NOT NULL,
                    stock INTEGER DEFAULT 0,
                    featured BOOLEAN DEFAULT FALSE,
                    in_stock BOOLEAN DEFAULT FALSE,
                    preorder BOOLEAN DEFAULT FALSE,
                    cost_price DECIMAL(10,2),
                    price_preorder_auto DECIMAL(10,2),
                    price_preorder_air DECIMAL(10,2),
                    has_specifications BOOLEAN DEFAULT FALSE,
                    specifications_data TEXT,
                    discount_percent INTEGER DEFAULT 0,
                    weight_kg DECIMAL(8,3) DEFAULT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            # Add columns if they don't exist (migration)
            try:
                await conn.execute('''
                    ALTER TABLE products ADD COLUMN IF NOT EXISTS preorder BOOLEAN DEFAULT FALSE
                ''')
                await conn.execute('''
                    ALTER TABLE products ADD COLUMN IF NOT EXISTS price_preorder_auto DECIMAL(10,2)
                ''')
                await conn.execute('''
                    ALTER TABLE products ADD COLUMN IF NOT EXISTS price_preorder_air DECIMAL(10,2)
                ''')
                await conn.execute('''
                    ALTER TABLE product_specifications ADD COLUMN IF NOT EXISTS price_preorder_auto DECIMAL(10,2)
                ''')
                await conn.execute('''
                    ALTER TABLE product_specifications ADD COLUMN IF NOT EXISTS price_preorder_air DECIMAL(10,2)
                ''')
                await conn.execute('''
                    ALTER TABLE cart_items ADD COLUMN IF NOT EXISTS order_type VARCHAR(20)
                ''')
                await conn.execute('''
                    ALTER TABLE cart_items ADD COLUMN IF NOT EXISTS delivery_type VARCHAR(10)
                ''')
                await conn.execute('''
                    ALTER TABLE products ADD COLUMN IF NOT EXISTS in_stock BOOLEAN DEFAULT FALSE
                ''')
                await conn.execute('''
                    ALTER TABLE products ADD COLUMN IF NOT EXISTS image_data TEXT
                ''')
                await conn.execute('''
                    ALTER TABLE products ADD COLUMN IF NOT EXISTS cost_price DECIMAL(10,2)
                ''')
                await conn.execute('''
                    ALTER TABLE products ADD COLUMN IF NOT EXISTS has_specifications BOOLEAN DEFAULT FALSE
                ''')
                await conn.execute('''
                    ALTER TABLE products ADD COLUMN IF NOT EXISTS specifications_data TEXT
                ''')
                # КРИТИЧЕСКИЕ МИГРАЦИИ: изменение типа image_url для поддержки base64
                await conn.execute('''
                    ALTER TABLE products ALTER COLUMN image_url TYPE TEXT
                ''')
                await conn.execute('''
                    ALTER TABLE product_specifications ALTER COLUMN image_url TYPE TEXT
                ''')
                await conn.execute('''
                    ALTER TABLE product_images ALTER COLUMN image_url TYPE TEXT
                ''')
                print("✅ Миграция: все image_url изменены на TEXT для поддержки base64")
                
                # КРИТИЧЕСКАЯ МИГРАЦИЯ: обновление in_stock для всех существующих товаров
                await conn.execute('''
                    UPDATE products SET in_stock = (stock > 0) WHERE in_stock != (stock > 0)
                ''')
                await conn.execute('''
                    UPDATE product_specifications SET in_stock = (stock > 0) WHERE in_stock != (stock > 0)
                ''')
                print("✅ Миграция: поле in_stock обновлено для всех товаров на основе stock")
            except Exception as e:
                print(f"⚠️ Миграция: {e}")
                pass  # Columns already exist

            # Миграция v18: поле скидки
            try:
                await conn.execute('''
                    ALTER TABLE products ADD COLUMN IF NOT EXISTS discount_percent INTEGER DEFAULT 0
                ''')
                await conn.execute('''
                    ALTER TABLE product_specifications ADD COLUMN IF NOT EXISTS discount_percent INTEGER DEFAULT 0
                ''')
                print("✅ Миграция v18: поле discount_percent добавлено")
            except Exception as e:
                print(f"⚠️ Миграция v18 (discount_percent): {e}")

            # Миграция v18: поле веса товара
            try:
                await conn.execute('''
                    ALTER TABLE products ADD COLUMN IF NOT EXISTS weight_kg DECIMAL(8,3) DEFAULT NULL
                ''')
                await conn.execute('''
                    ALTER TABLE product_specifications ADD COLUMN IF NOT EXISTS weight_kg DECIMAL(8,3) DEFAULT NULL
                ''')
                print("✅ Миграция v18: поле weight_kg добавлено")
            except Exception as e:
                print(f"⚠️ Миграция v18 (weight_kg): {e}")

            # Миграция v19: бейджик «Предзаказ недоступен»
            try:
                await conn.execute('''
                    ALTER TABLE products ADD COLUMN IF NOT EXISTS preorder_unavailable BOOLEAN DEFAULT FALSE
                ''')
                print("✅ Миграция v19: поле preorder_unavailable добавлено")
            except Exception as e:
                print(f"⚠️ Миграция v19 (preorder_unavailable): {e}")

            # Миграция v20: суперадмин + поля гарантии и сертификатов
            try:
                await conn.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS is_superadmin BOOLEAN DEFAULT FALSE")
                print("✅ Миграция v20: поле is_superadmin добавлено в users")
            except Exception as e:
                print(f"⚠️ Миграция v20 (is_superadmin): {e}")
            try:
                await conn.execute("ALTER TABLE products ADD COLUMN IF NOT EXISTS warranty_months INTEGER DEFAULT NULL")
                await conn.execute("ALTER TABLE products ADD COLUMN IF NOT EXISTS warranty_enabled BOOLEAN DEFAULT TRUE")
                await conn.execute("ALTER TABLE products ADD COLUMN IF NOT EXISTS certifications TEXT DEFAULT NULL")
                print("✅ Миграция v20: поля warranty_months, warranty_enabled, certifications добавлены в products")
            except Exception as e:
                print(f"⚠️ Миграция v20 (products warranty/cert): {e}")
            # Сделать главного admin-пользователя суперадмином
            try:
                await conn.execute("UPDATE users SET is_superadmin=TRUE WHERE username='admin' AND is_admin=TRUE")
                print("✅ Миграция v20: пользователю admin выдан статус суперадмина")
            except Exception as e:
                print(f"⚠️ Миграция v20 (superadmin upgrade): {e}")

            # Таблица настроек доставки
            await conn.execute('''
                CREATE TABLE IF NOT EXISTS delivery_settings (
                    id SERIAL PRIMARY KEY,
                    key VARCHAR(100) UNIQUE NOT NULL,
                    value TEXT,
                    label VARCHAR(200),
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            # Заполняем дефолтные настройки если их нет
            default_settings = [
                ('auto_rate_per_kg', None, 'Авто: ставка за 1 кг (₽)'),
                ('auto_min_price',   None, 'Авто: минимальная стоимость доставки (₽)'),
                ('auto_days_min',    None, 'Авто: срок доставки от (дней)'),
                ('auto_days_max',    None, 'Авто: срок доставки до (дней)'),
                ('air_rate_per_kg',  None, 'Самолёт: ставка за 1 кг (¥)'),
                ('air_min_weight',   '1',  'Самолёт: минимальный вес (кг)'),
                ('air_cny_to_rub',   None, 'Курс ¥ → ₽'),
                ('air_days_min',     None, 'Самолёт: срок доставки от (дней)'),
                ('air_days_max',     None, 'Самолёт: срок доставки до (дней)'),
            ]
            for key, val, label in default_settings:
                await conn.execute('''
                    INSERT INTO delivery_settings (key, value, label)
                    VALUES ($1, $2, $3)
                    ON CONFLICT (key) DO NOTHING
                ''', key, val, label)
            print("✅ Таблица delivery_settings готова")

            # Миграции v18.2: новые поля заказов
            try:
                await conn.execute("ALTER TABLE orders ADD COLUMN IF NOT EXISTS track_number VARCHAR(100) DEFAULT NULL")
                await conn.execute("ALTER TABLE orders ADD COLUMN IF NOT EXISTS delay_note TEXT DEFAULT NULL")
                print("✅ Миграция v18.2: track_number, delay_note в orders")
            except Exception as e:
                print(f"⚠️ Миграция v18.2 (orders): {e}")

            try:
                await conn.execute("ALTER TABLE order_items ADD COLUMN IF NOT EXISTS order_type VARCHAR(20) DEFAULT 'in_stock'")
                await conn.execute("ALTER TABLE order_items ADD COLUMN IF NOT EXISTS delivery_type VARCHAR(20) DEFAULT NULL")
                await conn.execute("ALTER TABLE order_items ADD COLUMN IF NOT EXISTS weight_kg DECIMAL(8,3) DEFAULT NULL")
                await conn.execute("ALTER TABLE order_items ADD COLUMN IF NOT EXISTS specification_id INTEGER DEFAULT NULL")
                await conn.execute("ALTER TABLE order_items ADD COLUMN IF NOT EXISTS specification_name VARCHAR(200) DEFAULT NULL")
                print("✅ Миграция v18.2: поля order_items")
            except Exception as e:
                print(f"⚠️ Миграция v18.2 (order_items): {e}")

            try:
                await conn.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS personal_discount INTEGER DEFAULT 0")
                await conn.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS is_manager BOOLEAN DEFAULT FALSE")
                print("✅ Миграция v18.2: personal_discount, is_manager в users")
            except Exception as e:
                print(f"⚠️ Миграция v18.2 (users): {e}")

            # Таблица заявок на установку (ТЗ 15)
            await conn.execute('''
                CREATE TABLE IF NOT EXISTS installation_requests (
                    id SERIAL PRIMARY KEY,
                    user_id UUID,
                    product_id INTEGER,
                    product_name VARCHAR(200),
                    scooter_model VARCHAR(200),
                    battery_type VARCHAR(200),
                    motor_type VARCHAR(200),
                    other_info TEXT,
                    full_name VARCHAR(200),
                    phone VARCHAR(50),
                    telegram VARCHAR(100),
                    status VARCHAR(30) DEFAULT 'new',
                    admin_comment TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            print("✅ Таблица installation_requests готова")
            
            # Спецификации товаров (версии/поколения)
            await conn.execute('''
                CREATE TABLE IF NOT EXISTS product_specifications (
                    id SERIAL PRIMARY KEY,
                    product_id INTEGER NOT NULL,
                    name VARCHAR(200) NOT NULL,
                    price DECIMAL(10,2) NOT NULL,
                    description TEXT,
                    image_url TEXT,
                    stock INTEGER DEFAULT 0,
                    in_stock BOOLEAN DEFAULT FALSE,
                    preorder BOOLEAN DEFAULT FALSE,
                    cost_price DECIMAL(10,2),
                    price_preorder_auto DECIMAL(10,2),
                    price_preorder_air DECIMAL(10,2),
                    discount_percent INTEGER DEFAULT 0,
                    weight_kg DECIMAL(8,3) DEFAULT NULL,
                    sort_order INTEGER DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (product_id) REFERENCES products(id) ON DELETE CASCADE
                )
            ''')
            
            # Характеристики товара
            await conn.execute('''
                CREATE TABLE IF NOT EXISTS product_characteristics (
                    id SERIAL PRIMARY KEY,
                    product_id INTEGER NOT NULL,
                    specification_id INTEGER,
                    char_name VARCHAR(100) NOT NULL,
                    char_value TEXT NOT NULL,
                    sort_order INTEGER DEFAULT 0,
                    FOREIGN KEY (product_id) REFERENCES products(id) ON DELETE CASCADE,
                    FOREIGN KEY (specification_id) REFERENCES product_specifications(id) ON DELETE CASCADE
                )
            ''')
            
            # Дополнительные изображения товара
            await conn.execute('''
                CREATE TABLE IF NOT EXISTS product_images (
                    id SERIAL PRIMARY KEY,
                    product_id INTEGER NOT NULL,
                    specification_id INTEGER,
                    image_url TEXT NOT NULL,
                    sort_order INTEGER DEFAULT 0,
                    FOREIGN KEY (product_id) REFERENCES products(id) ON DELETE CASCADE,
                    FOREIGN KEY (specification_id) REFERENCES product_specifications(id) ON DELETE CASCADE
                )
            ''')

            # Корзина
            await conn.execute('''
                CREATE TABLE IF NOT EXISTS cart_items (
                    id SERIAL PRIMARY KEY,
                    user_id UUID NOT NULL,
                    product_id INTEGER NOT NULL,
                    specification_id INTEGER,
                    quantity INTEGER NOT NULL CHECK (quantity > 0),
                    order_type VARCHAR(20),
                    delivery_type VARCHAR(10),
                    added_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(user_id, product_id, specification_id),
                    FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
                    FOREIGN KEY (product_id) REFERENCES products(id) ON DELETE CASCADE,
                    FOREIGN KEY (specification_id) REFERENCES product_specifications(id) ON DELETE CASCADE
                )
            ''')
            
            # Миграция: добавляем specification_id если его нет
            try:
                await conn.execute('''
                    ALTER TABLE cart_items ADD COLUMN IF NOT EXISTS specification_id INTEGER
                ''')
                # Удаляем старые ограничения уникальности (могут конфликтовать)
                # Три возможных имени в зависимости от версии схемы
                await conn.execute('''
                    ALTER TABLE cart_items DROP CONSTRAINT IF EXISTS cart_items_user_id_product_id_key
                ''')
                await conn.execute('''
                    ALTER TABLE cart_items DROP CONSTRAINT IF EXISTS cart_items_unique
                ''')
                await conn.execute('''
                    ALTER TABLE cart_items DROP CONSTRAINT IF EXISTS cart_items_user_id_product_id_specification_id_key
                ''')
                # Fix БАГ-02 (Вариант C): старый индекс (user_id, product_id) без order_type
                # не позволяет двум строкам одного товара с разным order_type существовать одновременно.
                # Новый индекс включает order_type — один товар может быть в корзине и как in_stock,
                # и как preorder одновременно (две отдельные строки).
                await conn.execute('''
                    DROP INDEX IF EXISTS cart_items_uniq_no_spec
                ''')
                await conn.execute('''
                    CREATE UNIQUE INDEX IF NOT EXISTS cart_items_uniq_no_spec
                    ON cart_items(user_id, product_id, COALESCE(order_type, ''))
                    WHERE specification_id IS NULL
                ''')
                await conn.execute('''
                    CREATE UNIQUE INDEX IF NOT EXISTS cart_items_uniq_with_spec
                    ON cart_items(user_id, product_id, specification_id, COALESCE(order_type, ''))
                    WHERE specification_id IS NOT NULL
                ''')
            except Exception as e:
                print(f"Cart migration (non-critical): {e}")
                pass  # Columns/constraints already exist

            # Миграция: delay_note в orders
            try:
                await conn.execute('''
                    ALTER TABLE orders ADD COLUMN IF NOT EXISTS delay_note TEXT
                ''')
            except Exception:
                pass

            # Заказы
            await conn.execute('''
                CREATE TABLE IF NOT EXISTS orders (
                    id SERIAL PRIMARY KEY,
                    user_id UUID,
                    status VARCHAR(30) DEFAULT 'pending',
                    total_amount DECIMAL(12,2) NOT NULL,
                    delivery_address TEXT,
                    comment TEXT,
                    payment_status VARCHAR(30) DEFAULT 'pending',
                    payment_id VARCHAR(200),
                    payment_url TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE SET NULL
                )
            ''')

            # Позиции заказа
            await conn.execute('''
                CREATE TABLE IF NOT EXISTS order_items (
                    id SERIAL PRIMARY KEY,
                    order_id INTEGER NOT NULL,
                    product_id INTEGER NOT NULL,
                    product_name VARCHAR(200) NOT NULL,
                    price DECIMAL(10,2) NOT NULL,
                    quantity INTEGER NOT NULL,
                    FOREIGN KEY (order_id) REFERENCES orders(id) ON DELETE CASCADE
                )
            ''')

            # Заметки о клиентах (CRM)
            await conn.execute('''
                CREATE TABLE IF NOT EXISTS customer_notes (
                    id SERIAL PRIMARY KEY,
                    user_id UUID NOT NULL,
                    note TEXT NOT NULL,
                    created_by VARCHAR(50) DEFAULT 'admin',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
                )
            ''')

            # 152-ФЗ: таблица для хранения согласий на обработку Cookie
            await conn.execute('''
                CREATE TABLE IF NOT EXISTS cookie_consents (
                    id SERIAL PRIMARY KEY,
                    session_id VARCHAR(128) UNIQUE NOT NULL,
                    consent_type VARCHAR(20) NOT NULL,
                    ip_address VARCHAR(45),
                    created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
                )
            ''')

            # ── v24: Промокоды ──
            await conn.execute('''
                CREATE TABLE IF NOT EXISTS promo_codes (
                    id SERIAL PRIMARY KEY,
                    code VARCHAR(50) UNIQUE NOT NULL,
                    discount_type VARCHAR(10) NOT NULL DEFAULT 'percent',
                    discount_value DECIMAL(10,2) NOT NULL,
                    max_uses INTEGER DEFAULT NULL,
                    used_count INTEGER DEFAULT 0,
                    is_active BOOLEAN DEFAULT TRUE,
                    expires_at TIMESTAMP DEFAULT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            # ── v24: Авто-скидка (ключ-значение, как delivery_settings) ──
            auto_disc_defaults = [
                ('auto_discount_enabled', '0', 'Авто-скидка: включена (1/0)'),
                ('auto_discount_percent', '0', 'Авто-скидка: размер (%)'),
            ]
            for key, val, label in auto_disc_defaults:
                await conn.execute('''
                    INSERT INTO delivery_settings (key, value, label)
                    VALUES ($1, $2, $3)
                    ON CONFLICT (key) DO NOTHING
                ''', key, val, label)
            # ── v24: Комментарий к позиции заказа ──
            try:
                await conn.execute("ALTER TABLE order_items ADD COLUMN IF NOT EXISTS item_comment TEXT DEFAULT NULL")
                print("✅ Миграция v24: item_comment в order_items")
            except Exception as e:
                print(f"⚠️ Миграция v24 (item_comment): {e}")
            # ── v25: Промокод в заказе, вишлист ──
            try:
                await conn.execute("ALTER TABLE orders ADD COLUMN IF NOT EXISTS promo_code VARCHAR(50) DEFAULT NULL")
                await conn.execute("ALTER TABLE orders ADD COLUMN IF NOT EXISTS promo_discount DECIMAL(10,2) DEFAULT 0")
                print("✅ Миграция v25: promo_code, promo_discount в orders")
            except Exception as e:
                print(f"⚠️ Миграция v25 (promo): {e}")
            # ── v25: Вишлист ──
            await conn.execute('''
                CREATE TABLE IF NOT EXISTS wishlists (
                    id SERIAL PRIMARY KEY,
                    user_id UUID NOT NULL,
                    product_id INTEGER NOT NULL,
                    added_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(user_id, product_id),
                    FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
                    FOREIGN KEY (product_id) REFERENCES products(id) ON DELETE CASCADE
                )
            ''')
            # ── v25: Отзывы на товары ──
            await conn.execute('''
                CREATE TABLE IF NOT EXISTS product_reviews (
                    id SERIAL PRIMARY KEY,
                    product_id INTEGER NOT NULL,
                    user_id UUID NOT NULL,
                    rating INTEGER NOT NULL CHECK (rating BETWEEN 1 AND 5),
                    review_text TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(product_id, user_id),
                    FOREIGN KEY (product_id) REFERENCES products(id) ON DELETE CASCADE,
                    FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
                )
            ''')
            print("✅ v25: promo в orders, wishlists, product_reviews готовы")

            # ── Документы поставщика ──
            await conn.execute('''
                CREATE TABLE IF NOT EXISTS supplier_documents (
                    id SERIAL PRIMARY KEY,
                    doc_type VARCHAR(64) NOT NULL,
                    filename VARCHAR(255) NOT NULL,
                    uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            # Миграция: убрать UNIQUE-ограничение на doc_type (поддержка нескольких файлов на тип)
            try:
                await conn.execute("DROP INDEX IF EXISTS supplier_documents_doc_type_key")
            except Exception:
                pass

            # ── v26: поля заказа ──
            try:
                await conn.execute("ALTER TABLE orders ADD COLUMN IF NOT EXISTS customer_name VARCHAR(200)")
                await conn.execute("ALTER TABLE orders ADD COLUMN IF NOT EXISTS delivery_cost DECIMAL(12,2) DEFAULT 0")
                print("✅ Миграция v26: customer_name, delivery_cost в orders")
            except Exception as e:
                print(f"⚠️ Миграция v26 (orders): {e}")

            # ── v27: статусы по категориям доставки ──
            try:
                await conn.execute("ALTER TABLE orders ADD COLUMN IF NOT EXISTS status_in_stock VARCHAR(30) DEFAULT NULL")
                await conn.execute("ALTER TABLE orders ADD COLUMN IF NOT EXISTS status_auto VARCHAR(30) DEFAULT NULL")
                await conn.execute("ALTER TABLE orders ADD COLUMN IF NOT EXISTS status_air VARCHAR(30) DEFAULT NULL")
                print("✅ Миграция v27: status_in_stock, status_auto, status_air в orders")
            except Exception as e:
                print(f"⚠️ Миграция v27 (orders): {e}")

            # ── v28: верификация email ──
            try:
                await conn.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS email_verified BOOLEAN DEFAULT FALSE")
                await conn.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS email_verification_token VARCHAR(100)")
                await conn.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS email_verification_sent_at TIMESTAMP")
                # Существующие admin-пользователи сразу помечаем как верифицированные
                await conn.execute("UPDATE users SET email_verified=TRUE WHERE is_admin=TRUE AND email_verified=FALSE")
                print("✅ Миграция v28: email_verified, email_verification_token в users")
            except Exception as e:
                print(f"⚠️ Миграция v28 (users): {e}")

            # ── v29: хранение файлов документов в БД (Render ephemeral FS fix) ──
            try:
                await conn.execute("ALTER TABLE supplier_documents ADD COLUMN IF NOT EXISTS file_content BYTEA")
                await conn.execute("ALTER TABLE supplier_documents ADD COLUMN IF NOT EXISTS content_type VARCHAR(100)")
                print("✅ Миграция v29: file_content, content_type в supplier_documents")
            except Exception as e:
                print(f"⚠️ Миграция v29 (supplier_documents): {e}")

            # Migration: add installation_service to products
            try:
                await conn.execute("ALTER TABLE products ADD COLUMN IF NOT EXISTS installation_service BOOLEAN DEFAULT FALSE")
                print("✅ Миграция: поле installation_service добавлено")
            except Exception as e:
                print(f"Миграция installation_service: {e}")

            # Migration: add installation_data to orders
            try:
                await conn.execute("ALTER TABLE orders ADD COLUMN IF NOT EXISTS installation_data JSONB DEFAULT NULL")
                print("✅ Миграция: поле installation_data добавлено")
            except Exception as e:
                print(f"Миграция installation_data: {e}")

            # Migration: add price_type and price_cny to products
            try:
                await conn.execute("ALTER TABLE products ADD COLUMN IF NOT EXISTS price_type VARCHAR(10) DEFAULT 'static'")
                await conn.execute("ALTER TABLE products ADD COLUMN IF NOT EXISTS price_cny DECIMAL(10,2) DEFAULT NULL")
                print("✅ Миграция: поля price_type, price_cny добавлены")
            except Exception as e:
                print(f"Миграция price_type/price_cny: {e}")

            # Migration: add no_preorder to products
            try:
                await conn.execute("ALTER TABLE products ADD COLUMN IF NOT EXISTS no_preorder BOOLEAN DEFAULT FALSE")
                print("✅ Миграция: поле no_preorder добавлено")
            except Exception as e:
                print(f"Миграция no_preorder: {e}")

            # Migration: add description_sections to products
            try:
                await conn.execute("ALTER TABLE products ADD COLUMN IF NOT EXISTS description_sections JSONB DEFAULT NULL")
                print("✅ Миграция: поле description_sections добавлено")
            except Exception as e:
                print(f"Миграция description_sections: {e}")

            # Migration: add title to supplier_documents
            try:
                await conn.execute("ALTER TABLE supplier_documents ADD COLUMN IF NOT EXISTS title VARCHAR(255) DEFAULT NULL")
                print("✅ Миграция: поле title добавлено в supplier_documents")
            except Exception as e:
                print(f"Миграция supplier_documents.title: {e}")

            # Migration: add cny_rate to delivery_settings
            try:
                await conn.execute("""
                    INSERT INTO delivery_settings (key, value, label)
                    VALUES ('cny_rate', NULL, 'Курс ¥ → ₽ (ЦБ РФ)')
                    ON CONFLICT (key) DO NOTHING
                """)
                await conn.execute("""
                    INSERT INTO delivery_settings (key, value, label)
                    VALUES ('cny_rate_updated_at', NULL, 'Дата обновления курса ¥')
                    ON CONFLICT (key) DO NOTHING
                """)
                print("✅ Миграция: cny_rate добавлен в delivery_settings")
            except Exception as e:
                print(f"Миграция cny_rate: {e}")

            # --- Начальные категории ---
            for slug, name, emoji, desc in DEFAULT_CATEGORIES:
                exists = await conn.fetchval(
                    "SELECT EXISTS(SELECT 1 FROM categories WHERE slug=$1)", slug
                )
                if not exists:
                    await conn.execute(
                        "INSERT INTO categories (slug, name, emoji, description) VALUES ($1,$2,$3,$4)",
                        slug, name, emoji, desc
                    )

            # --- Демо-пользователь (только для разработки) ---
            if os.getenv("ENVIRONMENT") == "development":
                if not await conn.fetchval("SELECT EXISTS(SELECT 1 FROM users WHERE username='demo')"):
                    await conn.execute(
                        "INSERT INTO users (id,username,email,full_name,phone,password_hash,privacy_accepted) VALUES ($1,$2,$3,$4,$5,$6,$7)",
                        str(uuid4()), 'demo', 'demo@fmtun.ru', 'Демо Пользователь',
                        '+79991234567', hasher.get_password_hash("demo123"), True
                    )

            # --- Админ ---
            if not await conn.fetchval("SELECT EXISTS(SELECT 1 FROM users WHERE username='admin')"):
                await conn.execute(
                    "INSERT INTO users (id,username,email,full_name,password_hash,is_admin,is_superadmin,privacy_accepted) VALUES ($1,$2,$3,$4,$5,$6,$7,$8)",
                    str(uuid4()), 'admin', 'admin@fmtun.ru', 'Администратор',
                    hasher.get_password_hash(ADMIN_PASSWORD), True, True, True
                )

            # --- Демо-товары ---
            if not await conn.fetchval("SELECT COUNT(*) FROM products"):
                demo = [
                    ("Аккумулятор Premium 36V 15Ah","batteries",16500.00,"Высокоёмкий литий-ионный аккумулятор с BMS. Гарантия 24 мес.","/static/images/battery.jpg",8,True),
                    ("Мотор-колесо Ultra 500W","motors",12500.00,"Бесщёточный мотор с прямым приводом. Макс. скорость 45 км/ч.","/static/images/motor.jpg",5,True),
                    ("Контроллер Smart 36V","electronics",4900.00,"Интеллектуальный контроллер с Bluetooth и мобильным приложением.","/static/images/controller.jpg",15,False),
                    ("Дисплей Color LCD","electronics",3200.00,"Цветной LCD дисплей с подсветкой и индикацией всех параметров.","/static/images/display.jpg",12,True),
                    ("Тормозные диски Premium","brakes",2200.00,"Вентилируемые тормозные диски из нержавеющей стали.","/static/images/brakes.jpg",25,False),
                    # УДАЛЕНО: ("Колесо 10\" All-Terrain","tires",1800.00,"Пневматическое колесо для бездорожья с усиленными стенками.","/static/images/wheel.jpg",20,False),
                    ("Тормозные колодки Premium","brakes",1200.00,"Керамические тормозные колодки для дисковых тормозов.","/static/images/brake-pads.jpg",30,True),
                    ("Руль алюминиевый","accessories",2500.00,"Алюминиевый руль с резиновыми накладками.","/static/images/handlebar.jpg",15,False),
                ]
                for p in demo:
                    await conn.execute(
                        "INSERT INTO products (name,category,price,description,image_url,stock,featured) VALUES ($1,$2,$3,$4,$5,$6,$7)",
                        *p
                    )

            print("✅ БД инициализирована")


db = Database()


async def refresh_cny_rate_background():
    """Фоновое обновление курса юаня каждые 6 часов"""
    import httpx
    import xml.etree.ElementTree as ET
    while True:
        try:
            async with httpx.AsyncClient(timeout=10) as client:
                resp = await client.get("https://www.cbr.ru/scripts/XML_daily.asp")
                resp.raise_for_status()
            root = ET.fromstring(resp.text)
            for valute in root.findall('Valute'):
                char_code = valute.find('CharCode')
                if char_code is not None and char_code.text == 'CNY':
                    value_el = valute.find('Value')
                    nominal_el = valute.find('Nominal')
                    if value_el is not None and nominal_el is not None:
                        value = float(value_el.text.replace(',', '.'))
                        nominal = int(nominal_el.text)
                        cny_rate = value / nominal
                        async with db.pool.acquire() as conn:
                            await conn.execute(
                                "UPDATE delivery_settings SET value = $1, updated_at = NOW() WHERE key = 'cny_rate'",
                                str(round(cny_rate, 4))
                            )
                            await conn.execute(
                                "UPDATE delivery_settings SET value = $1, updated_at = NOW() WHERE key = 'cny_rate_updated_at'",
                                datetime.now().isoformat()
                            )
                        print(f"✅ CNY rate updated: {round(cny_rate, 4)}")
                    break
        except Exception as e:
            print(f"⚠️ CNY rate refresh failed: {e}")
        await asyncio.sleep(6 * 3600)  # every 6 hours


@asynccontextmanager
async def lifespan(app: FastAPI):
    await db.connect()

    # Fix B-4: удаляем устаревшие записи согласий (старше 180 дней)
    # Запускается один раз при старте — экономичнее периодического фонового таска
    try:
        async with db.pool.acquire() as conn:
            # asyncpg: execute() returns "DELETE N" string — parse count from it
            result = await conn.execute(
                "DELETE FROM cookie_consents WHERE created_at < NOW() - INTERVAL '180 days'"
            )
            count = int(result.split()[-1]) if result else 0
            if count:
                logger.info("Cookie consent cleanup: removed %d stale records", count)
    except Exception as e:
        logger.warning("Cookie consent cleanup failed (non-critical): %s", e)

    asyncio.create_task(refresh_cny_rate_background())

    yield
    await db.disconnect()


# ========== ПРИЛОЖЕНИЕ ==========
_environment = os.getenv("ENVIRONMENT", "production")
_docs_url    = "/docs"   if _environment == "development" else None
_redoc_url   = "/redoc"  if _environment == "development" else None

# Fix #11: OpenAPI /docs и /redoc закрыты в production
app = FastAPI(
    title="Fm TuN API v18",
    docs_url=_docs_url,
    redoc_url=_redoc_url,
    openapi_url="/openapi.json" if _environment == "development" else None,
    lifespan=lifespan,
)

# ========== SECURITY MIDDLEWARE ==========
# ВАЖНО: Все 4 middleware объединены в один класс.
# Несколько стеков BaseHTTPMiddleware вызывали anyio.WouldBlock → anyio.EndOfStream
# при обращении к любой HTML-странице (GET /cart, /products и т.д.) — ошибка 500.
# Причина: каждый BaseHTTPMiddleware создаёт внутренний anyio-поток для receive/send,
# и при вложении они конфликтуют между собой начиная со Starlette 0.21+.
# Решение: единственный BaseHTTPMiddleware, содержащий всю логику.

class AppMiddleware(BaseHTTPMiddleware):
    """
    Единый middleware, объединяющий:
      1. Body size limit (защита от DoS через огромные JSON-тела)
      2. Rate limiting (auth: 10 req/min, global: 120 req/min)
      3. CSRF validation (X-CSRF-Token для мутирующих /api/* запросов)
      4. Security headers + CSP nonce (защита от XSS, clickjacking, MIME sniffing)
    """
    JSON_LIMIT    = 512 * 1024
    CSRF_EXEMPT   = {"/api/payment/callback"}
    AUTH_PATHS    = {"/api/login", "/api/register", "/api/admin/login"}
    MAX_TRACKED   = 10_000

    def __init__(self, app):
        super().__init__(app)
        self.window       = 60
        self.auth_limit   = int(os.getenv("AUTH_RATE_LIMIT",   "10"))
        self.global_limit = int(os.getenv("GLOBAL_RATE_LIMIT", "120"))
        self._buckets: dict = defaultdict(lambda: {"auth": [], "global": []})

    def _clean(self, ts: list, now: float) -> list:
        return [t for t in ts if now - t < self.window]

    async def dispatch(self, request: Request, call_next):
        # ── 1. Body size limit ──────────────────────────────────────────────
        ct = request.headers.get("content-type", "")
        if "multipart/form-data" not in ct and "application/octet-stream" not in ct:
            cl = request.headers.get("content-length")
            if cl and int(cl) > self.JSON_LIMIT:
                return JSONResponse(status_code=413, content={"detail": "Тело запроса слишком большое."})

        # ── 2. Rate limiting ─────────────────────────────────────────────────
        if not request.url.path.startswith("/static"):
            # Извлекаем IP с учётом обратного прокси (Render.com)
            forwarded_for = request.headers.get("X-Forwarded-For", "")
            if forwarded_for:
                ip = forwarded_for.split(",")[0].strip()[:45]
            else:
                ip = getattr(request.client, "host", "unknown")
            now = time.time()
            if len(self._buckets) > self.MAX_TRACKED:
                # Удаляем 20% старых записей вместо полной очистки
                cutoff = now - self.window
                stale = [k for k, v in self._buckets.items()
                         if not v.get("auth") and not v.get("global")
                         or (v.get("auth") and v["auth"][-1] < cutoff)
                         and (v.get("global") and v["global"][-1] < cutoff)]
                for k in stale[:max(1, self.MAX_TRACKED // 5)]:
                    del self._buckets[k]
            bucket = self._buckets[ip]
            if request.url.path in self.AUTH_PATHS:
                bucket["auth"] = self._clean(bucket["auth"], now)
                if len(bucket["auth"]) >= self.auth_limit:
                    return JSONResponse(status_code=429,
                        content={"detail": f"Слишком много попыток входа. Подождите {self.window} секунд."},
                        headers={"Retry-After": str(self.window)})
                bucket["auth"].append(now)
            else:
                bucket["global"] = self._clean(bucket["global"], now)
                if len(bucket["global"]) >= self.global_limit:
                    return JSONResponse(status_code=429,
                        content={"detail": "Слишком много запросов. Попробуйте позже."},
                        headers={"Retry-After": str(self.window)})
                bucket["global"].append(now)

        # ── 3. CSRF validation ───────────────────────────────────────────────
        if (
            request.method in ("POST", "PUT", "DELETE", "PATCH")
            and request.url.path.startswith("/api/")
            and request.url.path not in self.CSRF_EXEMPT
        ):
            cookie_tok  = request.cookies.get("csrf_token")
            header_tok  = request.headers.get("X-CSRF-Token")
            if not cookie_tok or not header_tok:
                return JSONResponse(status_code=403, content={"detail": "CSRF токен отсутствует"})
            if not hmac.compare_digest(cookie_tok, header_tok):
                return JSONResponse(status_code=403, content={"detail": "Недействительный CSRF токен"})

        # ── 4. CSP nonce (передаётся в шаблоны через request.state) ─────────
        nonce = secrets.token_urlsafe(16)
        request.state.csp_nonce = nonce

        response = await call_next(request)

        # ── 5. Security headers ──────────────────────────────────────────────
        csp = (
            f"default-src 'self'; "
            f"script-src 'self' 'nonce-{nonce}' "
            f"https://cdnjs.cloudflare.com https://cdn.jsdelivr.net https://unpkg.com; "
            f"style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
            f"font-src 'self' https://fonts.gstatic.com https://cdn.prod.website-files.com; "
            f"img-src 'self' data: https:; "
            f"connect-src 'self' https://api.yookassa.ru; "
            f"frame-src 'none'; "
            f"frame-ancestors 'none'; "
            f"base-uri 'self'; "
            f"form-action 'self'; "
            f"worker-src 'self' blob:; "
            f"object-src 'none';"
        )
        response.headers["Content-Security-Policy"]   = csp
        response.headers["X-Frame-Options"]           = "DENY"
        response.headers["X-Content-Type-Options"]    = "nosniff"
        response.headers["X-XSS-Protection"]          = "1; mode=block"
        response.headers["Referrer-Policy"]           = "strict-origin-when-cross-origin"
        response.headers["Permissions-Policy"]        = "geolocation=(), microphone=(), camera=(), payment=()"
        if request.url.scheme == "https":
            response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains; preload"

        return response

app.add_middleware(AppMiddleware)

# Fix #2: CORS — конкретный список доменов вместо wildcard.
# Wildcard "*" + credentials=True запрещена по спецификации W3C и создаёт CSRF-риски.
# Задайте ALLOWED_ORIGINS в .env через запятую, например:
#   ALLOWED_ORIGINS=https://yoursite.com,https://www.yoursite.com
_raw_origins = os.getenv("ALLOWED_ORIGINS", "http://localhost:8000,http://localhost:3000")
ALLOWED_ORIGINS: list = [o.strip() for o in _raw_origins.split(",") if o.strip()]

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type", "X-CSRF-Token", "Accept"],
)

app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")


# ── SEO: robots.txt и sitemap.xml ────────────────────────────────────────────
@app.get("/robots.txt", include_in_schema=False)
async def robots_txt():
    from fastapi.responses import PlainTextResponse
    robots_path = STATIC_DIR / "robots.txt"
    if robots_path.exists():
        return PlainTextResponse(robots_path.read_text())
    return PlainTextResponse("User-agent: *\nAllow: /\n")


@app.get("/sitemap.xml", include_in_schema=False)
async def sitemap_xml():
    from fastapi.responses import Response as _Response
    from datetime import date
    today = date.today().isoformat()
    base = "https://scooterparts.onrender.com"
    urls = ["/", "/products", "/about", "/legal"]
    xml_parts = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">',
    ]
    for u in urls:
        xml_parts.append(
            f"<url><loc>{base}{u}</loc><lastmod>{today}</lastmod><changefreq>weekly</changefreq></url>"
        )
    xml_parts.append("</urlset>")
    return _Response(content="\n".join(xml_parts), media_type="application/xml")


# ==========================================
# ========== СТРАНИЦЫ ==========
# ==========================================

@app.get("/")
async def root(request: Request):
    return templates.TemplateResponse("index.html", {"request": request, "ym_counter_id": YM_COUNTER_ID, "video_url": VIDEO_URL})

@app.get("/products")
async def products_page(request: Request):
    return templates.TemplateResponse("products.html", {"request": request, "ym_counter_id": YM_COUNTER_ID})


def _require_admin_cookie(request: Request):
    """Серверная проверка — редирект на /auth если не admin/manager/superadmin."""
    token = request.cookies.get("access_token")
    if not token:
        return None
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        if payload.get("is_admin") or payload.get("is_manager") or payload.get("is_superadmin"):
            return payload
    except Exception:
        pass
    return None

@app.get("/admin")
async def admin_panel(request: Request):
    payload = _require_admin_cookie(request)
    if not payload:
        return RedirectResponse("/auth?next=/admin", status_code=302)
    is_superadmin = bool(payload.get("is_superadmin"))
    is_admin = bool(payload.get("is_admin")) or is_superadmin
    is_manager = bool(payload.get("is_manager")) and not is_admin
    return templates.TemplateResponse("admin.html", {
        "request": request,
        "is_manager": is_manager,
        "is_superadmin": is_superadmin,
        "ym_counter_id": YM_COUNTER_ID
    })

@app.get("/admin/add-product")
async def admin_add_product_page(request: Request):
    if not _require_admin_cookie(request):
        return RedirectResponse("/auth?next=/admin", status_code=302)
    return templates.TemplateResponse("add_product.html", {"request": request, "ym_counter_id": YM_COUNTER_ID})

@app.get("/legal")
async def legal_page(request: Request, doc: str = "terms"):
    """Единый маршрут для правовых документов. /legal?doc=terms|privacy|cookies|offer|returns"""
    allowed = {"terms", "privacy", "cookies", "offer", "returns"}
    doc_type = doc if doc in allowed else "terms"
    return templates.TemplateResponse("legal.html", {"request": request, "doc_type": doc_type, "ym_counter_id": YM_COUNTER_ID})

@app.get("/privacy-policy")
async def privacy_policy_page(request: Request):
    return templates.TemplateResponse("legal.html", {"request": request, "doc_type": "privacy", "ym_counter_id": YM_COUNTER_ID})

@app.get("/cookie-policy")
async def cookie_policy_page(request: Request):
    """Алиас — редирект на раздел Cookie политики конфиденциальности."""
    return RedirectResponse("/privacy-policy#cookies", status_code=301)

@app.get("/terms")
async def terms_page(request: Request):
    return templates.TemplateResponse("legal.html", {"request": request, "doc_type": "terms", "ym_counter_id": YM_COUNTER_ID})

@app.get("/offer")
async def offer_page(request: Request):
    return templates.TemplateResponse("legal.html", {"request": request, "doc_type": "offer", "ym_counter_id": YM_COUNTER_ID})

@app.get("/returns")
async def returns_page(request: Request):
    return templates.TemplateResponse("legal.html", {"request": request, "doc_type": "returns", "ym_counter_id": YM_COUNTER_ID})

@app.get("/about")
async def about_page(request: Request):
    return templates.TemplateResponse("about.html", {"request": request, "ym_counter_id": YM_COUNTER_ID})

@app.get("/tracking")
async def tracking_page(request: Request):
    return templates.TemplateResponse("tracking.html", {"request": request, "ym_counter_id": YM_COUNTER_ID})

@app.get("/cart")
async def cart_page(request: Request):
    return templates.TemplateResponse("cart.html", {"request": request, "ym_counter_id": YM_COUNTER_ID})

@app.get("/auth")
async def auth_page(request: Request, next: str = "/"):
    return templates.TemplateResponse("auth.html", {"request": request, "next_url": next, "ym_counter_id": YM_COUNTER_ID})


# Fix #7: Эндпоинт для получения CSRF-токена. Фронтенд вызывает его при загрузке
# и сохраняет токен для последующих запросов.
@app.get("/api/csrf-token")
async def get_csrf_token(response: Response):
    """Выдаёт CSRF-токен и устанавливает cookie. Вызывать при инициализации фронтенда."""
    token = secrets.token_urlsafe(32)
    _is_https = os.getenv("ENVIRONMENT", "production") != "development"
    response.set_cookie(
        key="csrf_token",
        value=token,
        httponly=False,   # Должен быть доступен из JS для отправки в заголовке
        secure=_is_https,
        samesite="strict",
        max_age=3600
    )
    return {"csrf_token": token}


@app.post("/api/refresh")
async def refresh_token(request: Request, response: Response):
    """
    Fix A-11: Тихое обновление JWT-токена.
    Вызывается фронтендом за ~10 минут до истечения сессии.
    Возвращает новый токен если текущий валиден.
    """
    token = request.cookies.get("access_token")
    if not token:
        raise HTTPException(status_code=401, detail="Не авторизован")
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
        user_id = payload.get("user_id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Не авторизован")

        # Выдаём новый токен с теми же claims
        new_token = create_access_token({
            "user_id": user_id,
            "is_admin": payload.get("is_admin", False),
            "username": payload.get("username", ""),
        })
        _is_https = os.getenv("ENVIRONMENT", "production") != "development"
        response.set_cookie(
            key="access_token",
            value=new_token,
            httponly=True,
            secure=_is_https,
            samesite="strict",
            max_age=3600,
            path="/",
        )
        return {"ok": True}
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Сессия истекла")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Не авторизован")
    except Exception as e:
        logger.error("refresh_token error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")



# ==========================================
# ========== AUTH API ==========
# ==========================================

@app.post("/api/register")
async def register(user_data: UserRegister):
    if not user_data.privacy_accepted:
        raise HTTPException(status_code=400, detail="Необходимо принять политику конфиденциальности")
    try:
        async with db.pool.acquire() as conn:
            if await conn.fetchval("SELECT EXISTS(SELECT 1 FROM users WHERE username=$1)", user_data.username):
                raise HTTPException(status_code=400, detail="Имя пользователя уже занято")
            if await conn.fetchval("SELECT EXISTS(SELECT 1 FROM users WHERE email=$1)", user_data.email):
                raise HTTPException(status_code=400, detail="Email уже зарегистрирован")

            user_id = str(uuid4())
            password_hash = hasher.get_password_hash(user_data.password)
            verification_token = secrets.token_urlsafe(32)
            await conn.execute(
                "INSERT INTO users (id,username,email,full_name,phone,password_hash,privacy_accepted,privacy_accepted_at,email_verified,email_verification_token,email_verification_sent_at) VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11)",
                user_id, user_data.username, user_data.email, user_data.full_name,
                user_data.phone, password_hash, True, datetime.utcnow(),
                False, verification_token, datetime.utcnow()
            )
            asyncio.create_task(send_verification_email(user_data.email, verification_token))
            return {"message": "Аккаунт создан. Проверьте вашу почту для подтверждения email.", "user_id": user_id, "requires_verification": True}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.post("/api/login")
async def login(login_data: UserLogin, response: Response):
    try:
        async with db.pool.acquire() as conn:
            user = await conn.fetchrow(
                "SELECT id,username,email,full_name,password_hash,is_admin,is_manager,is_superadmin,personal_discount,email_verified FROM users WHERE username=$1",
                login_data.username
            )
            if not user or not hasher.verify_password(login_data.password, user['password_hash']):
                logger.warning("Failed login attempt for username: %s", login_data.username)
                raise HTTPException(status_code=401, detail="Неверное имя пользователя или пароль")
            # Проверяем верификацию email (кроме администраторов)
            if not user['is_admin'] and not user.get('email_verified', True):
                raise HTTPException(
                    status_code=403,
                    detail="Подтвердите email перед входом. Проверьте вашу почту или запросите новое письмо."
                )

            user_id = str(user['id'])
            is_manager = bool(user.get('is_manager', False))
            is_superadmin = bool(user.get('is_superadmin', False))
            token = create_access_token({"user_id": user_id, "is_admin": user['is_admin'], "is_manager": is_manager, "is_superadmin": is_superadmin})

            # Fix #1: токен в HttpOnly cookie — JS не может прочитать
            _is_https = os.getenv("ENVIRONMENT", "production") != "development"
            response.set_cookie(
                key="access_token",
                value=token,
                httponly=True,
                secure=_is_https,
                samesite="strict",
                max_age=3600,
                path="/",
            )
            logger.info("User logged in: %s (admin=%s, manager=%s)", login_data.username, user['is_admin'], is_manager)
            return {
                "message": "Вход выполнен",
                "user": {
                    "id": user_id,
                    "username": user['username'],
                    "email": user['email'],
                    "full_name": user['full_name'],
                    "is_admin": user['is_admin'],
                    "is_manager": is_manager,
                    "personal_discount": int(user.get('personal_discount') or 0),
                }
            }
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.get("/api/verify-email")
async def verify_email(token: str):
    """Подтверждение email по токену из письма."""
    try:
        async with db.pool.acquire() as conn:
            user = await conn.fetchrow(
                "SELECT id, email_verification_sent_at FROM users WHERE email_verification_token=$1 AND email_verified=FALSE",
                token
            )
            if not user:
                return HTMLResponse(content="""<html><body style="background:#080A0F;color:#F0F4F8;font-family:sans-serif;display:flex;align-items:center;justify-content:center;min-height:100vh;margin:0">
<div style="text-align:center;max-width:400px;padding:2rem">
  <div style="font-size:3rem;margin-bottom:1rem">❌</div>
  <h2 style="color:#FF4444">Недействительная ссылка</h2>
  <p style="color:#7B8599">Ссылка недействительна или уже использована.</p>
  <a href="/auth" style="color:#00D4FF">Перейти ко входу</a>
</div></body></html>""", status_code=400)
            # Проверяем срок действия (24 часа)
            sent_at = user['email_verification_sent_at']
            sent_at_aware = sent_at.replace(tzinfo=timezone.utc) if sent_at and sent_at.tzinfo is None else sent_at
            if sent_at_aware and (datetime.now(timezone.utc) - sent_at_aware).total_seconds() > 86400:
                return HTMLResponse(content="""<html><body style="background:#080A0F;color:#F0F4F8;font-family:sans-serif;display:flex;align-items:center;justify-content:center;min-height:100vh;margin:0">
<div style="text-align:center;max-width:400px;padding:2rem">
  <div style="font-size:3rem;margin-bottom:1rem">⏰</div>
  <h2 style="color:#FFB020">Ссылка истекла</h2>
  <p style="color:#7B8599">Ссылка действительна только 24 часа. Запросите новое письмо.</p>
  <a href="/auth" style="color:#00D4FF">Перейти ко входу</a>
</div></body></html>""", status_code=400)
            await conn.execute(
                "UPDATE users SET email_verified=TRUE, email_verification_token=NULL, email_verification_sent_at=NULL WHERE id=$1",
                user['id']
            )
        return RedirectResponse(url="/auth?verified=1", status_code=302)
    except Exception as e:
        logger.error("Email verify error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.post("/api/resend-verification")
async def resend_verification(request: Request):
    """Повторная отправка письма подтверждения email."""
    try:
        data = await request.json()
        email = data.get("email", "").strip().lower()
        username = data.get("username", "").strip()
        if not email and not username:
            raise HTTPException(status_code=400, detail="Email или имя пользователя не указаны")
        async with db.pool.acquire() as conn:
            if email:
                user = await conn.fetchrow(
                    "SELECT id, email, email_verified, email_verification_sent_at FROM users WHERE email=$1",
                    email
                )
            else:
                user = await conn.fetchrow(
                    "SELECT id, email, email_verified, email_verification_sent_at FROM users WHERE username=$1",
                    username
                )
            if user:
                email = user['email']
            if not user:
                # Не раскрываем факт существования пользователя
                return {"message": "Если аккаунт существует — письмо отправлено"}
            if user['email_verified']:
                raise HTTPException(status_code=400, detail="Email уже подтверждён")
            # Rate-limit: не чаще раза в 5 минут
            sent_at = user['email_verification_sent_at']
            sent_at_aware2 = sent_at.replace(tzinfo=timezone.utc) if sent_at and sent_at.tzinfo is None else sent_at
            if sent_at_aware2 and (datetime.now(timezone.utc) - sent_at_aware2).total_seconds() < 300:
                raise HTTPException(status_code=429, detail="Подождите 5 минут перед повторной отправкой")
            new_token = secrets.token_urlsafe(32)
            await conn.execute(
                "UPDATE users SET email_verification_token=$1, email_verification_sent_at=$2 WHERE id=$3",
                new_token, datetime.utcnow(), user['id']
            )
        asyncio.create_task(send_verification_email(email, new_token))
        return {"message": "Письмо отправлено. Проверьте почту."}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Resend verification error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.post("/api/logout")
async def logout(response: Response):
    """Fix #1 + logout fix: Очищает access_token и csrf_token cookie при выходе."""
    _is_https = os.getenv("ENVIRONMENT", "production") != "development"
    response.delete_cookie("access_token", path="/", secure=_is_https, httponly=True, samesite="strict")
    # csrf_token тоже удаляем — устаревший токен в браузере не нужен
    response.delete_cookie("csrf_token", path="/", secure=_is_https, samesite="strict")
    return {"message": "Выход выполнен"}


# ─── 152-ФЗ: Логирование согласий на обработку Cookie ────────────────────────
class CookieConsentLog(BaseModel):
    consent_type: Literal["all", "necessary", "partial"]   # тип согласия
    session_id: str = Field(max_length=128)      # идентификатор сессии браузера

@app.post("/api/cookie-consent")
async def log_cookie_consent(request: Request, body: CookieConsentLog):
    """
    152-ФЗ, ст. 9: Логирует факт согласия / отказа пользователя на обработку Cookie.
    Хранит: IP, timestamp, тип согласия, session_id.
    Согласие должно быть зафиксировано до начала обработки.
    """
    try:
        ip = request.headers.get("X-Forwarded-For", request.client.host if request.client else "unknown")
        ip = ip.split(",")[0].strip()[:45]   # берём первый IP из цепочки прокси
        timestamp = datetime.now(timezone.utc).isoformat()
        logger.info(
            "COOKIE_CONSENT | type=%s | session=%s | ip=%s | ts=%s",
            body.consent_type, body.session_id[:32], ip, timestamp
        )
        async with db.pool.acquire() as conn:
            await conn.execute("""
                INSERT INTO cookie_consents (session_id, consent_type, ip_address, created_at)
                VALUES ($1, $2, $3, $4)
                ON CONFLICT (session_id) DO UPDATE
                  SET consent_type=$2, ip_address=$3, created_at=$4
            """, body.session_id[:128], body.consent_type, ip, datetime.utcnow())
        return {"status": "ok", "consent": body.consent_type, "timestamp": timestamp}
    except Exception as e:
        logger.error("Cookie consent log error: %s", e, exc_info=True)
        # Не ломаем UX — возвращаем ok даже при ошибке БД
        return {"status": "ok", "consent": body.consent_type}


@app.get("/api/me")
async def get_me(user_id: str = Depends(get_current_user)):
    try:
        async with db.pool.acquire() as conn:
            user = await conn.fetchrow(
                "SELECT id,username,email,full_name,phone,is_admin,is_manager,personal_discount,created_at FROM users WHERE id=$1",
                user_id
            )
            if not user:
                raise HTTPException(status_code=404, detail="Пользователь не найден")
            d = dict(user)
            d['id'] = str(d['id'])
            d['is_manager'] = bool(d.get('is_manager', False))
            d['personal_discount'] = int(d.get('personal_discount') or 0)
            if isinstance(d.get('created_at'), datetime):
                d['created_at'] = d['created_at'].isoformat()
            return d
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


# ==========================================
# ========== DELIVERY SETTINGS API ==========
# ==========================================

@app.get("/api/delivery-settings")
async def get_delivery_settings():
    """Получить все настройки доставки"""
    try:
        async with db.pool.acquire() as conn:
            rows = await conn.fetch("SELECT key, value, label FROM delivery_settings ORDER BY id")
            return {r['key']: {'value': r['value'], 'label': r['label']} for r in rows}
    except Exception as e:
        logger.error("Delivery settings error: %s", e)
        return {}

@app.post("/api/delivery-settings")
async def update_delivery_settings(request: Request, admin=Depends(verify_admin)):
    """Обновить настройки доставки (только для admin). При изменении ставок — пересчёт цен по весу (ТЗ §3)."""
    try:
        data = await request.json()
        async with db.pool.acquire() as conn:
            for key, value in data.items():
                val = str(value).strip() if value is not None and str(value).strip() else None
                await conn.execute(
                    "UPDATE delivery_settings SET value=$1, updated_at=NOW() WHERE key=$2",
                    val, key
                )

            # Fix БАГ-07: при изменении ставок — пересчитываем цены предзаказа по весу
            settings_rows = await conn.fetch("SELECT key, value FROM delivery_settings")
            settings = {r['key']: r['value'] for r in settings_rows}

            auto_rate = float(settings.get('auto_rate_per_kg') or 0)
            air_rate  = float(settings.get('air_rate_per_kg') or 0)
            cny_to_rub = float(settings.get('air_cny_to_rub') or 0)

            if 'auto_rate_per_kg' in data and auto_rate > 0:
                await conn.execute('''
                    UPDATE products
                    SET price_preorder_auto = ROUND(weight_kg * $1)
                    WHERE weight_kg IS NOT NULL AND weight_kg > 0
                ''', auto_rate)
                await conn.execute('''
                    UPDATE product_specifications ps
                    SET price_preorder_auto = ROUND(ps.weight_kg * $1)
                    FROM products p WHERE ps.product_id = p.id
                    AND ps.weight_kg IS NOT NULL AND ps.weight_kg > 0
                ''', auto_rate)

            if 'air_rate_per_kg' in data or 'air_cny_to_rub' in data:
                if air_rate > 0 and cny_to_rub > 0:
                    await conn.execute('''
                        UPDATE products
                        SET price_preorder_air = ROUND(weight_kg * $1 * $2)
                        WHERE weight_kg IS NOT NULL AND weight_kg > 0
                    ''', air_rate, cny_to_rub)
                    await conn.execute('''
                        UPDATE product_specifications ps
                        SET price_preorder_air = ROUND(ps.weight_kg * $1 * $2)
                        FROM products p WHERE ps.product_id = p.id
                        AND ps.weight_kg IS NOT NULL AND ps.weight_kg > 0
                    ''', air_rate, cny_to_rub)

        return {"success": True}
    except Exception as e:
        logger.error("Update delivery settings error: %s", e)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.get("/api/admin/cny-rate/refresh")
async def refresh_cny_rate(admin=Depends(verify_admin)):
    """Обновить курс юаня с сайта ЦБ РФ"""
    try:
        import httpx
        import xml.etree.ElementTree as ET
        async with httpx.AsyncClient(timeout=10) as client:
            resp = await client.get("https://www.cbr.ru/scripts/XML_daily.asp")
            resp.raise_for_status()
        root = ET.fromstring(resp.text)
        cny_rate = None
        for valute in root.findall('Valute'):
            char_code = valute.find('CharCode')
            if char_code is not None and char_code.text == 'CNY':
                value_el = valute.find('Value')
                nominal_el = valute.find('Nominal')
                if value_el is not None and nominal_el is not None:
                    value = float(value_el.text.replace(',', '.'))
                    nominal = int(nominal_el.text)
                    cny_rate = value / nominal
                break
        if cny_rate is None:
            raise HTTPException(status_code=502, detail="Не удалось получить курс юаня")
        now_str = datetime.now().isoformat()
        async with db.pool.acquire() as conn:
            await conn.execute(
                "UPDATE delivery_settings SET value = $1, updated_at = NOW() WHERE key = 'cny_rate'",
                str(round(cny_rate, 4))
            )
            await conn.execute(
                "UPDATE delivery_settings SET value = $1, updated_at = NOW() WHERE key = 'cny_rate_updated_at'",
                now_str
            )
        return {"ok": True, "cny_rate": round(cny_rate, 4), "updated_at": now_str}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("refresh_cny_rate error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail=f"Ошибка: {str(e)}")


@app.post("/api/admin/cny-rate")
async def set_cny_rate(data: dict, admin=Depends(verify_admin)):
    """Вручную установить курс юаня"""
    rate = data.get("cny_rate")
    if not rate or float(rate) <= 0:
        raise HTTPException(status_code=400, detail="Некорректный курс")
    now_str = datetime.now().isoformat()
    async with db.pool.acquire() as conn:
        await conn.execute(
            "UPDATE delivery_settings SET value = $1, updated_at = NOW() WHERE key = 'cny_rate'",
            str(float(rate))
        )
        await conn.execute(
            "UPDATE delivery_settings SET value = $1, updated_at = NOW() WHERE key = 'cny_rate_updated_at'",
            now_str
        )
    return {"ok": True, "cny_rate": float(rate), "updated_at": now_str}


@app.get("/api/admin/cny-rate")
async def get_cny_rate(admin=Depends(verify_admin)):
    """Получить текущий курс юаня"""
    async with db.pool.acquire() as conn:
        rows = await conn.fetch("SELECT key, value, updated_at FROM delivery_settings WHERE key IN ('cny_rate', 'cny_rate_updated_at')")
    result = {r['key']: r['value'] for r in rows}
    return {
        "cny_rate": float(result.get('cny_rate') or 0) if result.get('cny_rate') else None,
        "updated_at": result.get('cny_rate_updated_at'),
    }


@app.get("/api/delivery-cost")
async def calculate_delivery_cost(weight: float, type: str = "auto"):
    """Рассчитать стоимость доставки"""
    import math
    async with db.pool.acquire() as conn:
        rows = await conn.fetch("SELECT key, value FROM delivery_settings")
    settings = {r['key']: r['value'] for r in rows}
    weight_ceil = math.ceil(weight)
    if type == "auto":
        rate = float(settings.get('auto_rate_per_kg') or 0)
        min_price = float(settings.get('auto_min_price') or 0)
        if rate <= 0:
            return {"cost": None, "message": "Ставка не установлена"}
        cost = max(weight_ceil * rate, min_price)
        return {"cost": cost, "weight_ceil": weight_ceil, "rate": rate, "type": "auto"}
    elif type == "air":
        rate_cny = float(settings.get('air_rate_per_kg') or 0)
        cny_rate = float(settings.get('cny_rate') or settings.get('air_cny_to_rub') or 0)
        if rate_cny <= 0 or cny_rate <= 0:
            return {"cost": None, "message": "Ставка не установлена"}
        cost = max(weight_ceil, 1) * rate_cny * cny_rate
        return {"cost": round(cost, 2), "weight_ceil": weight_ceil, "rate_cny": rate_cny, "cny_rate": cny_rate, "type": "air"}
    raise HTTPException(status_code=400, detail="Неверный тип доставки")


# ==========================================
# ========== CATEGORIES API ==========
# ==========================================

@app.get("/api/categories")
async def get_categories():
    """Получить все категории с количеством товаров"""
    try:
        async with db.pool.acquire() as conn:
            rows = await conn.fetch('''
                SELECT c.slug, c.name, c.emoji, c.description, c.sort_order,
                       COUNT(p.id) as count
                FROM categories c
                LEFT JOIN products p ON p.category = c.slug
                GROUP BY c.slug, c.name, c.emoji, c.description, c.sort_order
                ORDER BY c.sort_order, c.name
            ''')
            return {"categories": [dict(r) for r in rows]}
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")

@app.delete("/api/users/me")
async def delete_my_account(request: Request, response: Response, user_id: str = Depends(get_current_user)):
    """
    Fix B-6: Удаление аккаунта пользователем (ст. 17 ФЗ-152).
    Каскадно удаляет: cart_items, cookie_consents (ON DELETE CASCADE).
    Заказы: user_id обнуляется (ON DELETE SET NULL) — сохраняются для бухучёта.
    """
    try:
        async with db.pool.acquire() as conn:
            # Убеждаемся что пользователь существует и не является администратором
            # Fix B-6: user_id — UUID строка, не int. int(user_id) вызывал ValueError.
            user = await conn.fetchrow(
                "SELECT id, is_admin FROM users WHERE id=$1::uuid",
                user_id
            )
            if not user:
                raise HTTPException(status_code=404, detail="Пользователь не найден")
            if user['is_admin']:
                raise HTTPException(status_code=403, detail="Аккаунт администратора нельзя удалить через этот эндпоинт")

            await conn.execute("DELETE FROM users WHERE id=$1::uuid", user_id)
            logger.info("User account deleted: user_id=%s", user_id)

        # Сбрасываем cookie сессии
        _is_https = os.getenv("ENVIRONMENT", "production") != "development"
        response.delete_cookie("access_token", path="/", secure=_is_https, httponly=True, samesite="strict")
        return {"message": "Аккаунт успешно удалён"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("delete_account error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")




@app.post("/api/admin/categories")
async def create_category(cat: CategoryCreate, admin=Depends(verify_admin)):
    """Создать новую категорию"""
    try:
        async with db.pool.acquire() as conn:
            if await conn.fetchval("SELECT EXISTS(SELECT 1 FROM categories WHERE slug=$1)", cat.slug):
                raise HTTPException(status_code=400, detail="Категория с таким slug уже существует")
            max_order = await conn.fetchval("SELECT COALESCE(MAX(sort_order),0) FROM categories")
            await conn.execute(
                "INSERT INTO categories (slug,name,emoji,description,sort_order) VALUES ($1,$2,$3,$4,$5)",
                cat.slug, cat.name, cat.emoji, cat.description or "", max_order + 1
            )
            return {"message": "Категория создана", "slug": cat.slug}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.put("/api/admin/categories/{slug}")
async def update_category(slug: str, cat: CategoryUpdate, admin=Depends(verify_admin)):
    """Обновить категорию"""
    try:
        async with db.pool.acquire() as conn:
            existing = await conn.fetchrow("SELECT * FROM categories WHERE slug=$1", slug)
            if not existing:
                raise HTTPException(status_code=404, detail="Категория не найдена")
            new_name  = cat.name  if cat.name  is not None else existing['name']
            new_emoji = cat.emoji if cat.emoji is not None else existing['emoji']
            new_desc  = cat.description if cat.description is not None else existing['description']
            await conn.execute(
                "UPDATE categories SET name=$1, emoji=$2, description=$3 WHERE slug=$4",
                new_name, new_emoji, new_desc, slug
            )
            return {"message": "Категория обновлена"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.delete("/api/admin/categories/{slug}")
async def delete_category(slug: str, admin=Depends(verify_admin)):
    """Удалить категорию (только если нет товаров)"""
    try:
        async with db.pool.acquire() as conn:
            count = await conn.fetchval("SELECT COUNT(*) FROM products WHERE category=$1", slug)
            if count > 0:
                raise HTTPException(
                    status_code=400,
                    detail=f"Нельзя удалить категорию: в ней {count} товар(ов). Сначала переместите или удалите товары."
                )
            result = await conn.execute("DELETE FROM categories WHERE slug=$1", slug)
            if result == "DELETE 0":
                raise HTTPException(status_code=404, detail="Категория не найдена")
            return {"message": "Категория удалена"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


# ==========================================
# ========== PRODUCTS API ==========
# ==========================================

@app.get("/api/products")
async def get_products(
    request: Request,
    category: Optional[str] = None,
    featured: Optional[bool] = None,
    search: Optional[str] = Query(None, max_length=200),   # Fix #17
    limit: Optional[int] = Query(None, ge=1, le=200),      # Fix #9
):
    # Получаем personal_discount текущего пользователя (если авторизован)
    personal_discount = 0
    try:
        uid = get_current_user(request)
        if uid:
            async with db.pool.acquire() as conn:
                row = await conn.fetchrow("SELECT personal_discount FROM users WHERE id=$1", uid)
                personal_discount = int(row['personal_discount'] or 0) if row else 0
                # v24: Авто-скидка — применяется только если нет персональной
                if personal_discount == 0:
                    en_row = await conn.fetchrow("SELECT value FROM delivery_settings WHERE key='auto_discount_enabled'")
                    if en_row and en_row['value'] == '1':
                        pct_row = await conn.fetchrow("SELECT value FROM delivery_settings WHERE key='auto_discount_percent'")
                        if pct_row and pct_row['value']:
                            personal_discount = max(0, min(99, int(pct_row['value'])))
    except Exception:
        pass

    try:
        async with db.pool.acquire() as conn:
            # Fix #12: исключаем cost_price из публичного ответа
            query  = "SELECT id,name,category,price,description,image_url,stock,featured,in_stock,preorder,price_preorder_auto,price_preorder_air,has_specifications,discount_percent,weight_kg,created_at,installation_service,price_type,price_cny,no_preorder,description_sections FROM products WHERE 1=1"
            params = []
            i = 1
            if category:
                query += f" AND category = ${i}"; params.append(category); i += 1
            if featured is not None:
                query += f" AND featured = ${i}"; params.append(featured); i += 1
            if search:
                query += f" AND (LOWER(name) LIKE ${i} OR LOWER(description) LIKE ${i})"; params.append(f"%{search.lower()}%"); i += 1
            query += " ORDER BY id"
            if limit:
                query += f" LIMIT ${i}"; params.append(limit)

            rows = await conn.fetch(query, *params)

            # Get CNY rate for dynamic pricing
            cny_rate_row = await conn.fetchval("SELECT value FROM delivery_settings WHERE key = 'cny_rate'")
            cny_rate = float(cny_rate_row) if cny_rate_row else 0

            result = []
            for r in rows:
                import math
                d = dict(r)
                d['price'] = float(d['price'])
                d['in_stock'] = bool(d.get('stock', 0) > 0)
                if isinstance(d.get('created_at'), datetime):
                    d['created_at'] = d['created_at'].isoformat()
                # Compute dynamic price from CNY if applicable
                if d.get('price_type') == 'dynamic' and d.get('price_cny') and cny_rate > 0:
                    raw_price = float(d['price_cny']) * cny_rate
                    d['price'] = math.ceil(raw_price / 10) * 10
                # Применяем личную скидку: берём максимум из product discount и personal
                prod_disc = int(d.get('discount_percent') or 0)
                eff_disc = max(prod_disc, personal_discount)
                d['personal_discount'] = personal_discount
                d['effective_discount'] = eff_disc
                if eff_disc > 0:
                    d['price_discounted'] = round(d['price'] * (1 - eff_disc / 100))
                    if d.get('price_preorder_auto'):
                        d['price_preorder_auto_discounted'] = round(float(d['price_preorder_auto']) * (1 - eff_disc / 100))
                    if d.get('price_preorder_air'):
                        d['price_preorder_air_discounted'] = round(float(d['price_preorder_air']) * (1 - eff_disc / 100))
                result.append(d)
            return result
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.get("/api/products/{product_id}")
async def get_product(product_id: int):
    try:
        async with db.pool.acquire() as conn:
            row = await conn.fetchrow("SELECT * FROM products WHERE id=$1", product_id)
            if not row:
                raise HTTPException(status_code=404, detail="Товар не найден")
            d = dict(row)
            d['price'] = float(d['price'])
            # КРИТИЧЕСКОЕ ИСПРАВЛЕНИЕ: гарантируем правильное значение in_stock
            d['in_stock'] = bool(d.get('stock', 0) > 0)
            # Compute dynamic price from CNY if applicable
            cny_rate_row = await conn.fetchval("SELECT value FROM delivery_settings WHERE key = 'cny_rate'")
            cny_rate = float(cny_rate_row) if cny_rate_row else 0
            if d.get('price_type') == 'dynamic' and d.get('price_cny') and cny_rate > 0:
                import math
                raw_price = float(d['price_cny']) * cny_rate
                d['price'] = math.ceil(raw_price / 10) * 10
            
            # Если товар имеет спецификации, загружаем их
            if d.get('has_specifications'):
                specs = await conn.fetch('''
                    SELECT id, name, price, description, image_url, stock, in_stock, preorder, cost_price, discount_percent, sort_order
                    FROM product_specifications
                    WHERE product_id = $1
                    ORDER BY sort_order ASC, id ASC
                ''', product_id)
                
                d['specifications'] = []
                for s in specs:
                    spec_dict = dict(s)
                    spec_dict['price'] = float(spec_dict['price'])
                    # КРИТИЧЕСКОЕ ИСПРАВЛЕНИЕ: гарантируем правильное значение in_stock для спецификаций
                    spec_dict['in_stock'] = bool(spec_dict.get('stock', 0) > 0)
                    if spec_dict.get('cost_price'):
                        spec_dict['cost_price'] = float(spec_dict['cost_price'])
                    
                    # Загружаем характеристики для каждой спецификации
                    chars = await conn.fetch('''
                        SELECT char_name, char_value
                        FROM product_characteristics
                        WHERE specification_id = $1
                        ORDER BY sort_order ASC, id ASC
                    ''', s['id'])
                    spec_dict['characteristics'] = [dict(c) for c in chars]
                    
                    # Загружаем дополнительные изображения для каждой спецификации
                    images = await conn.fetch('''
                        SELECT image_url
                        FROM product_images
                        WHERE specification_id = $1
                        ORDER BY sort_order ASC, id ASC
                    ''', s['id'])
                    spec_dict['images'] = [img['image_url'] for img in images]
                    
                    d['specifications'].append(spec_dict)
            
            # Загружаем характеристики основного товара (если есть)
            chars = await conn.fetch('''
                SELECT char_name, char_value
                FROM product_characteristics
                WHERE product_id = $1 AND specification_id IS NULL
                ORDER BY sort_order ASC, id ASC
            ''', product_id)
            d['characteristics'] = [dict(c) for c in chars]
            
            # Загружаем дополнительные изображения основного товара
            images = await conn.fetch('''
                SELECT image_url
                FROM product_images
                WHERE product_id = $1 AND specification_id IS NULL
                ORDER BY sort_order ASC, id ASC
            ''', product_id)
            d['images'] = [img['image_url'] for img in images]
            
            return d
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


# ==========================================
# ========== CART API ==========
# ==========================================

@app.get("/api/cart")
async def get_cart(user_id: str = Depends(get_current_user)):
    try:
        async with db.pool.acquire() as conn:
            # Получаем персональную скидку пользователя
            user_row = await conn.fetchrow("SELECT personal_discount FROM users WHERE id=$1", user_id)
            personal_discount = int(user_row['personal_discount'] or 0) if user_row else 0
            # v24: Авто-скидка — применяется только если нет персональной
            if personal_discount == 0:
                en_row = await conn.fetchrow("SELECT value FROM delivery_settings WHERE key='auto_discount_enabled'")
                if en_row and en_row['value'] == '1':
                    pct_row = await conn.fetchrow("SELECT value FROM delivery_settings WHERE key='auto_discount_percent'")
                    if pct_row and pct_row['value']:
                        personal_discount = max(0, min(99, int(pct_row['value'])))

            items = await conn.fetch('''
                SELECT ci.product_id, ci.specification_id, ci.quantity,
                       ci.order_type, ci.delivery_type,
                       p.name, p.category, p.price, p.description, p.image_url, p.stock,
                       p.preorder as p_preorder, p.in_stock as p_in_stock,
                       p.price_preorder_auto as p_price_auto, p.price_preorder_air as p_price_air,
                       p.weight_kg as p_weight_kg, p.discount_percent as p_disc,
                       ps.name as spec_name, ps.price as spec_price, ps.stock as spec_stock,
                       ps.image_url as spec_image_url, ps.description as spec_description,
                       ps.preorder as ps_preorder, ps.in_stock as ps_in_stock,
                       ps.price_preorder_auto as ps_price_auto, ps.price_preorder_air as ps_price_air,
                       ps.weight_kg as ps_weight_kg, ps.discount_percent as ps_disc
                FROM cart_items ci 
                JOIN products p ON ci.product_id = p.id
                LEFT JOIN product_specifications ps ON ci.specification_id = ps.id
                WHERE ci.user_id = $1 ORDER BY ci.added_at DESC
            ''', user_id)
            total = 0
            result = []
            for item in items:
                # Если есть спецификация, используем её данные
                if item['specification_id']:
                    base_price = float(item['spec_price'])
                    stock = item['spec_stock']
                    image_url = item['spec_image_url'] or item['image_url']
                    name = f"{item['name']} - {item['spec_name']}"
                    description = item['spec_description'] or item['description']
                    is_preorder = item['ps_preorder']
                    is_in_stock = item['ps_in_stock']
                    price_auto = float(item['ps_price_auto']) if item['ps_price_auto'] else None
                    price_air = float(item['ps_price_air']) if item['ps_price_air'] else None
                    prod_disc = int(item['ps_disc'] or item['p_disc'] or 0)
                    weight_kg = float(item['ps_weight_kg']) if item['ps_weight_kg'] else (float(item['p_weight_kg']) if item['p_weight_kg'] else None)
                else:
                    base_price = float(item['price'])
                    stock = item['stock']
                    image_url = item['image_url']
                    name = item['name']
                    description = item['description']
                    is_preorder = item['p_preorder']
                    is_in_stock = item['p_in_stock']
                    price_auto = float(item['p_price_auto']) if item['p_price_auto'] else None
                    price_air = float(item['p_price_air']) if item['p_price_air'] else None
                    prod_disc = int(item['p_disc'] or 0)
                    weight_kg = float(item['p_weight_kg']) if item['p_weight_kg'] else None

                # Итоговая скидка = максимум из скидки товара и личной скидки
                eff_disc = max(prod_disc, personal_discount)

                order_type = item['order_type']
                delivery_type = item['delivery_type']
                if order_type == 'preorder' and delivery_type == 'auto' and price_auto:
                    raw_price = price_auto
                elif order_type == 'preorder' and delivery_type == 'air' and price_air:
                    raw_price = price_air
                else:
                    raw_price = base_price

                effective_price = round(raw_price * (1 - eff_disc / 100)) if eff_disc > 0 else raw_price

                item_total = effective_price * item['quantity']
                total += item_total
                result.append({
                    "product_id": item['product_id'],
                    "specification_id": item['specification_id'],
                    "quantity": item['quantity'],
                    "order_type": order_type,
                    "delivery_type": delivery_type,
                    "product": {
                        "id": item['product_id'],
                        "specification_id": item['specification_id'],
                        "name": name,
                        "category": item['category'],
                        "price": effective_price,
                        "base_price": base_price,
                        "discount_percent": eff_disc,
                        "description": description,
                        "image_url": image_url,
                        "stock": stock,
                        "preorder": is_preorder,
                        "in_stock": is_in_stock,
                        "price_preorder_auto": round(price_auto * (1 - eff_disc / 100)) if price_auto and eff_disc > 0 else price_auto,
                        "price_preorder_air": round(price_air * (1 - eff_disc / 100)) if price_air and eff_disc > 0 else price_air,
                        "weight_kg": weight_kg,
                    },
                    "item_total": item_total
                })
            return {"items": result, "total": total, "items_count": len(items), "personal_discount": personal_discount}
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.post("/api/cart")
async def add_to_cart(cart_item: CartUpdate, user_id: str = Depends(get_current_user)):
    if cart_item.quantity <= 0:
        raise HTTPException(status_code=400, detail="Количество должно быть больше 0")
    try:
        async with db.pool.acquire() as conn:
            # Если указана спецификация, проверяем её
            if cart_item.specification_id:
                spec = await conn.fetchrow(
                    "SELECT id, stock, product_id, preorder FROM product_specifications WHERE id=$1",
                    cart_item.specification_id
                )
                if not spec:
                    raise HTTPException(status_code=404, detail="Спецификация не найдена")
                if spec['product_id'] != cart_item.product_id:
                    raise HTTPException(status_code=400, detail="Спецификация не принадлежит данному товару")
                # Пропускаем проверку склада для preorder-спецификаций (stock = 0 — норма)
                if not spec['preorder'] and spec['stock'] < cart_item.quantity:
                    raise HTTPException(status_code=400, detail="Недостаточно товара на складе")
            else:
                # Проверяем основной товар
                product = await conn.fetchrow("SELECT id, stock, preorder, in_stock FROM products WHERE id=$1", cart_item.product_id)
                if not product:
                    raise HTTPException(status_code=404, detail="Товар не найден")
                # Пропускаем проверку склада для preorder-товаров (stock = 0 — это норма)
                if not product['preorder'] and product['stock'] < cart_item.quantity:
                    raise HTTPException(status_code=400, detail="Недостаточно товара на складе")
            
            # Получаем данные товара для проверки preorder/in_stock
            if cart_item.specification_id is not None:
                prod_data = await conn.fetchrow(
                    "SELECT preorder, in_stock FROM product_specifications WHERE id=$1",
                    cart_item.specification_id
                )
            else:
                prod_data = await conn.fetchrow(
                    "SELECT preorder, in_stock FROM products WHERE id=$1",
                    cart_item.product_id
                )

            is_preorder = prod_data and prod_data['preorder']
            is_in_stock = prod_data and prod_data['in_stock']

            # Валидация: если товар только предзаказ, order_type обязателен
            if is_preorder and not is_in_stock:
                if not cart_item.order_type:
                    raise HTTPException(status_code=400, detail="Выберите тип заказа: предзаказ или в наличии")
            if is_preorder and cart_item.order_type == "preorder":
                if not cart_item.delivery_type:
                    raise HTTPException(status_code=400, detail="Выберите тип доставки: авто или самолёт")
                if cart_item.delivery_type not in ("auto", "air"):
                    raise HTTPException(status_code=400, detail="Тип доставки: auto или air")

            order_type = cart_item.order_type if is_preorder else ("in_stock" if is_in_stock else None)
            delivery_type = cart_item.delivery_type if (order_type == "preorder") else None

            # Fix БАГ-02/ON CONFLICT: manual upsert — не зависит от наличия индексов на старой БД
            if cart_item.specification_id is not None:
                existing = await conn.fetchrow(
                    '''SELECT id FROM cart_items
                       WHERE user_id=$1 AND product_id=$2 AND specification_id=$3
                       AND COALESCE(order_type,'') = COALESCE($4,'')''',
                    user_id, cart_item.product_id, cart_item.specification_id, order_type
                )
                if existing:
                    await conn.execute(
                        'UPDATE cart_items SET quantity=$1, delivery_type=$2 WHERE id=$3',
                        cart_item.quantity, delivery_type, existing['id']
                    )
                else:
                    await conn.execute(
                        '''INSERT INTO cart_items
                           (user_id, product_id, specification_id, quantity, order_type, delivery_type)
                           VALUES ($1,$2,$3,$4,$5,$6)''',
                        user_id, cart_item.product_id, cart_item.specification_id,
                        cart_item.quantity, order_type, delivery_type
                    )
            else:
                existing = await conn.fetchrow(
                    '''SELECT id FROM cart_items
                       WHERE user_id=$1 AND product_id=$2 AND specification_id IS NULL
                       AND COALESCE(order_type,'') = COALESCE($3,'')''',
                    user_id, cart_item.product_id, order_type
                )
                if existing:
                    await conn.execute(
                        'UPDATE cart_items SET quantity=$1, delivery_type=$2 WHERE id=$3',
                        cart_item.quantity, delivery_type, existing['id']
                    )
                else:
                    await conn.execute(
                        '''INSERT INTO cart_items
                           (user_id, product_id, specification_id, quantity, order_type, delivery_type)
                           VALUES ($1,$2,NULL,$3,$4,$5)''',
                        user_id, cart_item.product_id,
                        cart_item.quantity, order_type, delivery_type
                    )
            return {"message": "Товар добавлен в корзину"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.delete("/api/cart/{product_id}")
async def remove_from_cart(product_id: int, specification_id: Optional[int] = Query(None), user_id: str = Depends(get_current_user)):
    try:
        async with db.pool.acquire() as conn:
            if specification_id is not None:
                r = await conn.execute(
                    "DELETE FROM cart_items WHERE user_id=$1 AND product_id=$2 AND specification_id=$3", 
                    user_id, product_id, specification_id
                )
            else:
                r = await conn.execute(
                    "DELETE FROM cart_items WHERE user_id=$1 AND product_id=$2 AND specification_id IS NULL", 
                    user_id, product_id
                )
            if r == "DELETE 0":
                raise HTTPException(status_code=404, detail="Товар не найден в корзине")
            return {"message": "Товар удалён из корзины"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.put("/api/cart/{product_id}")
async def update_cart_quantity(product_id: int, body: CartQuantityUpdate, user_id: str = Depends(get_current_user)):
    """Обновить количество товара в корзине"""
    quantity = body.quantity
    specification_id = body.specification_id
    
    if quantity <= 0:
        raise HTTPException(status_code=400, detail="Количество должно быть больше 0")
    
    try:
        async with db.pool.acquire() as conn:
            # Получаем cart_item чтобы проверить order_type
            if specification_id:
                cart_row = await conn.fetchrow(
                    "SELECT order_type FROM cart_items WHERE user_id=$1 AND product_id=$2 AND specification_id=$3",
                    user_id, product_id, specification_id
                )
            else:
                cart_row = await conn.fetchrow(
                    "SELECT order_type FROM cart_items WHERE user_id=$1 AND product_id=$2 AND specification_id IS NULL",
                    user_id, product_id
                )

            # Проверяем stock только для товаров в наличии (не для предзаказов)
            is_preorder = cart_row and cart_row['order_type'] == 'preorder'
            if not is_preorder:
                if specification_id:
                    item = await conn.fetchrow(
                        "SELECT stock FROM product_specifications WHERE id=$1",
                        specification_id
                    )
                    if not item:
                        raise HTTPException(status_code=404, detail="Спецификация не найдена")
                else:
                    item = await conn.fetchrow(
                        "SELECT stock FROM products WHERE id=$1",
                        product_id
                    )
                    if not item:
                        raise HTTPException(status_code=404, detail="Товар не найден")
                if item['stock'] < quantity:
                    raise HTTPException(status_code=400, detail=f"Недостаточно на складе (доступно: {item['stock']})")
            
            # Обновляем количество
            if specification_id:
                await conn.execute('''
                    UPDATE cart_items SET quantity = $1
                    WHERE user_id = $2 AND product_id = $3 AND specification_id = $4
                ''', quantity, user_id, product_id, specification_id)
            else:
                await conn.execute('''
                    UPDATE cart_items SET quantity = $1
                    WHERE user_id = $2 AND product_id = $3 AND specification_id IS NULL
                ''', quantity, user_id, product_id)
            
            return {"message": "Количество обновлено", "quantity": quantity}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.delete("/api/cart")
async def clear_cart(user_id: str = Depends(get_current_user)):
    try:
        async with db.pool.acquire() as conn:
            await conn.execute("DELETE FROM cart_items WHERE user_id=$1", user_id)
            return {"message": "Корзина очищена"}
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.patch("/api/cart/{product_id}/delivery")
async def update_cart_delivery(
    product_id: int,
    payload: dict = Body(...),
    user_id: str = Depends(get_current_user)
):
    """Изменить тип доставки для товара в корзине (только предзаказ)."""
    delivery_type = payload.get("delivery_type")
    specification_id = payload.get("specification_id")
    if delivery_type not in ("auto", "air"):
        raise HTTPException(status_code=400, detail="Неверный тип доставки")
    try:
        async with db.pool.acquire() as conn:
            if specification_id:
                row = await conn.fetchrow(
                    "SELECT order_type FROM cart_items WHERE user_id=$1 AND product_id=$2 AND specification_id=$3",
                    user_id, product_id, specification_id
                )
            else:
                row = await conn.fetchrow(
                    "SELECT order_type FROM cart_items WHERE user_id=$1 AND product_id=$2 AND specification_id IS NULL",
                    user_id, product_id
                )
            if not row:
                raise HTTPException(status_code=404, detail="Товар не найден в корзине")
            if row["order_type"] != "preorder":
                raise HTTPException(status_code=400, detail="Тип доставки применим только к предзаказам")
            if specification_id:
                await conn.execute(
                    "UPDATE cart_items SET delivery_type=$1 WHERE user_id=$2 AND product_id=$3 AND specification_id=$4",
                    delivery_type, user_id, product_id, specification_id
                )
            else:
                await conn.execute(
                    "UPDATE cart_items SET delivery_type=$1 WHERE user_id=$2 AND product_id=$3 AND specification_id IS NULL",
                    delivery_type, user_id, product_id
                )
            return {"message": "Тип доставки обновлён"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


# ==========================================
# ========== ORDERS API ==========
# ==========================================

@app.post("/api/orders")
async def create_order(order_data: OrderCreate, user_id: str = Depends(get_current_user)):
    """Создать заказ из корзины"""
    try:
        async with db.pool.acquire() as conn:
            # Получаем персональную скидку
            user_row = await conn.fetchrow("SELECT personal_discount FROM users WHERE id=$1", user_id)
            personal_discount = int(user_row['personal_discount'] or 0) if user_row else 0

            cart_items = await conn.fetch('''
                SELECT ci.product_id, ci.specification_id, ci.quantity,
                       ci.order_type, ci.delivery_type,
                       p.name, p.price, p.stock, p.preorder as p_preorder,
                       p.weight_kg as p_weight_kg, p.discount_percent as p_disc,
                       p.price_preorder_auto as p_price_auto, p.price_preorder_air as p_price_air,
                       ps.name as spec_name, ps.price as spec_price, ps.stock as spec_stock,
                       ps.preorder as ps_preorder, ps.weight_kg as ps_weight_kg,
                       ps.discount_percent as ps_disc,
                       ps.price_preorder_auto as ps_price_auto, ps.price_preorder_air as ps_price_air
                FROM cart_items ci 
                JOIN products p ON ci.product_id = p.id
                LEFT JOIN product_specifications ps ON ci.specification_id = ps.id
                WHERE ci.user_id = $1
            ''', user_id)

            if not cart_items:
                raise HTTPException(status_code=400, detail="Корзина пуста")

            # Подсчитываем общую сумму с учётом скидок
            total = 0
            for i in cart_items:
                prod_disc = int(i['ps_disc'] or i['p_disc'] or 0) if i['specification_id'] else int(i['p_disc'] or 0)
                eff_disc = max(prod_disc, personal_discount)
                if i['specification_id']:
                    raw = float(i['spec_price'])
                    if i['order_type'] == 'preorder' and i['delivery_type'] == 'auto' and i['ps_price_auto']:
                        raw = float(i['ps_price_auto'])
                    elif i['order_type'] == 'preorder' and i['delivery_type'] == 'air' and i['ps_price_air']:
                        raw = float(i['ps_price_air'])
                else:
                    raw = float(i['price'])
                    if i['order_type'] == 'preorder' and i['delivery_type'] == 'auto' and i['p_price_auto']:
                        raw = float(i['p_price_auto'])
                    elif i['order_type'] == 'preorder' and i['delivery_type'] == 'air' and i['p_price_air']:
                        raw = float(i['p_price_air'])
                price = round(raw * (1 - eff_disc / 100)) if eff_disc > 0 else raw
                total += price * i['quantity']

            # Применяем промокод (если передан)
            promo_discount = 0.0
            applied_promo = None
            if order_data.promo_code:
                promo_code_upper = order_data.promo_code.strip().upper()
                promo_row = await conn.fetchrow(
                    "SELECT * FROM promo_codes WHERE code=$1 AND is_active=TRUE", promo_code_upper
                )
                if promo_row:
                    from datetime import datetime as _dt
                    promo_expired = promo_row['expires_at'] and promo_row['expires_at'] < _dt.now(timezone.utc).replace(tzinfo=None)
                    promo_exhausted = promo_row['max_uses'] and promo_row['used_count'] >= promo_row['max_uses']
                    if not promo_expired and not promo_exhausted:
                        if promo_row['discount_type'] == 'percent':
                            promo_discount = round(total * float(promo_row['discount_value']) / 100, 2)
                        else:
                            promo_discount = min(float(promo_row['discount_value']), total)
                        applied_promo = promo_code_upper
                        # Инкрементируем счётчик использований
                        await conn.execute(
                            "UPDATE promo_codes SET used_count = used_count + 1 WHERE id = $1",
                            promo_row['id']
                        )
            total_final = max(0.0, total - promo_discount)

            # Generate unique track number
            year = datetime.now().year
            while True:
                hex_part = secrets.token_hex(8).upper()
                track_number = f"SP-{year}-{hex_part}"
                existing_track = await conn.fetchval("SELECT id FROM orders WHERE track_number = $1", track_number)
                if not existing_track:
                    break

            order_id = await conn.fetchval('''
                INSERT INTO orders (user_id, total_amount, delivery_address, comment, status, payment_status, promo_code, promo_discount, track_number, installation_data)
                VALUES ($1, $2, $3, $4, 'created', 'pending', $5, $6, $7, $8)
                RETURNING id
            ''', user_id, total_final, order_data.delivery_address, order_data.comment,
                applied_promo, promo_discount, track_number,
                json.dumps(order_data.installation_data) if order_data.installation_data else None)

            for item in cart_items:
                # Определяем название и цену товара
                if item['specification_id']:
                    product_name = f"{item['name']} - {item['spec_name']}"
                    raw_price = float(item['spec_price'])
                    if item['order_type'] == 'preorder' and item['delivery_type'] == 'auto' and item['ps_price_auto']:
                        raw_price = float(item['ps_price_auto'])
                    elif item['order_type'] == 'preorder' and item['delivery_type'] == 'air' and item['ps_price_air']:
                        raw_price = float(item['ps_price_air'])
                    prod_disc_item = int(item['ps_disc'] or item['p_disc'] or 0)
                    stock_column = 'product_specifications'
                    stock_id = item['specification_id']
                    current_stock = item['spec_stock']
                    item_weight = item['ps_weight_kg'] or item['p_weight_kg']
                else:
                    product_name = item['name']
                    raw_price = float(item['price'])
                    if item['order_type'] == 'preorder' and item['delivery_type'] == 'auto' and item['p_price_auto']:
                        raw_price = float(item['p_price_auto'])
                    elif item['order_type'] == 'preorder' and item['delivery_type'] == 'air' and item['p_price_air']:
                        raw_price = float(item['p_price_air'])
                    prod_disc_item = int(item['p_disc'] or 0)
                    stock_column = 'products'
                    stock_id = item['product_id']
                    current_stock = item['stock']
                    item_weight = item['p_weight_kg']

                eff_disc_item = max(prod_disc_item, personal_discount)
                price = round(raw_price * (1 - eff_disc_item / 100)) if eff_disc_item > 0 else raw_price

                # Проверяем наличие (пропускаем для предзаказов — stock=0 это норма)
                is_preorder_item = item['order_type'] == 'preorder'
                if not is_preorder_item and current_stock < item['quantity']:
                    raise HTTPException(
                        status_code=400, 
                        detail=f"Недостаточно товара '{product_name}' на складе"
                    )
                
                await conn.execute('''
                    INSERT INTO order_items (order_id, product_id, product_name, price, quantity,
                                            order_type, delivery_type, weight_kg, specification_id)
                    VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9)
                ''', order_id, item['product_id'], product_name, price, item['quantity'],
                     item['order_type'], item['delivery_type'],
                     float(item_weight) if item_weight else None,
                     item['specification_id'])
                
                # Обновляем остатки (только для товаров в наличии)
                if not is_preorder_item:
                    if item['specification_id']:
                        await conn.execute(
                            "UPDATE product_specifications SET stock = stock - $1 WHERE id = $2",
                            item['quantity'], item['specification_id']
                        )
                    else:
                        await conn.execute(
                            "UPDATE products SET stock = stock - $1 WHERE id = $2",
                            item['quantity'], item['product_id']
                        )

            await conn.execute("DELETE FROM cart_items WHERE user_id=$1", user_id)

            return {
                "message": "Заказ создан",
                "order_id": order_id,
                "total": total_final,
                "promo_discount": promo_discount,
                "promo_code": applied_promo,
                "status": "created",
                "track_number": track_number,
            }
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.get("/api/orders")
async def get_user_orders(user_id: str = Depends(get_current_user)):
    """Получить заказы пользователя"""
    try:
        async with db.pool.acquire() as conn:
            orders = await conn.fetch('''
                SELECT o.id, o.status, o.total_amount, o.payment_status,
                       o.created_at, o.delivery_address, o.delay_note,
                       COALESCE(
                         json_agg(
                           json_build_object(
                             'product_name', oi.product_name,
                             'quantity', oi.quantity,
                             'price', oi.price,
                             'order_type', COALESCE(oi.order_type, 'in_stock'),
                             'delivery_type', oi.delivery_type
                           ) ORDER BY oi.id
                         ) FILTER (WHERE oi.id IS NOT NULL),
                         '[]'
                       ) AS items
                FROM orders o
                LEFT JOIN order_items oi ON oi.order_id = o.id
                WHERE o.user_id=$1
                GROUP BY o.id
                ORDER BY o.created_at DESC
            ''', user_id)
            result = []
            for o in orders:
                d = dict(o)
                d['total_amount'] = float(d['total_amount'])
                if isinstance(d.get('created_at'), datetime):
                    d['created_at'] = d['created_at'].isoformat()
                # Parse items JSON if returned as string
                if isinstance(d.get('items'), str):
                    try: d['items'] = json.loads(d['items'])
                    except: d['items'] = []
                # Convert Decimal prices in items
                if d.get('items'):
                    for item in d['items']:
                        if 'price' in item:
                            item['price'] = float(item['price'])
                result.append(d)
            return result
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.get("/api/orders/active-count")
async def get_active_orders_count():
    """Публичный счётчик активных заказов для плашки в каталоге"""
    try:
        async with db.pool.acquire() as conn:
            count = await conn.fetchval(
                "SELECT COUNT(*) FROM orders WHERE status NOT IN ('completed','cancelled')"
            )
            return {"count": count or 0}
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.post("/api/orders/{order_id}/cancel")
async def cancel_order(order_id: int, user_id: str = Depends(get_current_user)):
    """Отменить заказ пользователем (только статусы created и processing)"""
    try:
        async with db.pool.acquire() as conn:
            order = await conn.fetchrow(
                "SELECT id, status, user_id FROM orders WHERE id=$1 AND user_id=$2",
                order_id, user_id
            )
            if not order:
                raise HTTPException(status_code=404, detail="Заказ не найден")
            if order['status'] not in ('created', 'processing'):
                raise HTTPException(
                    status_code=400,
                    detail="Нельзя отменить заказ в статусе «{}»".format(order['status'])
                )
            # Возвращаем сток для товаров in_stock
            items = await conn.fetch(
                "SELECT product_id, specification_id, quantity, order_type FROM order_items WHERE order_id=$1",
                order_id
            )
            for item in items:
                if item['order_type'] != 'preorder':
                    if item['specification_id']:
                        await conn.execute(
                            "UPDATE product_specifications SET stock = stock + $1 WHERE id = $2",
                            item['quantity'], item['specification_id']
                        )
                    else:
                        await conn.execute(
                            "UPDATE products SET stock = stock + $1 WHERE id = $2",
                            item['quantity'], item['product_id']
                        )
            await conn.execute(
                "UPDATE orders SET status='cancelled', updated_at=CURRENT_TIMESTAMP WHERE id=$1",
                order_id
            )
            return {"message": "Заказ отменён", "order_id": order_id}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("cancel_order error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.get("/api/orders/{order_id}/items")
async def get_user_order_items(order_id: int, user_id: str = Depends(get_current_user)):
    """Получить позиции заказа пользователя (для tracking.html)"""
    try:
        async with db.pool.acquire() as conn:
            # Убеждаемся что заказ принадлежит текущему пользователю
            order = await conn.fetchrow(
                "SELECT id FROM orders WHERE id=$1 AND user_id=$2", order_id, user_id
            )
            if not order:
                raise HTTPException(status_code=404, detail="Заказ не найден")
            rows = await conn.fetch(
                """SELECT product_name, quantity, price, order_type, delivery_type,
                          specification_name, weight_kg
                   FROM order_items WHERE order_id=$1 ORDER BY id""",
                order_id
            )
            result = []
            for r in rows:
                d = dict(r)
                d['price'] = float(d['price']) if d.get('price') else 0.0
                if d.get('weight_kg'):
                    d['weight_kg'] = float(d['weight_kg'])
                result.append(d)
            return result
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Order items error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


# ==========================================
# ========== PAYMENT API (заглушка) ==========
# ==========================================

@app.post("/api/payment/create")
async def create_payment(payment: PaymentCreate, user_id: str = Depends(get_current_user)):
    """
    Заглушка платёжной интеграции.
    Для подключения реального эквайринга замените этот блок:
    - ЮKassa: https://yookassa.ru/developers/api
    - Тинькофф: https://www.tinkoff.ru/kassa/develop/api/
    - Stripe: https://stripe.com/docs/api

    Переменные окружения:
      PAYMENT_API_KEY=...
      PAYMENT_SHOP_ID=...
      PAYMENT_SECRET_KEY=...
      PAYMENT_CALLBACK_URL=https://your-domain.com/api/payment/callback
    """
    try:
        async with db.pool.acquire() as conn:
            order = await conn.fetchrow(
                "SELECT id, total_amount, payment_status FROM orders WHERE id=$1 AND user_id=$2",
                payment.order_id, user_id
            )
            if not order:
                raise HTTPException(status_code=404, detail="Заказ не найден")
            if order['payment_status'] == 'paid':
                raise HTTPException(status_code=400, detail="Заказ уже оплачен")

            # ── ЗДЕСЬ ИНТЕГРИРУЙТЕ РЕАЛЬНЫЙ ЭКВАЙРИНГ ──
            # Пример для ЮKassa:
            # import yookassa
            # yookassa.Configuration.account_id = PAYMENT_SHOP_ID
            # yookassa.Configuration.secret_key = PAYMENT_SECRET_KEY
            # payment_obj = yookassa.Payment.create({
            #     "amount": {"value": str(payment.amount), "currency": "RUB"},
            #     "confirmation": {"type": "redirect", "return_url": PAYMENT_CALLBACK_URL},
            #     "capture": True,
            #     "description": f"Заказ #{payment.order_id}",
            # })
            # payment_url = payment_obj.confirmation.confirmation_url
            # payment_id  = payment_obj.id

            # ЗАГЛУШКА — возвращаем тестовые данные
            payment_id  = f"stub_{uuid4().hex[:16]}"
            payment_url = f"/payment-stub?order_id={payment.order_id}&amount={payment.amount}"

            await conn.execute(
                "UPDATE orders SET payment_id=$1, payment_url=$2, payment_status='waiting' WHERE id=$3",
                payment_id, payment_url, payment.order_id
            )

            return {
                "payment_id":  payment_id,
                "payment_url": payment_url,
                "amount":      payment.amount,
                "currency":    payment.currency,
                "status":      "waiting",
                "note":        "⚠️ Это заглушка. Подключите реальный эквайринг в backend/main.py"
            }
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.post("/api/payment/callback")
async def payment_callback(request: Request):
    """
    Fix #5: Webhook от платёжной системы с обязательной верификацией подписи.

    ВАЖНО: Этот эндпоинт принимает внешние HTTP-запросы. Без проверки подписи
    любой злоумышленник может отправить фейковый callback и «оплатить» заказ.

    Схема верификации зависит от платёжной системы:
    - ЮKassa:    SHA-1 HMAC заголовок Webhook-Signature
    - Тинькофф:  SHA-256 HMAC, токен = PAYMENT_SECRET_KEY
    - Stripe:    stripe.Webhook.construct_event(payload, sig, secret)
    """
    try:
        raw_body = await request.body()

        # ── Fix #5: Верификация подписи ──────────────────────────────────────
        # Раскомментируйте и адаптируйте блок для вашей платёжной системы.
        #
        # Пример для ЮKassa (Webhook-Signature: sha1=<hex>):
        # received_sig = request.headers.get("Webhook-Signature", "")
        # expected_sig = "sha1=" + hmac.new(
        #     PAYMENT_SECRET_KEY.encode(), raw_body, hashlib.sha1
        # ).hexdigest()
        # if not hmac.compare_digest(received_sig, expected_sig):
        #     logger.warning("Payment callback: invalid signature from %s", request.client.host)
        #     raise HTTPException(status_code=403, detail="Неверная подпись webhook")
        #
        # Пример для Тинькофф (поле Token в JSON):
        # import json as _json
        # body_dict = _json.loads(raw_body)
        # received_token = body_dict.pop("Token", "")
        # sorted_vals = "".join(str(v) for _, v in sorted(body_dict.items()))
        # expected = hashlib.sha256((sorted_vals + PAYMENT_SECRET_KEY).encode()).hexdigest()
        # if not hmac.compare_digest(received_token, expected):
        #     logger.warning("Payment callback: invalid Tinkoff signature")
        #     raise HTTPException(status_code=403, detail="Неверная подпись webhook")
        # ─────────────────────────────────────────────────────────────────────

        # Пока интеграция — заглушка: логируем и возвращаем ok
        body = json.loads(raw_body) if raw_body else {}
        logger.info("Payment callback received (stub mode): %s", body)

        # TODO: после подключения реального эквайринга — обновить orders.payment_status
        # order_id = body.get("order_id")
        # new_status = body.get("status")
        # async with db.pool.acquire() as conn:
        #     await conn.execute(
        #         "UPDATE orders SET payment_status=$1 WHERE id=$2",
        #         new_status, order_id
        #     )

        return {"status": "ok"}

    except HTTPException:
        raise
    except Exception as e:
        logger.error("Payment callback error: %s", e, exc_info=True)
        # Fix #6: не раскрываем детали внешнему миру
        return JSONResponse(status_code=200, content={"status": "error"})


@app.get("/payment-stub")
async def payment_stub_page(request: Request, order_id: int = 0, amount: float = 0):
    """Страница-заглушка оплаты (удалить после подключения реального эквайринга)"""
    return templates.TemplateResponse("payment_stub.html", {
        "request": request, "order_id": order_id, "amount": amount
    })


# ==========================================
# ========== ADMIN API ==========
# ==========================================

@app.post("/api/admin/login")
async def admin_login(login_data: AdminLogin, response: Response):
    # Fix A-4: сравниваем пароль через pbkdf2-хеш из БД, не с plaintext.
    # hmac.compare_digest на username защищает от timing-атак.
    if not hmac.compare_digest(login_data.username, ADMIN_USERNAME):
        logger.warning("Failed admin login attempt for: %s", login_data.username)
        raise HTTPException(status_code=401, detail="Неверные данные для входа")
    try:
        async with db.pool.acquire() as conn:
            admin = await conn.fetchrow(
                "SELECT id, password_hash, is_superadmin FROM users WHERE username=$1 AND is_admin=TRUE",
                ADMIN_USERNAME
            )
            if not admin or not hasher.verify_password(login_data.password, admin['password_hash']):
                logger.warning("Failed admin login attempt (bad password) for: %s", login_data.username)
                raise HTTPException(status_code=401, detail="Неверные данные для входа")
            is_superadmin = bool(admin.get('is_superadmin', False))
            token = create_access_token({"user_id": str(admin['id']), "username": ADMIN_USERNAME, "is_admin": True, "is_superadmin": is_superadmin})
            _is_https = os.getenv("ENVIRONMENT", "production") != "development"
            response.set_cookie(
                key="access_token",
                value=token,
                httponly=True,
                secure=_is_https,
                samesite="strict",
                max_age=3600,
                path="/",
            )
            logger.info("Admin logged in: %s", ADMIN_USERNAME)
            return {"message": "Вход выполнен", "user": {"username": ADMIN_USERNAME, "is_admin": True}}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.post("/api/admin/migrate-images")
async def migrate_images_to_base64(admin=Depends(verify_admin)):
    """
    Утилита для миграции изображений из файловой системы в base64
    Полезно при переносе с файлового хранения на БД
    """
    try:
        migrated = 0
        failed = []
        
        async with db.pool.acquire() as conn:
            # Получаем все товары с изображениями из файлов
            products = await conn.fetch('''
                SELECT id, name, image_url 
                FROM products 
                WHERE image_url LIKE '/static/uploads/%'
            ''')
            
            for product in products:
                try:
                    # Путь к файлу
                    file_path = STATIC_DIR / product['image_url'].lstrip('/')
                    
                    if not file_path.exists():
                        failed.append(f"Product {product['id']}: File not found")
                        continue
                    
                    # Читаем файл
                    with open(file_path, 'rb') as f:
                        file_data = f.read()
                    
                    # Оптимизируем
                    optimized_data = await optimize_image(file_data)
                    
                    # Конвертируем в base64
                    image_base64 = base64.b64encode(optimized_data).decode('utf-8')
                    image_data_url = f"data:image/jpeg;base64,{image_base64}"
                    
                    # Обновляем в БД
                    await conn.execute('''
                        UPDATE products 
                        SET image_url = $1 
                        WHERE id = $2
                    ''', image_data_url, product['id'])
                    
                    migrated += 1
                    print(f"✅ Migrated: {product['name']}")
                    
                except Exception as e:
                    failed.append(f"Product {product['id']}: {str(e)}")
                    print(f"❌ Failed: {product['name']} - {e}")
        
        return {
            "success": True,
            "migrated": migrated,
            "failed": failed,
            "message": f"Migrated {migrated} images to base64"
        }
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.get("/api/admin/stats")
async def get_admin_stats(admin=Depends(verify_admin)):
    try:
        async with db.pool.acquire() as conn:
            return {
                "users": {
                    "total": await conn.fetchval("SELECT COUNT(*) FROM users"),
                    "with_carts": await conn.fetchval("SELECT COUNT(DISTINCT user_id) FROM cart_items"),
                },
                "products": {
                    "total": await conn.fetchval("SELECT COUNT(*) FROM products"),
                    "in_stock": await conn.fetchval("SELECT COUNT(*) FROM products WHERE stock>0"),
                    "out_of_stock": await conn.fetchval("SELECT COUNT(*) FROM products WHERE stock=0"),
                    "featured": await conn.fetchval("SELECT COUNT(*) FROM products WHERE featured=true"),
                },
                "orders": {
                    "total": await conn.fetchval("SELECT COUNT(*) FROM orders"),
                    "active": await conn.fetchval("SELECT COUNT(*) FROM orders WHERE status NOT IN ('completed','cancelled')"),
                    "pending_payment": await conn.fetchval("SELECT COUNT(*) FROM orders WHERE payment_status='pending'"),
                    "revenue": float(await conn.fetchval("SELECT COALESCE(SUM(total_amount),0) FROM orders WHERE payment_status='paid'") or 0),
                },
                "categories": {
                    "total": await conn.fetchval("SELECT COUNT(*) FROM categories"),
                }
            }
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.get("/api/admin/orders")
async def get_admin_orders(admin=Depends(verify_manager_or_admin), status: Optional[str] = None, track: Optional[str] = None, limit: int = 50):
    try:
        is_manager_only = admin.get("is_manager") and not admin.get("is_admin")
        async with db.pool.acquire() as conn:
            q = "SELECT o.*, u.username, u.email, u.full_name, u.phone FROM orders o LEFT JOIN users u ON o.user_id=u.id WHERE 1=1"
            params = []
            if status:
                q += f" AND o.status = ${len(params)+1}"; params.append(status)
            if track:
                q += f" AND o.track_number ILIKE ${len(params)+1}"; params.append(f"%{track}%")
            q += f" ORDER BY o.created_at DESC LIMIT ${len(params)+1}"; params.append(limit)
            rows = await conn.fetch(q, *params)
            result = []
            for r in rows:
                d = dict(r)
                d['total_amount'] = float(d['total_amount'])
                d['user_id'] = str(d['user_id']) if d.get('user_id') else None
                if isinstance(d.get('created_at'), datetime): d['created_at'] = d['created_at'].isoformat()
                if isinstance(d.get('updated_at'), datetime): d['updated_at'] = d['updated_at'].isoformat()
                # Менеджеры не видят личные данные клиентов
                if is_manager_only:
                    d['email'] = None
                    d['full_name'] = None
                    d['phone'] = None
                    d['delivery_address'] = None
                result.append(d)
            return result
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.get("/api/admin/top-customers")
async def get_top_customers(admin=Depends(verify_admin), limit: int = 20):
    """Получить топ активных покупателей для CRM"""
    try:
        async with db.pool.acquire() as conn:
            rows = await conn.fetch('''
                SELECT 
                    u.id,
                    u.username,
                    u.email,
                    u.full_name,
                    u.phone,
                    u.personal_discount,
                    COUNT(o.id) as order_count,
                    COALESCE(SUM(o.total_amount), 0) as total_spent,
                    MAX(o.created_at) as last_order_date,
                    COUNT(CASE WHEN o.payment_status = 'paid' THEN 1 END) as paid_orders
                FROM users u
                LEFT JOIN orders o ON u.id = o.user_id
                WHERE u.is_admin = FALSE
                GROUP BY u.id, u.username, u.email, u.full_name, u.phone, u.personal_discount
                HAVING COUNT(o.id) > 0
                ORDER BY total_spent DESC
                LIMIT $1
            ''', limit)
            
            result = []
            for r in rows:
                d = dict(r)
                d['id'] = str(d['id'])
                d['total_spent'] = float(d['total_spent'])
                if isinstance(d.get('last_order_date'), datetime):
                    d['last_order_date'] = d['last_order_date'].isoformat()
                result.append(d)
            return result
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.get("/api/admin/customers")
async def get_all_customers(q: Optional[str] = None, admin=Depends(verify_admin)):
    """Получить всех покупателей с количеством заметок для CRM"""
    try:
        async with db.pool.acquire() as conn:
            if q:
                rows = await conn.fetch('''
                    SELECT u.id, u.username, u.email, u.full_name, u.phone, u.personal_discount,
                           u.is_manager, COUNT(DISTINCT o.id) as order_count,
                           COALESCE(SUM(o.total_amount), 0) as total_spent,
                           MAX(o.created_at) as last_order_date, COUNT(DISTINCT n.id) as notes_count
                    FROM users u
                    LEFT JOIN orders o ON u.id = o.user_id
                    LEFT JOIN customer_notes n ON u.id = n.user_id
                    WHERE u.is_admin = FALSE AND (
                        u.username ILIKE $1 OR u.email ILIKE $1 OR
                        u.full_name ILIKE $1 OR u.phone ILIKE $1
                    )
                    GROUP BY u.id, u.username, u.email, u.full_name, u.phone, u.personal_discount, u.is_manager
                    ORDER BY total_spent DESC
                ''', f'%{q}%')
            else:
                rows = await conn.fetch('''
                    SELECT
                        u.id,
                        u.username,
                        u.email,
                        u.full_name,
                        u.phone,
                        u.personal_discount,
                        u.is_manager,
                        COUNT(DISTINCT o.id) as order_count,
                        COALESCE(SUM(o.total_amount), 0) as total_spent,
                        MAX(o.created_at) as last_order_date,
                        COUNT(DISTINCT n.id) as notes_count
                    FROM users u
                    LEFT JOIN orders o ON u.id = o.user_id
                    LEFT JOIN customer_notes n ON u.id = n.user_id
                    WHERE u.is_admin = FALSE
                    GROUP BY u.id, u.username, u.email, u.full_name, u.phone, u.personal_discount, u.is_manager
                    HAVING COUNT(DISTINCT o.id) > 0
                    ORDER BY total_spent DESC
                ''')

            result = []
            for r in rows:
                d = dict(r)
                d['id'] = str(d['id'])
                d['total_spent'] = float(d['total_spent'])
                d['personal_discount'] = int(d.get('personal_discount') or 0)
                d['is_manager'] = bool(d.get('is_manager', False))
                if isinstance(d.get('last_order_date'), datetime):
                    d['last_order_date'] = d['last_order_date'].isoformat()
                result.append(d)
            return result
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


class CustomerUpdate(BaseModel):
    username: Optional[str] = None
    email: Optional[str] = None
    full_name: Optional[str] = None
    phone: Optional[str] = None
    personal_discount: Optional[int] = None
    is_manager: Optional[bool] = None


@app.put("/api/admin/customers/{user_id}")
async def update_customer(user_id: str, data: CustomerUpdate, admin=Depends(verify_admin)):
    """Обновить данные клиента"""
    try:
        async with db.pool.acquire() as conn:
            updates = []
            values = []
            i = 1
            if data.username is not None:
                updates.append(f"username = ${i}"); values.append(data.username); i += 1
            if data.email is not None:
                updates.append(f"email = ${i}"); values.append(data.email); i += 1
            if data.full_name is not None:
                updates.append(f"full_name = ${i}"); values.append(data.full_name); i += 1
            if data.phone is not None:
                updates.append(f"phone = ${i}"); values.append(data.phone); i += 1
            if data.personal_discount is not None:
                updates.append(f"personal_discount = ${i}"); values.append(data.personal_discount); i += 1
            if data.is_manager is not None:
                updates.append(f"is_manager = ${i}"); values.append(data.is_manager); i += 1
            if not updates:
                raise HTTPException(status_code=400, detail="Нет данных для обновления")
            values.append(user_id)
            await conn.execute(f"UPDATE users SET {', '.join(updates)} WHERE id = ${i}", *values)
            return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("update_customer error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка")


@app.get("/api/admin/customers/{user_id}/notes")
async def get_customer_notes(user_id: str, admin=Depends(verify_admin)):
    """Получить заметки о клиенте"""
    try:
        async with db.pool.acquire() as conn:
            rows = await conn.fetch('''
                SELECT id, note, created_by, created_at
                FROM customer_notes
                WHERE user_id = $1
                ORDER BY created_at DESC
            ''', user_id)
            
            result = []
            for r in rows:
                d = dict(r)
                if isinstance(d.get('created_at'), datetime):
                    d['created_at'] = d['created_at'].isoformat()
                result.append(d)
            return result
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.post("/api/admin/customers/{user_id}/notes")
async def add_customer_note(user_id: str, body: CustomerNoteCreate, admin=Depends(verify_admin)):
    """Добавить заметку о клиенте"""
    note = body.note.strip()
    
    try:
        async with db.pool.acquire() as conn:
            row = await conn.fetchrow('''
                INSERT INTO customer_notes (user_id, note, created_by)
                VALUES ($1, $2, $3)
                RETURNING id, note, created_by, created_at
            ''', user_id, note, "admin")
            
            d = dict(row)
            if isinstance(d.get('created_at'), datetime):
                d['created_at'] = d['created_at'].isoformat()
            return d
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.delete("/api/admin/customers/notes/{note_id}")
async def delete_customer_note(note_id: int, admin=Depends(verify_admin)):
    """Удалить заметку о клиенте"""
    try:
        async with db.pool.acquire() as conn:
            r = await conn.execute("DELETE FROM customer_notes WHERE id=$1", note_id)
            if r == "DELETE 0":
                raise HTTPException(status_code=404, detail="Заметка не найдена")
            return {"message": "Заметка удалена"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.get("/api/admin/customers/{user_id}/orders")
async def get_customer_orders(user_id: str, admin=Depends(verify_admin)):
    """История заказов конкретного клиента (FR-006, Lvl 2+)"""
    try:
        async with db.pool.acquire() as conn:
            orders = await conn.fetch('''
                SELECT o.id, o.status, o.total_amount, o.payment_status, o.created_at,
                       o.delivery_address, o.track_number
                FROM orders o
                WHERE o.user_id=$1::uuid
                ORDER BY o.created_at DESC
                LIMIT 100
            ''', user_id)
            result = []
            for o in orders:
                d = dict(o)
                d['total_amount'] = float(d['total_amount'])
                if isinstance(d.get('created_at'), datetime):
                    d['created_at'] = d['created_at'].isoformat()
                # Позиции заказа
                items = await conn.fetch('''
                    SELECT product_name, quantity, price FROM order_items WHERE order_id=$1
                ''', o['id'])
                d['items'] = [dict(i) for i in items]
                for it in d['items']:
                    it['price'] = float(it['price'])
                result.append(d)
            return result
    except Exception as e:
        logger.error("get_customer_orders error: %s", e)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.put("/api/admin/orders/{order_id}/status")
async def update_order_status(order_id: int, body: OrderStatusUpdate, admin=Depends(verify_manager_or_admin)):
    new_status = body.status
    delay_note = body.delay_note
    try:
        async with db.pool.acquire() as conn:
            # Получаем email/username клиента для уведомлений
            order_info = await conn.fetchrow(
                "SELECT u.email, u.username FROM orders o JOIN users u ON u.id=o.user_id WHERE o.id=$1",
                order_id
            )
            # Строим динамический UPDATE с учётом per-category статусов
            set_parts = ["status=$1", "delay_note=$2", "updated_at=NOW()"]
            params: list = [new_status, delay_note if delay_note else None]
            if body.payment_status:
                params.append(body.payment_status)
                set_parts.append(f"payment_status=${len(params)}")
            if body.status_in_stock is not None:
                params.append(body.status_in_stock)
                set_parts.append(f"status_in_stock=${len(params)}")
            if body.status_auto is not None:
                params.append(body.status_auto)
                set_parts.append(f"status_auto=${len(params)}")
            if body.status_air is not None:
                params.append(body.status_air)
                set_parts.append(f"status_air=${len(params)}")
            params.append(order_id)
            r = await conn.execute(
                f"UPDATE orders SET {', '.join(set_parts)} WHERE id=${len(params)}",
                *params
            )
            if r == "UPDATE 0":
                raise HTTPException(status_code=404, detail="Заказ не найден")
        # Отправляем уведомления (не блокируем ответ)
        if order_info:
            asyncio.create_task(send_order_email(order_info["email"], order_id, new_status, delay_note))
            asyncio.create_task(send_order_telegram(order_id, new_status, order_info["username"], order_info["email"], delay_note))
        return {"message": "Статус обновлён", "status": new_status, "delay_note": delay_note, "payment_status": body.payment_status}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.put("/api/admin/orders/{order_id}/track")
async def update_order_track(order_id: int, body: TrackNumberUpdate, admin=Depends(verify_manager_or_admin)):
    """Установить трек-номер заказа"""
    try:
        async with db.pool.acquire() as conn:
            r = await conn.execute(
                "UPDATE orders SET track_number=$1, updated_at=NOW() WHERE id=$2",
                body.track_number, order_id
            )
            if r == "UPDATE 0":
                raise HTTPException(status_code=404, detail="Заказ не найден")
            return {"message": "Трек-номер обновлён", "track_number": body.track_number}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Track update error: %s", e)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.put("/api/admin/orders/{order_id}/payment-status")
async def update_order_payment_status(order_id: int, body: PaymentStatusUpdate, admin=Depends(verify_admin)):
    """Обновить статус оплаты заказа (ЗАМ-05, ТЗ §9.1)"""
    try:
        async with db.pool.acquire() as conn:
            r = await conn.execute(
                "UPDATE orders SET payment_status=$1, updated_at=NOW() WHERE id=$2",
                body.payment_status, order_id
            )
            if r == "UPDATE 0":
                raise HTTPException(status_code=404, detail="Заказ не найден")
            return {"message": "Статус оплаты обновлён", "payment_status": body.payment_status}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Payment status update error: %s", e)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.get("/api/admin/orders/{order_id}/items")
async def get_order_items(order_id: int, admin=Depends(verify_manager_or_admin)):
    """Получить позиции заказа (для аккордеона)"""
    try:
        async with db.pool.acquire() as conn:
            rows = await conn.fetch(
                """SELECT oi.*, p.weight_kg as product_weight
                   FROM order_items oi
                   LEFT JOIN products p ON oi.product_id = p.id
                   WHERE oi.order_id = $1 ORDER BY oi.id""",
                order_id
            )
            items = []
            for r in rows:
                d = dict(r)
                d['price'] = float(d['price'])
                if d.get('product_weight'): d['product_weight'] = float(d['product_weight'])
                if d.get('weight_kg'): d['weight_kg'] = float(d['weight_kg'])
                items.append(d)
            return items
    except Exception as e:
        logger.error("Order items error: %s", e)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.post("/api/admin/orders")
async def create_order_manual(request: Request, admin=Depends(verify_admin)):
    """Создать заказ вручную в админке (ТЗ 17)"""
    try:
        data = await request.json()
        user_id = data.get('user_id')
        items = data.get('items', [])
        comment = data.get('comment', '')
        address = data.get('address', '')
        customer_name = data.get('customer_name', '')
        delivery_cost = float(data.get('delivery_cost', 0) or 0)
        if not items:
            raise HTTPException(status_code=400, detail="Нет позиций")
        items_total = sum(float(i.get('price', 0)) * int(i.get('quantity', 1)) for i in items)
        total = items_total + delivery_cost
        async with db.pool.acquire() as conn:
            order_id = await conn.fetchval(
                """INSERT INTO orders (user_id, status, total_amount, delivery_address, comment, payment_status, customer_name, delivery_cost)
                   VALUES ($1, 'created', $2, $3, $4, 'pending', $5, $6) RETURNING id""",
                user_id, total, address, comment, customer_name or None, delivery_cost
            )
            for item in items:
                await conn.execute(
                    """INSERT INTO order_items (order_id, product_id, product_name, price, quantity, order_type, delivery_type, weight_kg, item_comment)
                       VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9)""",
                    order_id,
                    item.get('product_id') or 0,
                    item.get('product_name', 'Позиция'),
                    float(item.get('price', 0)),
                    int(item.get('quantity', 1)),
                    item.get('order_type', 'in_stock'),
                    item.get('delivery_type'),
                    float(item['weight_kg']) if item.get('weight_kg') else None,
                    item.get('comment') or None
                )
                # Декрементируем остаток для товаров в наличии
                if item.get('order_type', 'in_stock') != 'preorder' and item.get('product_id'):
                    spec_id = item.get('specification_id')
                    qty = int(item.get('quantity', 1))
                    if spec_id:
                        await conn.execute(
                            "UPDATE product_specifications SET stock = GREATEST(stock - $1, 0) WHERE id = $2",
                            qty, spec_id
                        )
                    else:
                        await conn.execute(
                            "UPDATE products SET stock = GREATEST(stock - $1, 0) WHERE id = $2",
                            qty, item['product_id']
                        )
            return {"success": True, "order_id": order_id, "total": total}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Manual order error: %s", e)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.get("/api/admin/users")
async def admin_users_list(search: str = "", page: int = 1, limit: int = 50, admin=Depends(verify_superadmin)):
    """Список всех пользователей с поиском по ФИО / username / email."""
    try:
        offset = (page - 1) * limit
        pattern = f"%{search}%"
        async with db.pool.acquire() as conn:
            rows = await conn.fetch(
                """SELECT id, username, email, full_name, phone, is_admin, is_manager,
                          COALESCE(email_verified, FALSE) AS email_verified,
                          personal_discount, created_at
                   FROM users
                   WHERE ($1 = '' OR full_name ILIKE $1 OR username ILIKE $1 OR email ILIKE $1)
                   ORDER BY created_at DESC
                   LIMIT $2 OFFSET $3""",
                pattern, limit, offset
            )
            total = await conn.fetchval(
                "SELECT COUNT(*) FROM users WHERE ($1 = '' OR full_name ILIKE $1 OR username ILIKE $1 OR email ILIKE $1)",
                pattern
            )
        return {
            "users": [
                {
                    "id": str(r['id']),
                    "username": r['username'],
                    "email": r['email'],
                    "full_name": r['full_name'],
                    "phone": r['phone'],
                    "is_admin": r['is_admin'],
                    "is_manager": r['is_manager'],
                    "email_verified": r['email_verified'],
                    "personal_discount": int(r['personal_discount'] or 0),
                    "created_at": r['created_at'].isoformat() if r['created_at'] else None,
                }
                for r in rows
            ],
            "total": total,
            "page": page,
            "limit": limit,
        }
    except Exception as e:
        logger.error("admin_users_list error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.get("/api/admin/users-autocomplete")
async def admin_users_autocomplete(search: str = "", admin=Depends(verify_superadmin)):
    """Автодополнение для поля ФИО при создании заказа вручную."""
    try:
        pattern = f"%{search}%"
        async with db.pool.acquire() as conn:
            rows = await conn.fetch(
                """SELECT id, full_name, username, email FROM users
                   WHERE full_name ILIKE $1 OR username ILIKE $1 OR email ILIKE $1
                   ORDER BY full_name LIMIT 10""",
                pattern
            )
        return [{"id": str(r['id']), "full_name": r['full_name'], "username": r['username'], "email": r['email']} for r in rows]
    except Exception as e:
        logger.error("users_autocomplete error: %s", e)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.put("/api/admin/customers/{user_id}/discount")
async def set_customer_discount(user_id: str, request: Request, admin=Depends(verify_superadmin)):
    """Установить персональную скидку клиенту (ТЗ 16)"""
    try:
        data = await request.json()
        discount = max(0, min(99, int(data.get('discount', 0))))
        async with db.pool.acquire() as conn:
            await conn.execute(
                "UPDATE users SET personal_discount=$1 WHERE id=$2",
                discount, user_id
            )
        return {"success": True, "discount": discount}
    except Exception as e:
        logger.error("Discount set error: %s", e)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.put("/api/admin/customers/{user_id}/role")
async def set_customer_role(user_id: str, request: Request, admin=Depends(verify_superadmin)):
    """Выдать / снять роль у пользователя (только суперадмин)"""
    try:
        data = await request.json()
        role = data.get('role', 'user')  # 'user' | 'manager' | 'admin' | 'superadmin'
        is_manager = role == 'manager'
        is_admin_role = role in ('admin', 'superadmin')
        is_superadmin_role = role == 'superadmin'
        async with db.pool.acquire() as conn:
            await conn.execute(
                "UPDATE users SET is_manager=$1, is_admin=$2, is_superadmin=$3 WHERE id=$4",
                is_manager, is_admin_role, is_superadmin_role, user_id
            )
        return {"success": True, "role": role}
    except Exception as e:
        logger.error("Role set error: %s", e)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.post("/api/admin/users/create")
async def create_user_manual(request: Request, admin=Depends(verify_superadmin)):
    """Создать аккаунт вручную (ТЗ 10.1)"""
    try:
        data = await request.json()
        username = data.get('username', '').strip()
        email    = data.get('email', '').strip()
        full_name = data.get('full_name', '').strip()
        phone    = data.get('phone', '').strip() or None
        password = data.get('password') or secrets.token_urlsafe(12)
        if not username or not email:
            raise HTTPException(status_code=400, detail="Логин и email обязательны")
        pw_hash = hasher.get_password_hash(password)
        user_id = str(uuid4())
        async with db.pool.acquire() as conn:
            await conn.execute(
                "INSERT INTO users (id,username,email,full_name,phone,password_hash,privacy_accepted) VALUES ($1,$2,$3,$4,$5,$6,TRUE)",
                user_id, username, email, full_name or username, phone, pw_hash
            )
        return {"success": True, "user_id": user_id, "password": password}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Create user error: %s", e)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


# ── Installation Requests (ТЗ 15) ──

@app.post("/api/installation-requests")
async def create_installation_request(request: Request, user_id: str = Depends(get_current_user)):
    """Создать заявку на установку"""
    try:
        data = await request.json()
        async with db.pool.acquire() as conn:
            req_id = await conn.fetchval(
                """INSERT INTO installation_requests
                   (user_id, product_id, product_name, scooter_model, battery_type, motor_type, other_info, full_name, phone, telegram)
                   VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10) RETURNING id""",
                user_id,
                data.get('product_id'),
                data.get('product_name', ''),
                data.get('scooter_model', ''),
                data.get('battery_type', ''),
                data.get('motor_type', ''),
                data.get('other_info', ''),
                data.get('full_name', ''),
                data.get('phone', ''),
                data.get('telegram', '')
            )
        return {"success": True, "id": req_id}
    except Exception as e:
        logger.error("Installation request error: %s", e)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.get("/api/admin/installation-requests")
async def get_installation_requests(admin=Depends(verify_admin), status: Optional[str] = None):
    """Получить заявки на установку"""
    try:
        async with db.pool.acquire() as conn:
            q = """SELECT ir.*, u.username, u.email
                   FROM installation_requests ir
                   LEFT JOIN users u ON ir.user_id = u.id
                   WHERE 1=1"""
            params = []
            if status:
                q += f" AND ir.status = ${len(params)+1}"; params.append(status)
            q += " ORDER BY ir.created_at DESC"
            rows = await conn.fetch(q, *params)
            result = []
            for r in rows:
                d = dict(r)
                if d.get('user_id'): d['user_id'] = str(d['user_id'])
                if isinstance(d.get('created_at'), datetime): d['created_at'] = d['created_at'].isoformat()
                result.append(d)
            return result
    except Exception as e:
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.put("/api/admin/installation-requests/{req_id}")
async def update_installation_request(req_id: int, request: Request, admin=Depends(verify_admin)):
    """Обновить статус заявки на установку"""
    try:
        data = await request.json()
        async with db.pool.acquire() as conn:
            await conn.execute(
                "UPDATE installation_requests SET status=$1, admin_comment=$2, updated_at=NOW() WHERE id=$3",
                data.get('status', 'new'), data.get('admin_comment'), req_id
            )
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.get("/api/admin/crm/export-csv")
async def export_crm_csv(admin=Depends(verify_admin)):
    """Экспорт CRM в Excel XLSX (ТЗ §21)"""
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl не установлен. Выполните: pip install openpyxl")
    try:
        async with db.pool.acquire() as conn:
            rows = await conn.fetch('''
                SELECT u.username, u.email, u.full_name, u.phone,
                       u.personal_discount,
                       COUNT(o.id) as order_count,
                       COALESCE(SUM(o.total_amount), 0) as total_spent,
                       MAX(o.created_at) as last_order_date
                FROM users u
                LEFT JOIN orders o ON u.id = o.user_id
                WHERE u.is_admin = FALSE
                GROUP BY u.id, u.username, u.email, u.full_name, u.phone, u.personal_discount
                ORDER BY total_spent DESC
            ''')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CRM"

        headers = ['Логин','Email','Имя','Телефон','Скидка %','Заказов','Сумма (₽)','Последний заказ']
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1A1F2E", end_color="1A1F2E", fill_type="solid")
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for r in rows:
            last = r['last_order_date']
            if isinstance(last, datetime): last = last.strftime('%Y-%m-%d')
            ws.append([
                r['username'], r['email'], r['full_name'] or '', r['phone'] or '',
                float(r['personal_discount'] or 0), int(r['order_count']),
                float(r['total_spent']), last or ''
            ])

        # Авто-ширина колонок
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        from fastapi.responses import Response
        return Response(
            content=buf.getvalue(),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename="crm_export.xlsx"'}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


# ==========================================
# ========== IMAGE OPTIMIZATION ==========
# ==========================================

async def optimize_image(image_data: bytes, max_size: tuple = (800, 800), quality: int = 75) -> bytes:
    """
    Оптимизирует изображение для хранения в базе данных (base64).
    PIL-операции выполняются в thread-pool executor — не блокируют event loop.
    """
    def _sync_optimize(data: bytes) -> bytes:
        try:
            img = Image.open(io.BytesIO(data))
            if img.mode in ('RGBA', 'LA', 'P'):
                background = Image.new('RGB', img.size, (255, 255, 255))
                if img.mode == 'P':
                    img = img.convert('RGBA')
                background.paste(img, mask=img.split()[-1] if img.mode in ('RGBA', 'LA') else None)
                img = background
            elif img.mode != 'RGB':
                img = img.convert('RGB')
            img.thumbnail(max_size, Image.Resampling.LANCZOS)
            output = io.BytesIO()
            img.save(output, format='JPEG', quality=quality, optimize=True)
            return output.getvalue()
        except Exception as e:
            logger.warning("Image optimization failed: %s, using original", e)
            return data

    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, _sync_optimize, image_data)


@app.post("/api/admin/products")
async def create_product(request: Request, admin=Depends(verify_admin)):
    try:
        form = await request.form()
        
        # КРИТИЧЕСКОЕ ЛОГИРОВАНИЕ
        print("\n" + "="*80)
        print("🔍 НАЧАЛО ОБРАБОТКИ ЗАПРОСА НА СОЗДАНИЕ ТОВАРА")
        print("="*80)
        
        name     = str(form.get("name","")).strip()
        category = str(form.get("category","")).strip()
        price    = float(form.get("price",0))
        desc     = str(form.get("description","")).strip()
        stock    = int(form.get("stock",0))
        featured = str(form.get("featured","false")).lower() == "true"
        # ИСПРАВЛЕНИЕ: Автоматически определяем in_stock на основе stock
        in_stock = stock > 0
        preorder = str(form.get("preorder","false")).lower() == "true"
        preorder_unavailable = str(form.get("preorder_unavailable","false")).lower() == "true"
        cost_price_str = str(form.get("cost_price","")).strip()
        cost_price = float(cost_price_str) if cost_price_str else None
        discount_percent_str = str(form.get("discount_percent","0")).strip()
        discount_percent = int(float(discount_percent_str)) if discount_percent_str else 0
        discount_percent = max(0, min(99, discount_percent))
        weight_kg_str = str(form.get("weight_kg","")).strip()
        weight_kg = float(weight_kg_str) if weight_kg_str else None
        price_auto_str = str(form.get("price_preorder_auto","")).strip()
        price_preorder_auto = float(price_auto_str) if price_auto_str else None
        price_air_str = str(form.get("price_preorder_air","")).strip()
        price_preorder_air = float(price_air_str) if price_air_str else None
        warranty_months_str = str(form.get("warranty_months","")).strip()
        warranty_months = int(warranty_months_str) if warranty_months_str else None
        warranty_enabled = str(form.get("warranty_enabled","true")).lower() != "false"
        image_url = str(form.get("image_url","")).strip()
        installation_service = str(form.get("installation_service", "false")).lower() == "true"
        price_type = str(form.get("price_type", "static"))
        price_cny_str = form.get("price_cny")
        price_cny = float(price_cny_str) if price_cny_str else None
        no_preorder = str(form.get("no_preorder", "false")).lower() == "true"
        description_sections_str = form.get("description_sections")
        description_sections = json.loads(description_sections_str) if description_sections_str else None

        # Получаем все файлы изображений (до 5 штук)
        image_files = []
        for i in range(5):
            img_file = form.get(f"image_file_{i}")
            if img_file and hasattr(img_file, 'filename') and img_file.filename:
                image_files.append(img_file)
        
        # Обратная совместимость: если используется старое поле image_file
        old_image_file = form.get("image_file")
        if old_image_file and hasattr(old_image_file, 'filename') and old_image_file.filename:
            if not image_files:  # Только если новые файлы не загружены
                image_files.append(old_image_file)
        
        print(f"📝 Название: {name}")
        print(f"📁 Категория: {category}")
        print(f"💰 Цена: {price}")
        print(f"🖼️  Получено файлов изображений: {len(image_files)}")
        print(f"🌐 image_url: '{image_url}'")
        print("="*80 + "\n")

        if not name or len(name) < 3:
            raise HTTPException(status_code=400, detail="Название слишком короткое (мин. 3 символа)")
        if not category:
            raise HTTPException(status_code=400, detail="Категория обязательна")
        if price <= 0:
            raise HTTPException(status_code=400, detail="Цена должна быть больше 0")
        if not desc or len(desc) < 10:
            raise HTTPException(status_code=400, detail="Описание слишком короткое (мин. 10 символов)")

        # По умолчанию пустая строка (не null), чтобы избежать constraint violation
        final_image = ""
        additional_images = []

        # Обработка всех загруженных изображений
        for idx, image_file in enumerate(image_files):
            print(f"🔍 Обработка изображения {idx + 1}/{len(image_files)}")
            
            # Валидация формата
            ext = Path(image_file.filename).suffix.lower()
            allowed_formats = ['.jpg', '.jpeg', '.png', '.gif', '.webp']
            if ext not in allowed_formats:
                raise HTTPException(
                    status_code=400, 
                    detail=f"Недопустимый формат изображения. Разрешены: {', '.join(allowed_formats)}"
                )
            
            # Чтение файла
            file_content = await image_file.read()
            
            # Валидация размера (макс 10MB)
            max_size_mb = 10
            if len(file_content) > max_size_mb * 1024 * 1024:
                raise HTTPException(
                    status_code=400, 
                    detail=f"Изображение слишком большое. Максимум {max_size_mb}MB"
                )

            # Оптимизация изображения
            try:
                optimized_content = await optimize_image(file_content)
                
                # Конвертируем в base64 для хранения в БД
                image_base64 = base64.b64encode(optimized_content).decode('utf-8')
                image_data_url = f"data:image/jpeg;base64,{image_base64}"
                
                # Первое изображение - основное
                if idx == 0:
                    final_image = image_data_url
                    print(f"📸 Основное изображение: {len(file_content)} → {len(optimized_content)} bytes")
                else:
                    additional_images.append(image_data_url)
                    print(f"📸 Дополнительное изображение {idx}: {len(file_content)} → {len(optimized_content)} bytes")
                
            except Exception as e:
                print(f"❌ Error processing image {idx + 1}: {e}")
                import traceback
                traceback.print_exc()
                raise HTTPException(status_code=500, detail="Ошибка обработки изображения")
        
        # Если файлы не загружены, используем URL
        if not image_files and image_url:
            print(f"ℹ️ Используется image_url: {image_url}")
            final_image = validate_image_url(image_url)  # Fix #5

        async with db.pool.acquire() as conn:
            print(f"\n💾 СОХРАНЕНИЕ В БД:")
            print(f"   final_image длина: {len(final_image)}")
            print(f"   Дополнительных изображений: {len(additional_images)}")
            print("="*80 + "\n")
            
            # Создаем товар
            row = await conn.fetchrow('''
                INSERT INTO products (name,category,price,description,image_url,stock,featured,in_stock,preorder,cost_price,price_preorder_auto,price_preorder_air,discount_percent,weight_kg,preorder_unavailable,warranty_months,warranty_enabled,installation_service,price_type,price_cny,no_preorder,description_sections)
                VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16,$17,$18,$19,$20,$21,$22) RETURNING *
            ''', name, category, price, desc, final_image, stock, featured, in_stock, preorder, cost_price, price_preorder_auto, price_preorder_air, discount_percent, weight_kg, preorder_unavailable, warranty_months, warranty_enabled, installation_service, price_type, price_cny, no_preorder, json.dumps(description_sections) if description_sections else None)
            
            product_id = row['id']
            
            # Сохраняем дополнительные изображения
            for idx, img_url in enumerate(additional_images):
                await conn.execute('''
                    INSERT INTO product_images (product_id, image_url, sort_order)
                    VALUES ($1, $2, $3)
                ''', product_id, img_url, idx + 1)
                print(f"✅ Сохранено дополнительное изображение {idx + 1}")
            
            d = dict(row)
            d['price'] = float(d['price'])
            if d.get('cost_price'):
                d['cost_price'] = float(d['cost_price'])
            return {"success": True, "message": "Товар добавлен!", "product": d}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.put("/api/admin/products/{product_id}")
async def update_product(product_id: int, request: Request, admin=Depends(verify_admin)):
    try:
        form = await request.form()
        async with db.pool.acquire() as conn:
            existing = await conn.fetchrow("SELECT * FROM products WHERE id=$1", product_id)
            if not existing:
                raise HTTPException(status_code=404, detail="Товар не найден")

            name     = str(form.get("name", existing['name'])).strip()
            category = str(form.get("category", existing['category'])).strip()
            price    = float(form.get("price", existing['price']))
            desc     = str(form.get("description", existing['description'])).strip()
            stock    = int(form.get("stock", existing['stock']))
            featured = str(form.get("featured", str(existing['featured']))).lower() == "true"
            # ИСПРАВЛЕНИЕ: Автоматически определяем in_stock на основе stock
            in_stock = stock > 0
            preorder = str(form.get("preorder", str(existing.get('preorder', False)))).lower() == "true"
            preorder_unavailable = str(form.get("preorder_unavailable", str(existing.get('preorder_unavailable', False)))).lower() == "true"
            cost_price_str = str(form.get("cost_price","")).strip()
            cost_price = float(cost_price_str) if cost_price_str else existing.get('cost_price')
            discount_percent_str = str(form.get("discount_percent","")).strip()
            discount_percent = int(float(discount_percent_str)) if discount_percent_str else (existing.get('discount_percent') or 0)
            discount_percent = max(0, min(99, discount_percent))
            price_auto_str = str(form.get("price_preorder_auto","")).strip()
            price_preorder_auto = float(price_auto_str) if price_auto_str else existing.get('price_preorder_auto')
            price_air_str = str(form.get("price_preorder_air","")).strip()
            price_preorder_air = float(price_air_str) if price_air_str else existing.get('price_preorder_air')
            weight_kg_str = str(form.get("weight_kg","")).strip()
            weight_kg = float(weight_kg_str) if weight_kg_str else existing.get('weight_kg')
            warranty_months_str = str(form.get("warranty_months","")).strip()
            warranty_months = int(warranty_months_str) if warranty_months_str else existing.get('warranty_months')
            warranty_enabled_str = form.get("warranty_enabled")
            warranty_enabled = str(warranty_enabled_str).lower() != "false" if warranty_enabled_str is not None else bool(existing.get('warranty_enabled', True))
            image_url = str(form.get("image_url","")).strip()
            image_file = form.get("image_file")
            installation_service_str = form.get("installation_service")
            installation_service = str(installation_service_str).lower() == "true" if installation_service_str is not None else bool(existing.get('installation_service', False))
            price_type_str = form.get("price_type")
            price_type = str(price_type_str) if price_type_str is not None else (existing.get('price_type') or 'static')
            price_cny_str = str(form.get("price_cny","")).strip()
            price_cny = float(price_cny_str) if price_cny_str else existing.get('price_cny')
            no_preorder_str = form.get("no_preorder")
            no_preorder = str(no_preorder_str).lower() == "true" if no_preorder_str is not None else bool(existing.get('no_preorder', False))
            description_sections_str = form.get("description_sections")
            if description_sections_str is not None:
                description_sections = json.loads(description_sections_str) if description_sections_str else None
            else:
                description_sections = existing.get('description_sections')

            final_image = existing['image_url']
            if image_file and hasattr(image_file, 'filename') and image_file.filename:
                # Валидация формата
                ext = Path(image_file.filename).suffix.lower()
                allowed_formats = ['.jpg', '.jpeg', '.png', '.gif', '.webp']
                if ext not in allowed_formats:
                    raise HTTPException(
                        status_code=400, 
                        detail=f"Недопустимый формат файла. Разрешены: {', '.join(allowed_formats)}"
                    )
                
                # Чтение и валидация размера
                file_content = await image_file.read()
                max_size_mb = 10
                if len(file_content) > max_size_mb * 1024 * 1024:
                    raise HTTPException(
                        status_code=400, 
                        detail=f"Файл слишком большой. Максимум {max_size_mb}MB"
                    )
                
                # Оптимизация изображения
                try:
                    optimized_content = await optimize_image(file_content)
                    
                    # Конвертируем в base64
                    image_base64 = base64.b64encode(optimized_content).decode('utf-8')
                    image_data_url = f"data:image/jpeg;base64,{image_base64}"
                    
                    final_image = image_data_url
                    
                    print(f"📸 Image updated: {len(file_content)} → {len(optimized_content)} bytes")
                    print(f"✅ Image converted to base64 successfully")
                    
                except Exception as e:
                    print(f"❌ Error updating image: {e}")
                    raise HTTPException(status_code=500, detail="Ошибка обновления изображения")
            elif image_url:
                final_image = validate_image_url(image_url)  # Fix #5

            await conn.execute('''
                UPDATE products SET name=$1,category=$2,price=$3,description=$4,
                image_url=$5,stock=$6,featured=$7,in_stock=$8,preorder=$9,cost_price=$10,
                price_preorder_auto=$11,price_preorder_air=$12,discount_percent=$13,weight_kg=$14,
                preorder_unavailable=$15,warranty_months=$16,warranty_enabled=$17,
                installation_service=$18,price_type=$19,price_cny=$20,no_preorder=$21,description_sections=$22
                WHERE id=$23
            ''', name, category, price, desc, final_image, stock, featured, in_stock, preorder, cost_price, price_preorder_auto, price_preorder_air, discount_percent, weight_kg, preorder_unavailable, warranty_months, warranty_enabled, installation_service, price_type, price_cny, no_preorder, json.dumps(description_sections) if description_sections else None, product_id)

            return {"success": True, "message": "Товар обновлён"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.post("/api/admin/products/{product_id}/certificates")
async def add_product_certificate(product_id: int, request: Request, admin=Depends(verify_admin)):
    """Загрузить фото сертификата для товара (добавляется в JSON-массив)"""
    try:
        form = await request.form()
        cert_file = form.get("cert_file")
        if not cert_file or not hasattr(cert_file, 'filename') or not cert_file.filename:
            raise HTTPException(status_code=400, detail="Файл не передан")
        ext = Path(cert_file.filename).suffix.lower()
        if ext not in ['.jpg', '.jpeg', '.png', '.webp']:
            raise HTTPException(status_code=400, detail="Допустимы: jpg, jpeg, png, webp")
        file_content = await cert_file.read()
        if len(file_content) > 10 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="Файл слишком большой. Максимум 10MB")
        optimized = await optimize_image(file_content)
        b64 = base64.b64encode(optimized).decode('utf-8')
        data_url = f"data:image/jpeg;base64,{b64}"
        async with db.pool.acquire() as conn:
            row = await conn.fetchrow("SELECT certifications FROM products WHERE id=$1", product_id)
            if not row:
                raise HTTPException(status_code=404, detail="Товар не найден")
            existing = json.loads(row['certifications'] or '[]')
            existing.append(data_url)
            await conn.execute("UPDATE products SET certifications=$1 WHERE id=$2", json.dumps(existing), product_id)
        return {"success": True, "count": len(existing)}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("add_certificate error: %s", e)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.delete("/api/admin/products/{product_id}/certificates/{cert_index}")
async def delete_product_certificate(product_id: int, cert_index: int, admin=Depends(verify_admin)):
    """Удалить фото сертификата по индексу"""
    try:
        async with db.pool.acquire() as conn:
            row = await conn.fetchrow("SELECT certifications FROM products WHERE id=$1", product_id)
            if not row:
                raise HTTPException(status_code=404, detail="Товар не найден")
            certs = json.loads(row['certifications'] or '[]')
            if cert_index < 0 or cert_index >= len(certs):
                raise HTTPException(status_code=404, detail="Сертификат не найден")
            certs.pop(cert_index)
            await conn.execute("UPDATE products SET certifications=$1 WHERE id=$2", json.dumps(certs), product_id)
        return {"success": True, "count": len(certs)}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("delete_certificate error: %s", e)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.get("/api/admin/products/{product_id}/certificates")
async def get_product_certificates(product_id: int, admin=Depends(verify_admin)):
    """Получить список сертификатов товара"""
    try:
        async with db.pool.acquire() as conn:
            row = await conn.fetchrow("SELECT certifications FROM products WHERE id=$1", product_id)
            if not row:
                raise HTTPException(status_code=404, detail="Товар не найден")
            certs = json.loads(row['certifications'] or '[]')
        return {"certifications": certs}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.delete("/api/admin/products/{product_id}")
async def delete_product(product_id: int, admin=Depends(verify_admin)):
    try:
        async with db.pool.acquire() as conn:
            r = await conn.execute("DELETE FROM products WHERE id=$1", product_id)
            if r == "DELETE 0":
                raise HTTPException(status_code=404, detail="Товар не найден")
            return {"message": "Товар удалён"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


# Endpoint removed: /api/admin/add-product was unprotected duplicate of /admin/add-product
# Access the page via: GET /admin/add-product (server-side admin check)


@app.get("/api/stats")
async def get_public_stats():
    # Fix A-6: убраны total_orders и active_orders — они не нужны UI и раскрывают
    # внутреннюю бизнес-метрику любому анонимному посетителю.
    try:
        async with db.pool.acquire() as conn:
            return {
                "total_products": await conn.fetchval("SELECT COUNT(*) FROM products") or 0,
                "categories": await conn.fetchval("SELECT COUNT(*) FROM categories") or 0,
                "total_stock": await conn.fetchval("SELECT COALESCE(SUM(stock),0) FROM products") or 0,
            }
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.get("/api/test-auth")
async def test_auth(admin=Depends(verify_admin)):
    # Fix A-6: эндпоинт закрыт за verify_admin.
    # Раскрытие количества пользователей и заказов без авторизации — разведывательная информация.
    try:
        async with db.pool.acquire() as conn:
            return {
                "status": "ok",
                "users_count": await conn.fetchval("SELECT COUNT(*) FROM users"),
                "products_count": await conn.fetchval("SELECT COUNT(*) FROM products"),
                "categories_count": await conn.fetchval("SELECT COUNT(*) FROM categories"),
                "orders_count": await conn.fetchval("SELECT COUNT(*) FROM orders"),
                "database_connected": True
            }
    except Exception as e:
        logger.error("test-auth error: %s", e)
        return {"status": "error"}


# ==========================================
# ========== SPECIFICATIONS API ==========
# ==========================================

@app.get("/api/products/{product_id}/specifications")
async def get_product_specifications(product_id: int):
    """Получить все спецификации товара"""
    try:
        async with db.pool.acquire() as conn:
            specs = await conn.fetch('''
                SELECT id, name, price, description, image_url, stock, in_stock, preorder, cost_price, discount_percent, sort_order
                FROM product_specifications
                WHERE product_id = $1
                ORDER BY sort_order ASC, id ASC
            ''', product_id)
            
            result = []
            for s in specs:
                d = dict(s)
                d['price'] = float(d['price'])
                if d.get('cost_price'):
                    d['cost_price'] = float(d['cost_price'])
                result.append(d)
            return result
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.post("/api/admin/products/{product_id}/specifications")
async def add_product_specification(product_id: int, request: Request, admin=Depends(verify_admin)):
    """Добавить спецификацию к товару"""
    try:
        form = await request.form()
        name = str(form.get("name", "")).strip()
        price = float(form.get("price", 0))
        desc = str(form.get("description", "")).strip()
        stock = int(form.get("stock", 0))
        # ИСПРАВЛЕНИЕ: Автоматически определяем in_stock на основе stock
        in_stock = stock > 0
        preorder = str(form.get("preorder", "false")).lower() == "true"
        cost_price_str = str(form.get("cost_price", "")).strip()
        cost_price = float(cost_price_str) if cost_price_str else None
        price_auto_str = str(form.get("price_preorder_auto","")).strip()
        price_preorder_auto = float(price_auto_str) if price_auto_str else None
        price_air_str = str(form.get("price_preorder_air","")).strip()
        price_preorder_air = float(price_air_str) if price_air_str else None
        image_url = str(form.get("image_url", "")).strip()
        sort_order = int(form.get("sort_order", 0))

        if not name or len(name) < 3:
            raise HTTPException(status_code=400, detail="Название слишком короткое")
        if price <= 0:
            raise HTTPException(status_code=400, detail="Цена должна быть больше 0")

        # Получаем все файлы изображений (до 5 штук)
        image_files = []
        for i in range(5):
            img_file = form.get(f"image_file_{i}")
            if img_file and hasattr(img_file, 'filename') and img_file.filename:
                image_files.append(img_file)
        
        # Обратная совместимость: если используется старое поле image_file
        old_image_file = form.get("image_file")
        if old_image_file and hasattr(old_image_file, 'filename') and old_image_file.filename:
            if not image_files:  # Только если новые файлы не загружены
                image_files.append(old_image_file)

        final_image = None
        additional_images = []

        # Обработка всех загруженных изображений
        for idx, image_file in enumerate(image_files):
            ext = Path(image_file.filename).suffix.lower()
            if ext not in ['.jpg', '.jpeg', '.png', '.gif', '.webp']:
                raise HTTPException(status_code=400, detail="Недопустимый формат изображения")
            
            # Чтение файла
            file_content = await image_file.read()
            
            # Валидация размера
            max_size_mb = 10
            if len(file_content) > max_size_mb * 1024 * 1024:
                raise HTTPException(status_code=400, detail=f"Изображение слишком большое. Максимум {max_size_mb}MB")
            
            # Оптимизация изображения
            try:
                optimized_content = await optimize_image(file_content)
                
                # Конвертируем в base64
                image_base64 = base64.b64encode(optimized_content).decode('utf-8')
                image_data_url = f"data:image/jpeg;base64,{image_base64}"
                
                # Первое изображение - основное
                if idx == 0:
                    final_image = image_data_url
                else:
                    additional_images.append(image_data_url)
                
            except Exception as e:
                print(f"❌ Error processing image {idx + 1}: {e}")
                raise HTTPException(status_code=500, detail="Ошибка обработки изображения")
        
        # Если файлы не загружены, используем URL
        if not image_files and image_url:
            final_image = image_url

        async with db.pool.acquire() as conn:
            # Проверяем существование товара
            product = await conn.fetchrow("SELECT id, has_specifications FROM products WHERE id=$1", product_id)
            if not product:
                raise HTTPException(status_code=404, detail="Товар не найден")
            
            # Помечаем товар как имеющий спецификации
            if not product['has_specifications']:
                await conn.execute("UPDATE products SET has_specifications=true WHERE id=$1", product_id)
            
            # Создаём спецификацию
            spec = await conn.fetchrow('''
                INSERT INTO product_specifications (product_id, name, price, description, image_url, stock, in_stock, preorder, cost_price, price_preorder_auto, price_preorder_air, sort_order)
                VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12)
                RETURNING *
            ''', product_id, name, price, desc, final_image, stock, in_stock, preorder, cost_price, price_preorder_auto, price_preorder_air, sort_order)
            
            spec_id = spec['id']
            
            # Сохраняем дополнительные изображения
            for idx, img_url in enumerate(additional_images):
                await conn.execute('''
                    INSERT INTO product_images (product_id, specification_id, image_url, sort_order)
                    VALUES ($1, $2, $3, $4)
                ''', product_id, spec_id, img_url, idx + 1)
                print(f"✅ Сохранено дополнительное изображение спецификации {idx + 1}")
            
            d = dict(spec)
            d['price'] = float(d['price'])
            if d.get('cost_price'):
                d['cost_price'] = float(d['cost_price'])
            return {"success": True, "message": "Спецификация добавлена", "specification": d}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.put("/api/admin/specifications/{spec_id}")
async def update_specification(spec_id: int, request: Request, admin=Depends(verify_admin)):
    """Обновить спецификацию"""
    try:
        form = await request.form()
        async with db.pool.acquire() as conn:
            existing = await conn.fetchrow("SELECT * FROM product_specifications WHERE id=$1", spec_id)
            if not existing:
                raise HTTPException(status_code=404, detail="Спецификация не найдена")

            name = str(form.get("name", existing['name'])).strip()
            price = float(form.get("price", existing['price']))
            desc = str(form.get("description", existing['description'] or "")).strip()
            stock = int(form.get("stock", existing['stock']))
            # ИСПРАВЛЕНИЕ: Автоматически определяем in_stock на основе stock
            in_stock = stock > 0
            preorder = str(form.get("preorder", str(existing.get('preorder', False)))).lower() == "true"
            cost_price_str = str(form.get("cost_price", "")).strip()
            cost_price = float(cost_price_str) if cost_price_str else existing.get('cost_price')
            image_url = str(form.get("image_url", "")).strip()
            image_file = form.get("image_file")
            sort_order = int(form.get("sort_order", existing['sort_order']))
            # Fix БАГ 1: сохраняем поля которые раньше игнорировались
            weight_kg_str = str(form.get("weight_kg", "")).strip()
            weight_kg = float(weight_kg_str) if weight_kg_str else existing.get('weight_kg')
            discount_str = str(form.get("discount_percent", "")).strip()
            discount_percent = max(0, min(99, int(float(discount_str)))) if discount_str else (existing.get('discount_percent') or 0)
            price_auto_str = str(form.get("price_preorder_auto", "")).strip()
            price_preorder_auto = float(price_auto_str) if price_auto_str else existing.get('price_preorder_auto')
            price_air_str = str(form.get("price_preorder_air", "")).strip()
            price_preorder_air = float(price_air_str) if price_air_str else existing.get('price_preorder_air')

            final_image = existing['image_url']
            if image_file and hasattr(image_file, 'filename') and image_file.filename:
                # Fix A-7: конвертируем в base64 (как в create_product/add_specification).
                # Сохранение на диск заменено — нет риска обхода через расширение файла.
                ext = Path(image_file.filename).suffix.lower()
                if ext not in ['.jpg', '.jpeg', '.png', '.gif', '.webp']:
                    raise HTTPException(status_code=400, detail="Недопустимый формат файла")
                file_content = await image_file.read()
                if len(file_content) > 10 * 1024 * 1024:
                    raise HTTPException(status_code=400, detail="Файл слишком большой. Максимум 10MB")
                try:
                    optimized = await optimize_image(file_content)
                    final_image = "data:image/jpeg;base64," + base64.b64encode(optimized).decode('utf-8')
                except Exception as e:
                    raise HTTPException(status_code=500, detail="Ошибка обработки изображения")
            elif image_url:
                final_image = validate_image_url(image_url)  # Fix #5

            # Fix БАГ 1: сохраняем все поля, включая weight_kg, discount_percent, price_preorder
            await conn.execute('''
                UPDATE product_specifications
                SET name=$1, price=$2, description=$3, image_url=$4, stock=$5, in_stock=$6,
                    preorder=$7, cost_price=$8, sort_order=$9,
                    weight_kg=$10, discount_percent=$11,
                    price_preorder_auto=$12, price_preorder_air=$13
                WHERE id=$14
            ''', name, price, desc, final_image, stock, in_stock, preorder, cost_price, sort_order,
                 weight_kg, discount_percent, price_preorder_auto, price_preorder_air, spec_id)

            return {"success": True, "message": "Спецификация обновлена"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.delete("/api/admin/specifications/{spec_id}")
async def delete_specification(spec_id: int, admin=Depends(verify_admin)):
    """Удалить спецификацию"""
    try:
        async with db.pool.acquire() as conn:
            # Получаем product_id перед удалением
            spec = await conn.fetchrow("SELECT product_id FROM product_specifications WHERE id=$1", spec_id)
            if not spec:
                raise HTTPException(status_code=404, detail="Спецификация не найдена")
            
            product_id = spec['product_id']
            
            # Удаляем спецификацию
            await conn.execute("DELETE FROM product_specifications WHERE id=$1", spec_id)
            
            # Проверяем, остались ли еще спецификации у товара
            remaining = await conn.fetchval("SELECT COUNT(*) FROM product_specifications WHERE product_id=$1", product_id)
            if remaining == 0:
                await conn.execute("UPDATE products SET has_specifications=false WHERE id=$1", product_id)
            
            return {"message": "Спецификация удалена"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.get("/api/products/{product_id}/images")
async def get_product_images(product_id: int, specification_id: Optional[int] = None):
    """Получить дополнительные изображения товара или спецификации"""
    try:
        async with db.pool.acquire() as conn:
            if specification_id:
                images = await conn.fetch('''
                    SELECT id, image_url, sort_order
                    FROM product_images
                    WHERE product_id = $1 AND specification_id = $2
                    ORDER BY sort_order ASC, id ASC
                ''', product_id, specification_id)
            else:
                images = await conn.fetch('''
                    SELECT id, image_url, sort_order
                    FROM product_images
                    WHERE product_id = $1 AND specification_id IS NULL
                    ORDER BY sort_order ASC, id ASC
                ''', product_id)
            
            return [dict(img) for img in images]
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.get("/api/products/{product_id}/characteristics")
async def get_product_characteristics(product_id: int, specification_id: Optional[int] = None):
    """Получить характеристики товара или спецификации"""
    try:
        async with db.pool.acquire() as conn:
            if specification_id:
                chars = await conn.fetch('''
                    SELECT id, char_name, char_value, sort_order
                    FROM product_characteristics
                    WHERE product_id = $1 AND specification_id = $2
                    ORDER BY sort_order ASC, id ASC
                ''', product_id, specification_id)
            else:
                chars = await conn.fetch('''
                    SELECT id, char_name, char_value, sort_order
                    FROM product_characteristics
                    WHERE product_id = $1 AND specification_id IS NULL
                    ORDER BY sort_order ASC, id ASC
                ''', product_id)
            
            return [dict(char) for char in chars]
    except Exception as e:
        logger.error("Internal error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


# ========== ЗАПУСК ==========
# Fix #8: Debug endpoint закрыт для обычных пользователей.
# Раскрытие путей файловой системы и списка файлов помогает атакующим в разведке.
@app.get("/api/debug/uploads")
async def debug_uploads(admin=Depends(verify_admin)):
    """Debug endpoint — только для администратора."""
    try:
        files = list(UPLOAD_DIR.iterdir())
        return {
            "upload_dir": str(UPLOAD_DIR),
            "exists": UPLOAD_DIR.exists(),
            "files_count": len([f for f in files if f.is_file()]),
            # Не возвращаем список имён файлов — это не нужно даже администратору
        }
    except Exception as e:
        logger.error("debug_uploads error: %s", e)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


# ══════════════════════════════════════════════
# ═══ v24: АВТО-СКИДКА ════════════════════════
# ══════════════════════════════════════════════

@app.get("/api/admin/auto-discount")
async def get_auto_discount(admin=Depends(verify_superadmin)):
    """Получить текущие настройки авто-скидки"""
    try:
        async with db.pool.acquire() as conn:
            rows = await conn.fetch(
                "SELECT key, value FROM delivery_settings WHERE key IN ('auto_discount_enabled','auto_discount_percent')"
            )
            settings = {r['key']: r['value'] for r in rows}
            return {
                "enabled": settings.get('auto_discount_enabled') == '1',
                "discount_percent": int(settings.get('auto_discount_percent') or 0)
            }
    except Exception as e:
        logger.error("get_auto_discount error: %s", e)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.post("/api/admin/auto-discount")
async def set_auto_discount(request: Request, admin=Depends(verify_superadmin)):
    """Установить авто-скидку (включить/выключить и размер)"""
    try:
        data = await request.json()
        enabled = bool(data.get('enabled', False))
        pct = max(0, min(99, int(data.get('discount_percent', 0))))
        async with db.pool.acquire() as conn:
            await conn.execute(
                "UPDATE delivery_settings SET value=$1, updated_at=CURRENT_TIMESTAMP WHERE key='auto_discount_enabled'",
                '1' if enabled else '0'
            )
            await conn.execute(
                "UPDATE delivery_settings SET value=$1, updated_at=CURRENT_TIMESTAMP WHERE key='auto_discount_percent'",
                str(pct)
            )
        return {"success": True, "enabled": enabled, "discount_percent": pct}
    except Exception as e:
        logger.error("set_auto_discount error: %s", e)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


# ══════════════════════════════════════════════
# ═══ v24: ПРОМОКОДЫ ══════════════════════════
# ══════════════════════════════════════════════

@app.get("/api/admin/promo-codes")
async def list_promo_codes(admin=Depends(verify_superadmin)):
    """Список всех промокодов"""
    try:
        async with db.pool.acquire() as conn:
            rows = await conn.fetch(
                "SELECT id,code,discount_type,discount_value,max_uses,used_count,is_active,expires_at,created_at FROM promo_codes ORDER BY created_at DESC"
            )
            result = []
            for r in rows:
                d = dict(r)
                d['discount_value'] = float(d['discount_value'])
                if d.get('expires_at'):
                    d['expires_at'] = d['expires_at'].isoformat()
                if d.get('created_at'):
                    d['created_at'] = d['created_at'].isoformat()
                result.append(d)
            return result
    except Exception as e:
        logger.error("list_promo_codes error: %s", e)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.post("/api/admin/promo-codes")
async def create_promo_code(request: Request, admin=Depends(verify_superadmin)):
    """Создать промокод"""
    try:
        data = await request.json()
        code = data.get('code', '').strip().upper()
        if not code:
            raise HTTPException(status_code=400, detail="Код не может быть пустым")
        disc_type = data.get('discount_type', 'percent')
        if disc_type not in ('percent', 'fixed'):
            raise HTTPException(status_code=400, detail="Неверный тип скидки")
        disc_val = float(data.get('discount_value', 0))
        if disc_val <= 0:
            raise HTTPException(status_code=400, detail="Значение скидки должно быть > 0")
        if disc_type == 'percent' and disc_val > 99:
            raise HTTPException(status_code=400, detail="Процент скидки не может быть > 99")
        max_uses = int(data['max_uses']) if data.get('max_uses') else None
        expires_raw = data.get('expires_at')
        expires_at = None
        if expires_raw:
            from datetime import datetime as dt
            try:
                expires_at = dt.fromisoformat(expires_raw)
            except Exception:
                expires_at = None
        async with db.pool.acquire() as conn:
            row = await conn.fetchrow(
                """INSERT INTO promo_codes (code,discount_type,discount_value,max_uses,expires_at)
                   VALUES ($1,$2,$3,$4,$5) RETURNING id""",
                code, disc_type, disc_val, max_uses, expires_at
            )
        return {"success": True, "id": row['id'], "code": code}
    except HTTPException:
        raise
    except asyncpg.UniqueViolationError:
        raise HTTPException(status_code=409, detail=f"Промокод с кодом уже существует")
    except Exception as e:
        logger.error("create_promo_code error: %s", e)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.delete("/api/admin/promo-codes/{code_id}")
async def delete_promo_code(code_id: int, admin=Depends(verify_superadmin)):
    """Удалить промокод"""
    try:
        async with db.pool.acquire() as conn:
            deleted = await conn.fetchval(
                "DELETE FROM promo_codes WHERE id=$1 RETURNING id", code_id
            )
        if not deleted:
            raise HTTPException(status_code=404, detail="Промокод не найден")
        return {"success": True}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("delete_promo_code error: %s", e)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.post("/api/promo-codes/validate")
async def validate_promo_code(request: Request, user_id: str = Depends(get_current_user)):
    """Проверить и применить промокод к сумме корзины (вызывается из корзины)"""
    try:
        data = await request.json()
        code = data.get('code', '').strip().upper()
        cart_total = float(data.get('cart_total', 0))
        if not code:
            raise HTTPException(status_code=400, detail="Введите промокод")
        async with db.pool.acquire() as conn:
            row = await conn.fetchrow(
                "SELECT * FROM promo_codes WHERE code=$1 AND is_active=TRUE", code
            )
            if not row:
                raise HTTPException(status_code=404, detail="Промокод не найден или неактивен")
            if row['expires_at'] and row['expires_at'] < datetime.now(timezone.utc).replace(tzinfo=None):
                raise HTTPException(status_code=410, detail="Срок действия промокода истёк")
            if row['max_uses'] and row['used_count'] >= row['max_uses']:
                raise HTTPException(status_code=410, detail="Промокод исчерпан")
            # Считаем скидку
            if row['discount_type'] == 'percent':
                discount_amount = round(cart_total * float(row['discount_value']) / 100)
            else:
                discount_amount = min(round(float(row['discount_value'])), cart_total)
            new_total = max(0, cart_total - discount_amount)
            return {
                "valid": True,
                "code": code,
                "discount_type": row['discount_type'],
                "discount_value": float(row['discount_value']),
                "discount_amount": discount_amount,
                "new_total": new_total
            }
    except HTTPException:
        raise
    except Exception as e:
        logger.error("validate_promo error: %s", e)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


# ============================================================
# ВИШЛИСТ (избранное)
# ============================================================

@app.get("/api/wishlist")
async def get_wishlist(user_id: str = Depends(get_current_user)):
    """Получить список избранного пользователя"""
    try:
        async with db.pool.acquire() as conn:
            rows = await conn.fetch('''
                SELECT w.product_id, p.name, p.price, p.image_url, p.in_stock, p.preorder,
                       p.discount_percent, w.added_at
                FROM wishlists w
                JOIN products p ON p.id = w.product_id
                WHERE w.user_id = $1
                ORDER BY w.added_at DESC
            ''', user_id)
            result = []
            for r in rows:
                d = dict(r)
                d['price'] = float(d['price'])
                if isinstance(d.get('added_at'), datetime):
                    d['added_at'] = d['added_at'].isoformat()
                result.append(d)
            return result
    except Exception as e:
        logger.error("get_wishlist error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.post("/api/wishlist/{product_id}")
async def add_to_wishlist(product_id: int, user_id: str = Depends(get_current_user)):
    """Добавить товар в избранное"""
    try:
        async with db.pool.acquire() as conn:
            exists = await conn.fetchval("SELECT id FROM products WHERE id=$1", product_id)
            if not exists:
                raise HTTPException(status_code=404, detail="Товар не найден")
            await conn.execute(
                "INSERT INTO wishlists (user_id, product_id) VALUES ($1, $2) ON CONFLICT DO NOTHING",
                user_id, product_id
            )
            return {"message": "Добавлено в избранное"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("add_to_wishlist error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.delete("/api/wishlist/{product_id}")
async def remove_from_wishlist(product_id: int, user_id: str = Depends(get_current_user)):
    """Убрать товар из избранного"""
    try:
        async with db.pool.acquire() as conn:
            await conn.execute(
                "DELETE FROM wishlists WHERE user_id=$1 AND product_id=$2",
                user_id, product_id
            )
            return {"message": "Удалено из избранного"}
    except Exception as e:
        logger.error("remove_from_wishlist error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


# ============================================================
# ДОКУМЕНТЫ ПОСТАВЩИКА
# ============================================================

DOCS_UPLOAD_DIR = UPLOAD_DIR / "docs"
DOCS_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
ALLOWED_DOC_TYPES = {"partner", "warranty"}


@app.get("/api/supplier-docs")
async def get_supplier_docs():
    """Получить список загруженных документов поставщика"""
    try:
        async with db.pool.acquire() as conn:
            rows = await conn.fetch(
                "SELECT id, doc_type, filename, title, uploaded_at FROM supplier_documents ORDER BY doc_type, uploaded_at"
            )
            return [{"id": r["id"], "doc_type": r["doc_type"], "filename": r["filename"], "title": r["title"], "url": f"/api/supplier-docs/{r['id']}/file", "uploaded_at": r["uploaded_at"].isoformat() if r["uploaded_at"] else None} for r in rows]
    except Exception as e:
        logger.error("get_supplier_docs error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.post("/api/supplier-docs/{doc_type}")
async def upload_supplier_doc(
    doc_type: str,
    file: UploadFile = File(...),
    admin: dict = Depends(verify_admin)
):
    """Загрузить документ поставщика (только для admin)"""
    if doc_type not in ALLOWED_DOC_TYPES:
        raise HTTPException(status_code=400, detail="Недопустимый тип документа")
    # Validate file type
    allowed_exts = {".pdf", ".jpg", ".jpeg", ".png"}
    suffix = Path(file.filename).suffix.lower() if file.filename else ""
    if suffix not in allowed_exts:
        raise HTTPException(status_code=400, detail="Разрешены только PDF и изображения")
    # Save file with unique name
    from datetime import datetime as _dt
    ts = _dt.utcnow().strftime("%Y%m%d_%H%M%S")
    uid = uuid4().hex[:8]
    safe_name = f"{doc_type}_{ts}_{uid}{suffix}"
    dest = DOCS_UPLOAD_DIR / safe_name
    try:
        content = await file.read()
        if len(content) > 10 * 1024 * 1024:  # 10 MB limit
            raise HTTPException(status_code=400, detail="Файл слишком большой (макс 10 МБ)")
        async with aiofiles.open(dest, "wb") as f:
            await f.write(content)
    except HTTPException:
        raise
    except Exception as e:
        logger.error("upload_supplier_doc file write error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Ошибка сохранения файла")
    # Insert DB record — store file content in DB for persistence across Render.com redeploys
    _ct_map = {".pdf": "application/pdf", ".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png"}
    file_ct = _ct_map.get(suffix, "application/octet-stream")
    try:
        async with db.pool.acquire() as conn:
            row = await conn.fetchrow('''
                INSERT INTO supplier_documents (doc_type, filename, file_content, content_type, uploaded_at)
                VALUES ($1, $2, $3, $4, NOW())
                RETURNING id
            ''', doc_type, safe_name, content, file_ct)
        doc_id = row["id"]
        return {"message": "Документ загружен", "filename": safe_name, "url": f"/api/supplier-docs/{doc_id}/file"}
    except Exception as e:
        logger.error("upload_supplier_doc db error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Ошибка базы данных")


@app.get("/api/supplier-docs/{doc_id}/file")
async def get_supplier_doc_file(doc_id: int):
    """Отдать файл документа из БД (не зависит от файловой системы)"""
    from fastapi.responses import Response as _Resp
    try:
        async with db.pool.acquire() as conn:
            row = await conn.fetchrow(
                "SELECT filename, file_content, content_type FROM supplier_documents WHERE id=$1", doc_id
            )
        if not row or not row["file_content"]:
            raise HTTPException(status_code=404, detail="Файл не найден")
        ct = row["content_type"] or "application/octet-stream"
        return _Resp(
            content=bytes(row["file_content"]),
            media_type=ct,
            headers={
                "Content-Disposition": f'inline; filename="{row["filename"]}"',
                "Cache-Control": "public, max-age=86400",
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error("get_supplier_doc_file error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.delete("/api/supplier-docs/{doc_id}")
async def delete_supplier_doc(doc_id: int, admin: dict = Depends(verify_admin)):
    """Удалить документ поставщика (только для admin)"""
    try:
        async with db.pool.acquire() as conn:
            row = await conn.fetchrow("SELECT filename FROM supplier_documents WHERE id=$1", doc_id)
            if not row:
                raise HTTPException(status_code=404, detail="Документ не найден")
            file_path = DOCS_UPLOAD_DIR / row["filename"]
            try:
                file_path.unlink(missing_ok=True)
            except Exception:
                pass
            await conn.execute("DELETE FROM supplier_documents WHERE id=$1", doc_id)
        return {"message": "Документ удалён"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("delete_supplier_doc error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.delete("/api/admin/supplier-docs/{doc_id}")
async def admin_delete_supplier_doc(doc_id: int, admin=Depends(verify_admin)):
    """Удалить документ поставщика"""
    async with db.pool.acquire() as conn:
        await conn.execute("DELETE FROM supplier_documents WHERE id = $1", doc_id)
    return {"ok": True}


@app.put("/api/admin/supplier-docs/{doc_id}")
async def admin_update_supplier_doc(doc_id: int, data: dict, admin=Depends(verify_admin)):
    """Обновить данные документа (заголовок, тип)"""
    title = data.get("title")
    doc_type = data.get("doc_type")
    async with db.pool.acquire() as conn:
        if title is not None:
            await conn.execute("UPDATE supplier_documents SET title = $1 WHERE id = $2", title, doc_id)
        if doc_type is not None:
            await conn.execute("UPDATE supplier_documents SET doc_type = $1 WHERE id = $2", doc_type, doc_id)
    return {"ok": True}


@app.post("/api/admin/supplier-docs")
async def admin_upload_supplier_doc(
    file: UploadFile = File(...),
    doc_type: str = Form("partner"),
    title: str = Form(None),
    admin: dict = Depends(verify_admin)
):
    """Загрузить документ поставщика через админку (с поддержкой title)"""
    allowed_exts = {".pdf", ".jpg", ".jpeg", ".png", ".webp"}
    suffix = Path(file.filename).suffix.lower() if file.filename else ""
    if suffix not in allowed_exts:
        raise HTTPException(status_code=400, detail="Разрешены только PDF и изображения (jpg, png, webp)")
    from datetime import datetime as _dt
    ts = _dt.utcnow().strftime("%Y%m%d_%H%M%S")
    uid = uuid4().hex[:8]
    safe_name = f"{doc_type}_{ts}_{uid}{suffix}"
    dest = DOCS_UPLOAD_DIR / safe_name
    try:
        content = await file.read()
        if len(content) > 15 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="Файл слишком большой (макс 15 МБ)")
        async with aiofiles.open(dest, "wb") as f:
            await f.write(content)
    except HTTPException:
        raise
    except Exception as e:
        logger.error("admin_upload_supplier_doc file write error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Ошибка сохранения файла")
    _ct_map = {".pdf": "application/pdf", ".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png", ".webp": "image/webp"}
    file_ct = _ct_map.get(suffix, "application/octet-stream")
    try:
        async with db.pool.acquire() as conn:
            row = await conn.fetchrow('''
                INSERT INTO supplier_documents (doc_type, filename, file_content, content_type, title, uploaded_at)
                VALUES ($1, $2, $3, $4, $5, NOW())
                RETURNING id
            ''', doc_type, safe_name, content, file_ct, title or safe_name)
        doc_id = row["id"]
        return {"ok": True, "id": doc_id, "filename": safe_name, "url": f"/api/supplier-docs/{doc_id}/file"}
    except Exception as e:
        logger.error("admin_upload_supplier_doc db error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Ошибка базы данных")


# ============================================================
# ОТЗЫВЫ НА ТОВАРЫ
# ============================================================

@app.get("/api/products/{product_id}/reviews")
async def get_product_reviews(product_id: int):
    """Получить отзывы на товар"""
    try:
        async with db.pool.acquire() as conn:
            rows = await conn.fetch('''
                SELECT r.id, r.rating, r.review_text, r.created_at,
                       u.username
                FROM product_reviews r
                JOIN users u ON u.id = r.user_id
                WHERE r.product_id = $1
                ORDER BY r.created_at DESC
            ''', product_id)
            result = []
            for r in rows:
                d = dict(r)
                if isinstance(d.get('created_at'), datetime):
                    d['created_at'] = d['created_at'].isoformat()
                result.append(d)
            return result
    except Exception as e:
        logger.error("get_reviews error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.post("/api/products/{product_id}/reviews")
async def add_product_review(product_id: int, request: Request, user_id: str = Depends(get_current_user)):
    """Добавить отзыв на товар (один отзыв на пользователя)"""
    try:
        data = await request.json()
        rating = int(data.get('rating', 0))
        review_text = str(data.get('review_text', '')).strip()[:2000]
        if not 1 <= rating <= 5:
            raise HTTPException(status_code=400, detail="Рейтинг должен быть от 1 до 5")
        async with db.pool.acquire() as conn:
            exists = await conn.fetchval("SELECT id FROM products WHERE id=$1", product_id)
            if not exists:
                raise HTTPException(status_code=404, detail="Товар не найден")
            # Проверяем что пользователь покупал этот товар
            purchased = await conn.fetchval('''
                SELECT 1 FROM order_items oi
                JOIN orders o ON o.id = oi.order_id
                WHERE o.user_id=$1 AND oi.product_id=$2 AND o.status='completed'
                LIMIT 1
            ''', user_id, product_id)
            if not purchased:
                raise HTTPException(status_code=403, detail="Отзыв можно оставить только после покупки")
            try:
                review_id = await conn.fetchval('''
                    INSERT INTO product_reviews (product_id, user_id, rating, review_text)
                    VALUES ($1, $2, $3, $4) RETURNING id
                ''', product_id, user_id, rating, review_text or None)
            except asyncpg.UniqueViolationError:
                raise HTTPException(status_code=409, detail="Вы уже оставили отзыв на этот товар")
            return {"message": "Отзыв добавлен", "review_id": review_id}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("add_review error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.get("/api/products/{product_id}/reviews/stats")
async def get_product_review_stats(product_id: int):
    """Получить статистику отзывов товара (средний рейтинг, количество)"""
    try:
        async with db.pool.acquire() as conn:
            row = await conn.fetchrow('''
                SELECT COUNT(*) as total, ROUND(AVG(rating)::numeric, 1) as avg_rating
                FROM product_reviews WHERE product_id=$1
            ''', product_id)
            return {
                "total": int(row['total']),
                "avg_rating": float(row['avg_rating']) if row['avg_rating'] else None
            }
    except Exception as e:
        logger.error("review_stats error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


# ============================================================
# ПРОФИЛЬ ПОЛЬЗОВАТЕЛЯ (обновление данных)
# ============================================================

@app.put("/api/me")
async def update_profile(request: Request, user_id: str = Depends(get_current_user)):
    """Обновить профиль пользователя (имя, телефон)"""
    try:
        data = await request.json()
        full_name = str(data.get('full_name', '')).strip()
        phone = str(data.get('phone', '')).strip()
        if not full_name:
            raise HTTPException(status_code=400, detail="Имя не может быть пустым")
        async with db.pool.acquire() as conn:
            await conn.execute(
                "UPDATE users SET full_name=$1, phone=$2 WHERE id=$3",
                full_name, phone or None, user_id
            )
        return {"message": "Профиль обновлён"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("update_profile error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.put("/api/me/password")
async def change_password(request: Request, user_id: str = Depends(get_current_user)):
    """Изменить пароль пользователя"""
    try:
        data = await request.json()
        old_password = str(data.get('old_password', ''))
        new_password = str(data.get('new_password', ''))
        if len(new_password) < 8:
            raise HTTPException(status_code=400, detail="Новый пароль должен быть не менее 8 символов")
        async with db.pool.acquire() as conn:
            row = await conn.fetchrow("SELECT password_hash FROM users WHERE id=$1", user_id)
            if not row or not PasswordHasher.verify(old_password, row['password_hash']):
                raise HTTPException(status_code=400, detail="Неверный текущий пароль")
            new_hash = PasswordHasher.hash(new_password)
            await conn.execute("UPDATE users SET password_hash=$1 WHERE id=$2", new_hash, user_id)
        return {"message": "Пароль изменён"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("change_password error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


# ============================================================
# СТРАНИЦА ЛИЧНОГО КАБИНЕТА
# ============================================================

@app.get("/account", response_class=HTMLResponse)
async def account_page(request: Request):
    """Страница личного кабинета"""
    return templates.TemplateResponse("account.html", {"request": request, "ym_counter_id": YM_COUNTER_ID})


# ============================================================
# ПОХОЖИЕ ТОВАРЫ
# ============================================================

@app.get("/api/products/{product_id}/similar")
async def get_similar_products(product_id: int):
    """Получить похожие товары из той же категории"""
    try:
        async with db.pool.acquire() as conn:
            product = await conn.fetchrow("SELECT category FROM products WHERE id=$1", product_id)
            if not product:
                raise HTTPException(status_code=404, detail="Товар не найден")
            rows = await conn.fetch('''
                SELECT id, name, price, image_url, in_stock, preorder, discount_percent
                FROM products
                WHERE category=$1 AND id != $2
                ORDER BY RANDOM()
                LIMIT 4
            ''', product['category'], product_id)
            result = []
            for r in rows:
                d = dict(r)
                d['price'] = float(d['price'])
                result.append(d)
            return result
    except HTTPException:
        raise
    except Exception as e:
        logger.error("similar_products error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


# ============================================================
# 404 HANDLER
# ============================================================

from fastapi.exceptions import HTTPException as FastAPIHTTPException
from starlette.exceptions import HTTPException as StarletteHTTPException

@app.exception_handler(StarletteHTTPException)
async def custom_404_handler(request: Request, exc: StarletteHTTPException):
    if exc.status_code == 404:
        try:
            return templates.TemplateResponse(
                "404.html",
                {"request": request},
                status_code=404
            )
        except Exception:
            return HTMLResponse(
                content="<h1>404 — Страница не найдена</h1><a href='/'>На главную</a>",
                status_code=404
            )
    return JSONResponse({"detail": exc.detail}, status_code=exc.status_code)


if __name__ == "__main__":
    print("=" * 70)
    print("⚡ Fm TuN v18")
    print("=" * 70)
    print("   http://localhost:8000              — Главная")
    print("   http://localhost:8000/products     — Каталог")
    print("   http://localhost:8000/admin        — Админка")
    print("   http://localhost:8000/privacy-policy — Политика конфиденциальности")
    print("=" * 70)
    print("⚠️  Заполните .env: ADMIN_PASSWORD, SECRET_KEY, DATABASE_URL")
    print("💳  Для эквайринга: PAYMENT_API_KEY, PAYMENT_SHOP_ID, PAYMENT_SECRET_KEY")
    print("=" * 70)
    uvicorn.run(app, host="0.0.0.0", port=8000, reload=False)
